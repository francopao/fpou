"""
ZODA Dashboard v2 — Generador de HTML autónomo
================================================
Uso:  python zoda_dashboard_v2.py
Output: ZODA_Dashboard_<DD.MM.YY>.html  (misma carpeta que ZODA.xlsm)

Lee directamente desde ZODA.xlsm (BD, DQ_IN, Parametría, Financials, Cons y Const, Min y Uti).
No modifica ningún archivo fuente. No toca ZODA_v7.py.
"""

import os, datetime, warnings, json, base64, time
import numpy as np
import pandas as pd
warnings.filterwarnings("ignore")

# ─── CONFIGURACIÓN ──────────────────────────────────────────────────────────
RUTA_ZODA = (
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA"
    r"\Docs privados\ZODA\shortcut\ZODA.xlsm"
)
RUTA_SALIDA = os.path.dirname(RUTA_ZODA)
RUTA_LOGO   = os.path.join(os.path.dirname(RUTA_ZODA), "integra_logo.png")
RUTA_QEST   = (
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA"
    r"\Docs privados\ZODA\shortcut\Benchmarking\QEST_BD.xlsm"
)
# CSV intermedio generado por ZODA_03_06_2026.py (fuente de múltiplos para Plot 05)
RUTA_MULTIPLOS_CSV = os.path.join(os.path.dirname(RUTA_ZODA), "Multiplos_diarios.csv")

# ─── PALETA AFP INTEGRA (exacta del template corporativo) ───────────────────
BRAND = {
    "navy":   "#1E2E6E",   # azul marino principal (header/fondo oscuro)
    "dark":   "#002060",   # azul oscuro (footer bar)
    "teal":   "#00AECB",   # cyan acento corporativo
    "yellow": "#E3E829",   # amarillo accent bar
    "gray1":  "#DCDDDE",   # gris claro
    "gray2":  "#7E8083",   # gris medio
    "white":  "#FFFFFF",
    # Surface system (dark dashboard)
    "bg":     "#07101f",
    "surf1":  "#0b1730",
    "surf2":  "#0f1e3a",
    "border": "#1a2f52",
    "text":   "#e8eef8",
    "muted":  "#6b85a8",
    # Semantic
    "green":  "#2ec97a",
    "red":    "#e84343",
    "amber":  "#e0a320",
}

SECTOR_COLOR = {
    "Financials":             "#2f81f7",
    "Mining":                 "#f0883e",
    "Utilities":              "#3fb950",
    "Consumer":               "#bc8cff",
    "Construction Materials": "#ffa657",
    "Industrials":            "#79c0ff",
    "Health Care":            "#ff7b72",
    "Media":                  "#d2a8ff",
    "Index":                  "#6a85a8",
    "Other":                  "#6a85a8",
}

# ─── CONSTANTS ──────────────────────────────────────────────────────────────
_VACIOS  = ["#N/A N/A","NM","(Invalid Formula Name)","NA","#VALUE!"]
_COLS_ALL = ["IQ_NI","IQ_TOTAL_REV","IQ_TOTAL_EQUITY","IQ_EBITDA",
             "Px_last","CUR_MKT_CAP","CURR_ENTP_VAL"]
MULTIPLOS_LABELS = {
    "Trail_PE":       "P/E Trailing",
    "Fwd_PE":         "P/E Forward",
    "Trail_PS":       "P/S Trailing",
    "Fwd_PS":         "P/S Forward",
    "Trail_PB":       "P/B Trailing",
    "Fwd_PB":         "P/B Forward",
    "Trail_EVEBITDA": "EV/EBITDA Trailing",
    "Fwd_EVEBITDA":   "EV/EBITDA Forward",
}

# ─── LECTURA ────────────────────────────────────────────────────────────────
def _read(sheet, **kw):
    return pd.read_excel(RUTA_ZODA, sheet_name=sheet,
                         engine="openpyxl", dtype=object, **kw)

def cargar_bd():
    df = _read("BD", header=0)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"]).sort_values("Fecha").reset_index(drop=True)
    for c in df.columns[1:]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    z = [c for c in df.columns if c.startswith("Z_") or c.endswith("Index")]
    df[z] = df[z].replace(0, np.nan)
    return df

def cargar_parametria():
    df = _read("Parametría", header=None)
    p  = df.iloc[3:, 1:7].copy()
    p.columns = ["alias","z_col","basket_col","bvl_col","bbg_ticker","sbs"]
    p  = p.dropna(subset=["alias"]).reset_index(drop=True)
    for c in p.columns: p[c] = p[c].astype(str).str.strip()
    return p

def cargar_gics():
    df = _read("GICS", header=0)
    df = df[["BBG","GICS","Empresa"]].dropna(subset=["BBG"]).copy()
    for c in ["BBG","GICS","Empresa"]: df[c] = df[c].astype(str).str.strip()
    return df.set_index("BBG")

def _load_dq_in():
    """
    Carga DQ_IN intentando primero QEST_BD.xlsm (fuente oficial);
    si no está disponible, cae en ZODA.xlsm como fallback.
    Devuelve DataFrame con header=0 donde la fila 1 tiene los SBS codes.
    En QEST_BD.xlsm los SBS codes están en la fila 11 (filas 1-10 = metadata);
    en ZODA.xlsm están en fila 1 (ya como header de pandas).
    """
    # Intentar QEST_BD.xlsm
    if os.path.exists(RUTA_QEST):
        try:
            raw = pd.read_excel(RUTA_QEST, sheet_name="DQ_IN",
                                header=None, engine="openpyxl", dtype=object)
            # Row 11 (0-indexed: row 10) has SBS codes
            raw.columns = [str(v) if v is not None else f"col_{i}"
                           for i, v in enumerate(raw.iloc[10].tolist())]
            df = raw.iloc[11:].copy().reset_index(drop=True)
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
            return df, True
        except Exception:
            pass
    # Fallback: ZODA.xlsm
    df = _read("DQ_IN", header=0)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    return df, False

def cargar_exposiciones():
    """
    Construye: {fecha_str → {sbs_code → exposure_PEN}}
    Sumando Fondo1+Fondo2+Fondo3 PxQ para cada fecha y cada SBS.
    """
    df, _ = _load_dq_in()
    pxq   = df[df["Dato"].astype(str) == "PxQ"].dropna(subset=["Fecha"]).copy()
    sbs_cols = [c for c in df.columns
                if c not in ("column_0","Fondo","Dato","Fecha","SBS","Fecha_dt","")
                and not str(c).startswith("col_")
                and str(c) not in ("nan","None","TEST")]

    result = {}
    for fecha, grupo in pxq.groupby("Fecha"):
        row_d = {}
        for sbs in sbs_cols:
            if sbs in grupo.columns:
                val = pd.to_numeric(grupo[sbs], errors="coerce").fillna(0).sum()
                if val > 0:
                    row_d[str(sbs)] = float(val)
        if row_d:
            result[fecha.strftime("%Y-%m-%d")] = row_d
    return result  # {date_str: {sbs: exposure_PEN}}

def cargar_posiciones():
    """Legacy wrapper — kept for backward compat with build_bubble_data."""
    exposiciones = cargar_exposiciones()
    if not exposiciones:
        return pd.DataFrame()
    records = []
    for fecha_str, sbs_dict in exposiciones.items():
        row = {"Fecha": pd.Timestamp(fecha_str)}
        row.update({sbs: v/100_000 for sbs, v in sbs_dict.items()})
        records.append(row)
    return pd.DataFrame(records).set_index("Fecha").sort_index()

# ─── MÚLTIPLOS (para Plot 05) ────────────────────────────────────────────────
def _limpiar(df):
    for v in _VACIOS: df.replace(v, 0, inplace=True)

def _forwards_one_year(base, names):
    fechas = pd.Series(base.index.tolist())
    dtoi   = {d: i for i, d in enumerate(fechas)}
    maxf   = fechas.max()
    for n in names:
        ci = names.index(n); vals = [None] * len(fechas)
        for idx, dia in enumerate(fechas):
            fwd  = (dia + pd.DateOffset(months=12)).replace(hour=0, minute=0, second=0)
            prev = fechas[fechas < fwd]
            if fwd in dtoi:
                vals[idx] = base.iloc[dtoi[fwd], ci]
            elif not prev.empty:
                vals[idx] = base.iloc[dtoi[prev.max()], ci]
        base[n] = vals
    return base

def _cd(ciq_col, ciq, fechas_q, fd):
    """CIQ trimestral → diario vía Period Q map con dedup."""
    s = ciq[ciq_col].copy()
    s.index = pd.DatetimeIndex(s.index).to_period("Q")
    s = s.groupby(level=0).last()          # dedup: 2026Q1 puede aparecer 2 veces
    return pd.Series(s.reindex(fechas_q).ffill().values, index=fd, name=ciq_col)

def _ratio(num, den):
    v = num.replace(0, np.nan) / den.replace(0, np.nan)
    v[np.isinf(v)] = np.nan
    return v.round(3)

def compute_multiplos_sheet(base, fechas_d, fechas_tri):
    """Calcula los 8 múltiplos diarios para todos los tickers de una hoja."""
    fd      = pd.DatetimeIndex(fechas_d)
    fechas_q = fd.to_period("Q")
    results  = {}
    tickers  = list(dict.fromkeys(
        [v for v in base.iloc[4].tolist()
         if isinstance(v, str) and "Equity" in v]
    ))

    for nombre in tickers:
        try:
            data_emp = base.loc[:, base.iloc[4] == nombre]
            if len(data_emp.columns) < 5: continue
            n_cols   = len(data_emp.columns)
            cols     = [_COLS_ALL[i] if i < n_cols and i < len(_COLS_ALL)
                        else f"col_{i}" for i in range(n_cols)]
            data_emp_= data_emp.iloc[5:].copy()
            data_emp_.columns = cols

            # CIQ
            ciq = data_emp_.iloc[:, :4].dropna(axis=0, how="all").fillna(0)
            n   = min(len(fechas_tri), len(ciq))
            if n == 0: continue
            ciq = ciq.iloc[:n].copy()
            ciq.insert(0, "Fecha", fechas_tri.values[:n])
            ciq.set_index("Fecha", inplace=True)
            _limpiar(ciq)
            ciq = ciq.astype(np.float64)
            ciq = _forwards_one_year(
                ciq, ["Fwd_IQ_NI","Fwd_IQ_TOTAL_REV",
                       "Fwd_IQ_TOTAL_EQUITY","Fwd_IQ_EBITDA"]
            )

            # BBG (no dropna)
            bbg = data_emp_.iloc[:, 4:]
            n2  = min(len(fechas_d), len(bbg))
            bbg = bbg.iloc[:n2].copy()
            bbg.insert(0, "Fecha", fechas_d.values[:n2])
            bbg.set_index("Fecha", inplace=True)
            _limpiar(bbg)
            bbg = bbg.apply(pd.to_numeric, errors="coerce").fillna(0)

            ev = bbg["CURR_ENTP_VAL"].reindex(fd)
            mc = bbg["CUR_MKT_CAP"].reindex(fd)

            m = pd.DataFrame({
                "Trail_PE":       _ratio(mc, _cd("IQ_NI",           ciq, fechas_q, fd)),
                "Fwd_PE":         _ratio(mc, _cd("Fwd_IQ_NI",       ciq, fechas_q, fd)),
                "Trail_PS":       _ratio(mc, _cd("IQ_TOTAL_REV",    ciq, fechas_q, fd)),
                "Fwd_PS":         _ratio(mc, _cd("Fwd_IQ_TOTAL_REV",ciq, fechas_q, fd)),
                "Trail_PB":       _ratio(mc, _cd("IQ_TOTAL_EQUITY", ciq, fechas_q, fd)),
                "Fwd_PB":         _ratio(mc, _cd("Fwd_IQ_TOTAL_EQUITY",ciq,fechas_q,fd)),
                "Trail_EVEBITDA": _ratio(ev, _cd("IQ_EBITDA",       ciq, fechas_q, fd)),
                "Fwd_EVEBITDA":   _ratio(mc, _cd("Fwd_IQ_EBITDA",   ciq, fechas_q, fd)),
            }, index=fd)

            # Hybrid resample: monthly for history > 5yr, bi-weekly for recent 5yr
            # Keeps ~400 pts/series vs 800 weekly → half the JSON size
            cutoff = pd.Timestamp.now() - pd.DateOffset(years=5)
            recent = m[m.index >= cutoff].resample("2W-FRI").last()
            older  = m[m.index <  cutoff].resample("ME").last()
            m = pd.concat([older, recent]).sort_index()
            m = m[~m.index.duplicated(keep="last")]

            ticker_data = {}
            for col in m.columns:
                clean = m[col].dropna()
                if len(clean) > 10:
                    ticker_data[col] = {
                        "x": clean.index.strftime("%Y-%m-%d").tolist(),
                        "y": clean.tolist(),
                    }
            if ticker_data:
                results[nombre] = ticker_data
        except Exception:
            pass
    return results


def cargar_multiplos_csv() -> tuple[dict, bool]:
    """
    Carga el CSV de múltiplos pre-calculados generado por ZODA_03_06_2026.py.

    Devuelve (mult_data, desde_csv) donde:
    - mult_data : dict {ticker → {multiplo → {"x": [...], "y": [...]}}}
                  Mismo formato que produce build_multiplos_data().
    - desde_csv : True si se leyó del CSV, False si el CSV no existe
                  (en cuyo caso el caller debe usar build_multiplos_data()).

    DISEÑO:
    - Aplica el mismo resample híbrido (mensual >5yr, bi-weekly reciente)
      que usa build_multiplos_data(), para consistencia en el dashboard.
    - Filas con todos los múltiplos NaN se descartan (tickers sin dato).
    - La columna "GeneradoEn" se usa para mostrar en el dashboard cuándo
      fue la última ejecución de ZODA_03_06_2026.py.
    """
    if not os.path.exists(RUTA_MULTIPLOS_CSV):
        return {}, False

    try:
        df = pd.read_csv(RUTA_MULTIPLOS_CSV, encoding="utf-8-sig",
                         parse_dates=["Fecha"], low_memory=False)
    except Exception as e:
        print(f"  ⚠ No se pudo leer {os.path.basename(RUTA_MULTIPLOS_CSV)}: {e}")
        return {}, False

    mult_cols = ["Trail_PE", "Fwd_PE", "Trail_PS", "Fwd_PS",
                 "Trail_PB", "Fwd_PB", "Trail_EVEBITDA", "Fwd_EVEBITDA"]
    # Columnas presentes en el CSV (pueden ser un subconjunto)
    mult_cols = [c for c in mult_cols if c in df.columns]

    # Timestamp de generación (para footer del dashboard)
    generado_en = ""
    if "GeneradoEn" in df.columns:
        generado_en = str(df["GeneradoEn"].dropna().iloc[0]) if not df["GeneradoEn"].dropna().empty else ""

    cutoff = pd.Timestamp.now() - pd.DateOffset(years=5)
    result = {}

    for ticker, grupo in df.groupby("Ticker"):
        grupo = grupo.sort_values("Fecha").set_index("Fecha")
        # Mantener solo columnas de múltiplos
        grupo = grupo[mult_cols]
        # Convertir a numérico (el CSV puede traer strings por NaN)
        grupo = grupo.apply(pd.to_numeric, errors="coerce")
        # Descartar tickers completamente vacíos
        if grupo.dropna(how="all").empty:
            continue
        # Resample híbrido: mensual para historia >5yr, bi-weekly para reciente
        reciente = grupo[grupo.index >= cutoff].resample("2W-FRI").last()
        antiguo  = grupo[grupo.index <  cutoff].resample("ME").last()
        grupo    = pd.concat([antiguo, reciente]).sort_index()
        grupo    = grupo[~grupo.index.duplicated(keep="last")]

        ticker_data = {}
        for col in mult_cols:
            clean = grupo[col].dropna()
            if len(clean) > 10:
                ticker_data[col] = {
                    "x": clean.index.strftime("%Y-%m-%d").tolist(),
                    "y": clean.round(3).tolist(),
                }
        if ticker_data:
            result[ticker] = ticker_data

    return result, True

def build_multiplos_data():
    """Carga las 3 hojas y construye la tabla de múltiplos para Plot 05."""
    all_m   = {}
    ticker_sheet = {}
    for sheet in ["Financials", "Cons y Const", "Min y Uti"]:
        df = _read(sheet, header=None, nrows=4020)
        df.columns = [f"c{i}" for i in range(len(df.columns))]
        fechas_tri = pd.to_datetime(df["c1"].iloc[5:], errors="coerce").dropna()
        mask_px    = df.iloc[2] == "Px_last"
        px_col     = df.columns[mask_px][0]
        fechas_d   = pd.to_datetime(df.loc[5:, px_col], errors="coerce").dropna()
        data       = compute_multiplos_sheet(df, fechas_d, fechas_tri)
        all_m.update(data)
        for t in data: ticker_sheet[t] = sheet
        print(f"  {sheet}: {len(data)} tickers")
    return all_m, ticker_sheet

# ─── INTERPRETACIÓN ─────────────────────────────────────────────────────────
def interpretar_z(z_val, percentil, empresa):
    if z_val is None or np.isnan(z_val):
        return "Sin dato", BRAND["muted"], "Sin dato suficiente", "No hay historia suficiente."
    pt = f"percentil {percentil:.0f}" if percentil is not None else ""
    if z_val > 1.5:
        return ("CARA", BRAND["red"],
                f"Z = {z_val:+.2f} — Valorización elevada ({pt})",
                f"{empresa} cotiza con prima histórica significativa. Múltiplo {z_val:.1f}σ "
                f"sobre su media. Revisar si existe catalizador justificado (M&A, guidance, momentum sectorial).")
    elif z_val > 0.5:
        return ("MODERADAMENTE CARA", BRAND["amber"],
                f"Z = {z_val:+.2f} — Leve presión alcista ({pt})",
                f"{empresa} muestra múltiplo por encima de su media. Monitorear si el avance "
                f"responde a momentum de precio o a revisión real de fundamentales.")
    elif z_val > -0.5:
        return ("NEUTRAL", BRAND["gray2"],
                f"Z = {z_val:+.2f} — Dentro de rango histórico ({pt})",
                f"{empresa} cotiza en zona de valorización neutra. "
                f"El Z Basket y Z BVL son más informativos en este rango.")
    elif z_val > -1.5:
        return ("MODERADAMENTE BARATA", BRAND["teal"],
                f"Z = {z_val:+.2f} — Descuento moderado ({pt})",
                f"{empresa} cotiza con descuento respecto a su historia. "
                f"Validar con Z Basket para confirmar la señal.")
    else:
        return ("BARATA", BRAND["green"],
                f"Z = {z_val:+.2f} — Descuento histórico significativo ({pt})",
                f"{empresa} en zona de descuento profundo. Señal de valor sostenida — "
                f"verificar que no esté justificada por deterioro estructural de fundamentales.")

# ─── BUILD DATA FUNCTIONS ────────────────────────────────────────────────────
# ── Z per window (para Chart 1 reponderación dinámica) ─────────────────────
# Pesos de múltiplos: idénticos a ZODA_03_06_2026.py
_PESOS_FINAN    = [0.33, 0, 0.33, 0, 0.33, 0, 0,    0]
_PESOS_EX_FINAN = [0.25, 0, 0.25, 0, 0.25, 0, 0.25, 0]
_MULT_COLS      = ["Trail_PE","Fwd_PE","Trail_PS","Fwd_PS",
                   "Trail_PB","Fwd_PB","Trail_EVEBITDA","Fwd_EVEBITDA"]

def _z_por_ventana_desde_multiplos(mult_dict: dict, es_finan: bool) -> dict:
    """
    Calcula el Z por ventana (1Y, 3Y, 5Y, 10Y) a partir de los múltiplos
    reales del ticker — la misma lógica que Creacion_Z_Final() en ZODA_03_06_2026.py,
    aplicada ventana a ventana de forma separada para permitir reponderación en JS.

    mult_dict: {col_name → {"x": [dates], "y": [values]}}  (formato de cargar_multiplos_csv)
    es_finan:  True para bancos/financieras, False para el resto

    Con datos bi-semanales (26 puntos/año) los mínimos de ventana son:
      1Y = 26 pts, 3Y = 78 pts, 5Y = 130 pts, 10Y = 260 pts
    """
    pesos = _PESOS_FINAN if es_finan else _PESOS_EX_FINAN

    # Construir Series individuales por múltiplo
    series_mult = {}
    for i, col in enumerate(_MULT_COLS):
        if pesos[i] == 0 or col not in mult_dict:
            continue
        d = mult_dict[col]
        if not d or not d.get("x"):
            continue
        s = pd.Series(d["y"], index=pd.to_datetime(d["x"]), name=col).sort_index()
        s = pd.to_numeric(s, errors="coerce")
        series_mult[col] = (s, pesos[i])

    if not series_mult:
        return {}

    # Índice unificado (unión de todas las fechas disponibles)
    idx_all = pd.DatetimeIndex([])
    for s, _ in series_mult.values():
        idx_all = idx_all.union(s.dropna().index)
    if idx_all.empty:
        return {}

    resultado = {}
    # 26 puntos/año con datos bi-semanales
    pts_por_anio = 26
    for anios, etiqueta in [(1,"1Y"),(3,"3Y"),(5,"5Y"),(10,"10Y")]:
        window = anios * pts_por_anio
        z_parts = []
        for col, (s, w) in series_mult.items():
            s_r = s.reindex(idx_all)
            m   = s_r.rolling(window, min_periods=window).mean()
            std = s_r.rolling(window, min_periods=window).std()
            z   = ((s_r - m) / std * w)
            z_parts.append(z)
        if not z_parts:
            continue
        combined = pd.concat(z_parts, axis=1).sum(axis=1, min_count=1).dropna()
        if len(combined) >= 15:
            resultado[etiqueta] = {
                "x": combined.index.strftime("%Y-%m-%d").tolist(),
                "y": combined.round(4).tolist(),
            }
    return resultado

def build_historico(bd, param, gics, mult_data: dict = None):
    """
    Construye los datos del Chart 1.
    mult_data: dict {bbg_ticker → {multiplo → {x,y}}} — opcional.
      Si está presente, calcula z_windows (Z por ventana 1Y/3Y/5Y/10Y) desde
      los múltiplos reales, permitiendo al usuario reponderarlas en el dashboard.
      Si no está, z_windows queda vacío y el Chart 1 solo muestra el Z compuesto
      pre-calculado de BD (sin reponderación dinámica).
    """
    traces, interps = [], {}
    for _, r in param.iterrows():
        alias, z_col, bk, bvl = r["alias"], r["z_col"], r["basket_col"], r["bvl_col"]
        bbg = r["bbg_ticker"]
        if z_col not in bd.columns or z_col.startswith("nan"): continue
        sector = gics.loc[bbg,"GICS"] if bbg in gics.index else "Index"
        nombre = gics.loc[bbg,"Empresa"] if bbg in gics.index else alias
        serie  = bd[["Fecha",z_col]].dropna(subset=[z_col]).copy()
        if serie.empty: continue
        z_act  = serie[z_col].iloc[-1]
        pct    = float((serie[z_col] < z_act).mean() * 100)
        bk_ok  = bk if bk in bd.columns else None
        bvl_ok = bvl if bvl in bd.columns else None

        # Z por ventana desde los múltiplos reales (para reponderación en JS)
        # Usa el bbg_ticker para buscar en mult_data (mismo ticker que en ZODA_03_06_2026)
        z_windows = {}
        if mult_data and bbg in mult_data:
            es_finan = sector == "Financials"
            z_windows = _z_por_ventana_desde_multiplos(mult_data[bbg], es_finan)

        traces.append(dict(
            alias=alias,
            x=serie["Fecha"].dt.strftime("%Y-%m-%d").tolist(),
            y=serie[z_col].round(4).tolist(),
            y_basket=(bd[["Fecha",bk_ok]].dropna(subset=[bk_ok])[bk_ok].round(4).tolist() if bk_ok else []),
            x_basket=(bd[["Fecha",bk_ok]].dropna(subset=[bk_ok])["Fecha"].dt.strftime("%Y-%m-%d").tolist() if bk_ok else []),
            y_bvl=(bd[["Fecha",bvl_ok]].dropna(subset=[bvl_ok])[bvl_ok].round(4).tolist() if bvl_ok else []),
            x_bvl=(bd[["Fecha",bvl_ok]].dropna(subset=[bvl_ok])["Fecha"].dt.strftime("%Y-%m-%d").tolist() if bvl_ok else []),
            color=SECTOR_COLOR.get(sector, BRAND["muted"]),
            sector=sector, empresa=nombre,
            z_actual=round(float(z_act),4) if not np.isnan(z_act) else None,
            percentil=round(pct,1),
            z_windows=z_windows,   # {"1Y":{x,y}, "3Y":{x,y}, "5Y":{x,y}, "10Y":{x,y}}
        ))
        señal, sc, tc, tl = interpretar_z(z_act, pct, nombre)
        interps[alias] = dict(señal=señal, color=sc, corto=tc, largo=tl,
                              z_actual=round(float(z_act),4) if not np.isnan(z_act) else None,
                              percentil=round(pct,1))
    return traces, interps

def build_bd_col_sbs_map(param):
    """
    Construye un dict {bd_col → sbs_code} a partir de Parametría.
    Columna C (z_col), D (basket_col), E (bvl_col) → todas apuntan al mismo SBS (col G).
    Índices sin SBS (índices de mercado) quedan excluidos.
    """
    mapping = {}
    for _, r in param.iterrows():
        sbs = str(r["sbs"]).strip() if pd.notna(r.get("sbs")) else ""
        if not sbs or sbs in ("nan","None",""):
            continue
        for col_key in ("z_col","basket_col","bvl_col"):
            col_val = str(r.get(col_key,"")).strip()
            if col_val and col_val not in ("nan","None",""):
                mapping[col_val] = sbs
    return mapping

def _safe_z(val):
    """Extrae un float de un valor que puede ser Series, ndarray o scalar.
    Devuelve np.nan si no es numérico — previene que Plotly interprete
    objetos pandas como timestamps en el eje X."""
    try:
        v = float(val)
        return np.nan if np.isnan(v) or np.isinf(v) else v
    except (TypeError, ValueError):
        return np.nan

def build_bubble_data(bd, param, gics, pos):
    """
    Construye {fecha_str → [lista de puntos]} para el bubble chart.

    Cada punto contiene:
      alias, empresa, sector, color — identificación y estilo
      z_abs    — eje X: Z-score absoluto (float puro, nunca timestamp)
      z_basket — eje Y opción 1: Z vs Basket
      z_bvl    — eje Y opción 2: Z vs BVL
      size     — tamaño ∝ exposición en S/. millones (Fondo1+2+3 PxQ)

    BUG CORREGIDO: row.get() devuelve un elemento de Series que puede
    tener metadatos de índice datetime. _safe_z() lo convierte a float
    puro antes de pasarlo a Plotly, evitando que el eje X muestre
    timestamps (1969-12-31 18:59:59.xxx) en lugar de Z-scores.
    """
    # Columnas Z absolutas (no basket, no BVL, no índices MXLA)
    z_emp = [c for c in bd.columns if c.startswith("Z_")
             and "Basket" not in c and "BVL" not in c
             and "MXLA" not in c and "Index" not in c]

    # Fechas con al menos un Z no-NaN — últimas 400 para no saturar el HTML
    fechas = sorted(
        bd.loc[bd[z_emp].notna().any(axis=1), "Fecha"].dt.normalize().unique()
    )[-400:]

    # Todos los activos mapeados en Parametría (empresas + índices de referencia)
    ems = [r for _, r in param.iterrows()
           if str(r["z_col"]) in bd.columns
           and not str(r["z_col"]).startswith("nan")]

    # Fechas disponibles en exposiciones (dict date_str → {sbs → PEN})
    exp_dates = sorted(pos.keys()) if isinstance(pos, dict) else []

    result = {}
    for fecha in fechas:
        fkey = fecha.strftime("%Y-%m-%d")

        # Fila de BD para esta fecha (última disponible ≤ fecha)
        mask = bd["Fecha"].dt.normalize() == fecha
        if not mask.any():
            continue
        # Extraer como dict Python puro para evitar cualquier metadato pandas
        row_series = bd[mask].iloc[-1]
        row = {col: row_series[col] for col in bd.columns}

        # Exposición más cercana disponible ≤ fecha del chart
        pr = None
        if exp_dates:
            avail = [d for d in exp_dates if d <= fkey]
            pr = pos[avail[-1]] if avail else None

        pts = []
        for r in ems:
            alias   = r["alias"]
            z_col   = r["z_col"]
            bk      = r["basket_col"]
            bvl     = r["bvl_col"]
            sbs     = str(r["sbs"]).strip()
            bbg     = r["bbg_ticker"]

            # Extraer Z-scores como floats puros (fix del bug de timestamps)
            z_abs = _safe_z(row.get(z_col, np.nan))
            z_bk  = _safe_z(row.get(bk,    np.nan)) if bk  in bd.columns else np.nan
            z_bvl = _safe_z(row.get(bvl,   np.nan)) if bvl in bd.columns else np.nan

            # Saltar si el Z absoluto no existe para esta fecha
            if np.isnan(z_abs):
                continue

            sector = gics.loc[bbg, "GICS"]    if bbg in gics.index else "Index"
            nombre = gics.loc[bbg, "Empresa"] if bbg in gics.index else alias

            # Exposición: Fondo1+2+3 PxQ en S/. millones
            # Índices (sin SBS): burbuja fija de referencia (sz=0 → tamaño mínimo)
            sz = 0.0
            if sbs and sbs not in ("nan", "None", "") and pr is not None:
                if isinstance(pr, dict) and sbs in pr and pr[sbs] > 0:
                    sz = float(pr[sbs]) / 1e6      # S/. millones

            pts.append(dict(
                alias=alias, empresa=nombre, sector=sector,
                color=SECTOR_COLOR.get(sector, BRAND["muted"]),
                z_abs=round(z_abs, 4),
                z_basket=round(z_bk,  4) if not np.isnan(z_bk)  else None,
                z_bvl=round(z_bvl, 4) if not np.isnan(z_bvl) else None,
                size=round(sz, 2),              # S/. millones
            ))

        if pts:
            result[fkey] = pts

    return result

def build_heatmap_data(bd, param, gics):
    ems = [r for _,r in param.iterrows()
           if str(r["z_col"]) in bd.columns
           and not str(r["z_col"]).startswith("nan")
           and "Index" not in str(r["z_col"])]
    last   = bd["Fecha"].max()
    start  = last - pd.DateOffset(months=12)
    weeks  = pd.date_range(start=start, end=last, freq="W-FRI")
    bds    = bd.sort_values("Fecha")
    rows, labels, secs, nombres = [], [], [], []
    for r in ems:
        z_col, bbg = r["z_col"], r["bbg_ticker"]
        sector = gics.loc[bbg,"GICS"] if bbg in gics.index else "Other"
        nombre = gics.loc[bbg,"Empresa"] if bbg in gics.index else r["alias"]
        s = bds[["Fecha",z_col]].dropna(subset=[z_col]).set_index("Fecha")[z_col]
        if s.empty: continue
        rv = [round(float(s[s.index<=v].iloc[-1]),4) if not s[s.index<=v].empty else None for v in weeks]
        if sum(v is not None for v in rv) < len(weeks)//2: continue
        rows.append(rv); labels.append(r["alias"]); secs.append(sector); nombres.append(nombre)
    return dict(z=rows, x=[f.strftime("%Y-%m-%d") for f in weeks],
                y=labels, sectores=secs, empresas=nombres)

def build_eqrv_data(bd, param, gics):
    SO = ["Financials","Consumer","Construction Materials",
          "Industrials","Health Care","Media","Mining","Utilities","Other"]
    results = []
    for _, r in param.iterrows():
        z_col, bbg = r["z_col"], r["bbg_ticker"]
        if z_col not in bd.columns or z_col.startswith("nan") or "Index" in z_col: continue
        s = bd[z_col].dropna()
        if len(s) < 52: continue
        sector = gics.loc[bbg,"GICS"] if bbg in gics.index else "Other"
        nombre = gics.loc[bbg,"Empresa"] if bbg in gics.index else r["alias"]
        cur    = float(s.iloc[-1]); pct = float((s < cur).mean()*100)
        results.append(dict(alias=r["alias"], empresa=nombre, sector=sector,
                            color=SECTOR_COLOR.get(sector, BRAND["muted"]),
                            current=round(cur,3),
                            p10=round(float(s.quantile(.10)),3),
                            p25=round(float(s.quantile(.25)),3),
                            p50=round(float(s.quantile(.50)),3),
                            p75=round(float(s.quantile(.75)),3),
                            p90=round(float(s.quantile(.90)),3),
                            pct_rank=round(pct,1), n_obs=int(len(s))))
    results.sort(key=lambda x: (SO.index(x["sector"]) if x["sector"] in SO else 99, -x["current"]))
    return results

# ─── HTML ────────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ZODA · AFP Integra</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Calibri:wght@300;400;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root {
  --navy:   #1E2E6E;
  --dark:   #002060;
  --teal:   #00AECB;
  --yellow: #E3E829;
  --bg:     #FFFFFF;
  --surf1:  #F5F7FA;
  --surf2:  #EEF1F6;
  --border: #D0D7E5;
  --text:   #1a2340;
  --muted:  #6B7A99;
  --green:  #1a8a4a;
  --red:    #c0392b;
  --amber:  #b37a00;
  --mono:   'IBM Plex Mono', monospace;
  --sans:   'Calibri', 'Segoe UI', sans-serif;
}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html { scroll-behavior: smooth; }
body { background: var(--bg); color: var(--text); font-family: var(--sans);
       font-size: 13.5px; min-height: 100vh; }

/* ── HEADER ── */
.header {
  background: var(--navy);
  border-bottom: 4px solid var(--yellow);
  padding: 0 32px;
  height: 56px;
  display: flex; align-items: center; gap: 20px;
  position: sticky; top: 0; z-index: 200;
  box-shadow: 0 2px 20px rgba(0,0,0,0.5);
}
.header-logo { height: 30px; width: auto; object-fit: contain; flex-shrink: 0; }
.header-logo-ph {
  height: 30px; padding: 0 10px;
  background: rgba(255,255,255,0.1); border-radius: 3px;
  display: flex; align-items: center;
  font-family: var(--mono); font-size: 10px;
  color: rgba(255,255,255,0.5); letter-spacing: 0.1em;
}
.header-sep { width: 1px; height: 28px; background: rgba(255,255,255,0.18); flex-shrink: 0; }
.header-brand { display: flex; flex-direction: column; gap: 1px; }
.header-brand-title {
  font-family: var(--mono); font-size: 13px; font-weight: 500;
  color: #fff; letter-spacing: 0.08em;
}
.header-brand-sub {
  font-size: 10px; color: rgba(255,255,255,0.45);
  letter-spacing: 0.04em;
}
.header-space { flex: 1; }
.header-right { display: flex; flex-direction: column; align-items: flex-end; gap: 1px; }
.header-updated {
  font-family: var(--mono); font-size: 9.5px;
  color: var(--teal); letter-spacing: 0.05em;
}
.header-author { font-size: 9px; color: rgba(255,255,255,0.2); font-style: italic; }

/* ── NAV ── */
.nav {
  background: #FFFFFF;
  border-bottom: 2px solid var(--navy);
  padding: 0 32px;
  display: flex; gap: 0;
  overflow-x: auto;
}
.nav-tab {
  font-family: var(--mono); font-size: 10px; letter-spacing: 0.06em;
  text-transform: uppercase; color: var(--muted);
  padding: 10px 18px; border: none; background: none;
  cursor: pointer; white-space: nowrap; border-bottom: 2px solid transparent;
  transition: all .15s;
}
.nav-tab:hover { color: var(--text); }
.nav-tab.active { color: var(--teal); border-bottom-color: var(--teal); }

/* ── MAIN ── */
.main { padding: 28px 32px; max-width: 1480px; margin: 0 auto; }
.section { margin-bottom: 56px; }

.section-header {
  display: flex; align-items: center; gap: 14px;
  margin-bottom: 18px; padding-bottom: 12px;
  border-bottom: 1px solid var(--border);
}
.section-num {
  font-family: var(--mono); font-size: 9px; font-weight: 500;
  letter-spacing: 0.15em; text-transform: uppercase;
  color: var(--teal);
  background: rgba(0,174,203,0.09);
  border: 1px solid rgba(0,174,203,0.22);
  padding: 3px 8px; border-radius: 3px;
}
.section-title { font-size: 14px; font-weight: 700; letter-spacing: 0.01em; }
.section-desc  { font-size: 11px; color: var(--muted); margin-left: auto; font-style: italic; }

/* ── CONTROLS ── */
.controls { display: flex; gap: 14px; align-items: flex-end; flex-wrap: wrap; margin-bottom: 16px; }
.ctrl-group { display: flex; flex-direction: column; gap: 5px; }
.ctrl-label {
  font-family: var(--mono); font-size: 9px; text-transform: uppercase;
  letter-spacing: 0.1em; color: var(--muted);
}
select, input[type=date] {
  background: #FFFFFF; border: 1px solid var(--border);
  color: var(--text); font-family: var(--mono); font-size: 11.5px;
  padding: 7px 11px; border-radius: 4px; outline: none; cursor: pointer;
  min-width: 200px; transition: border-color .15s, box-shadow .15s;
}
select:hover, input[type=date]:hover,
select:focus, input[type=date]:focus {
  border-color: var(--teal); box-shadow: 0 0 0 2px rgba(0,174,203,0.15);
}
.tog { display: flex; border: 1px solid var(--border); border-radius: 4px; overflow: hidden; }
.tog-btn {
  background: #F5F7FA; border: none; color: var(--muted);
  font-family: var(--mono); font-size: 10px; padding: 7px 14px;
  cursor: pointer; transition: all .15s; white-space: nowrap;
}
.tog-btn.active { background: var(--navy); color: var(--teal); font-weight: 600; border-bottom: 2px solid var(--teal); }
.tog-btn:hover:not(.active) { background: #D0D7E5; color: var(--navy); }

/* ── PLOT CARD ── */
.plot-card {
  background: #FFFFFF; border: 1px solid var(--border);
  border-radius: 6px; overflow: hidden;
  box-shadow: 0 1px 6px rgba(30,46,110,0.08);
}
.plot-card-accent { border-top: 3px solid var(--navy); }

/* ── INTERPRETATION ── */
.interp {
  margin-top: 12px; background: #F0F8FC;
  border: 1px solid #B8DCE8; border-left: 3px solid var(--teal);
  border-radius: 6px; padding: 15px 20px;
  display: grid; grid-template-columns: auto 1fr; gap: 16px; align-items: start;
}
.interp-badge {
  font-family: var(--mono); font-size: 9.5px; font-weight: 600;
  letter-spacing: 0.1em; padding: 4px 10px; border-radius: 3px;
  text-transform: uppercase; white-space: nowrap;
}
.interp-short { font-size: 12px; font-weight: 700; margin-bottom: 5px; color: var(--text); }
.interp-long  { font-size: 11px; color: var(--muted); line-height: 1.65; }
.interp-meta  { font-family: var(--mono); font-size: 9.5px; color: var(--muted); margin-top: 8px; }

/* ── LEGEND ROW ── */
.leg-row { display: flex; flex-wrap: wrap; gap: 14px; margin-bottom: 12px; }
.leg-item {
  display: flex; align-items: center; gap: 6px;
  font-size: 10.5px; color: var(--text); cursor: pointer; user-select: none;
}
.leg-dot { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }

/* ── FOOTNOTE ── */
.footnote {
  font-size: 9px; color: var(--muted); line-height: 1.7;
  margin-top: 11px; padding-left: 8px;
  border-left: 2px solid var(--teal);
}
.footnote em { color: rgba(255,255,255,0.35); }

/* ── WEIGHT PANEL (Chart 1) ── */
.weight-panel {
  background: #F0F8FC; border: 1px solid #B8DCE8;
  border-left: 3px solid var(--navy);
  border-radius: 5px; padding: 10px 16px;
  margin-bottom: 14px; display: flex;
  flex-direction: column; gap: 8px;
}
.weight-label-title {
  font-family: var(--mono); font-size: 9px; text-transform: uppercase;
  letter-spacing: 0.1em; color: var(--navy); font-weight: 600;
}
.weight-warn { color: #c0392b; font-weight: 700; margin-left: 8px; }
.weight-row { display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }
.weight-item { display: flex; flex-direction: column; gap: 3px; }
.weight-lbl {
  font-family: var(--mono); font-size: 9px; color: var(--muted);
  text-align: center; letter-spacing: 0.05em;
}
.weight-inp {
  width: 68px; text-align: center; font-family: var(--mono);
  font-size: 12px; font-weight: 600; color: var(--navy);
  background: #fff; border: 1px solid var(--border);
  border-radius: 4px; padding: 5px 6px; outline: none;
  transition: border-color .15s, box-shadow .15s;
}
.weight-inp:focus { border-color: var(--navy); box-shadow: 0 0 0 2px rgba(30,46,110,0.12); }
.weight-inp:hover { border-color: var(--teal); }
.weight-reset {
  font-family: var(--mono); font-size: 10px; color: var(--muted);
  background: none; border: 1px solid var(--border); border-radius: 4px;
  padding: 5px 10px; cursor: pointer; transition: all .15s; white-space: nowrap;
}
.weight-reset:hover { background: var(--navy); color: #fff; border-color: var(--navy); }
.weight-sum {
  font-family: var(--mono); font-size: 11px; font-weight: 700;
  color: var(--navy); margin-left: 4px;
}

/* ── New chart selector control ── */
.chart-sel-row { display:flex; gap:12px; align-items:flex-end; flex-wrap:wrap; margin-bottom:16px; }
.period-tog .tog-btn { padding:5px 12px; font-size:9.5px; }

/* ── FOOTER ── */
.footer {
  background: var(--dark); border-top: 4px solid var(--yellow);
  padding: 14px 32px; display: flex; align-items: center; gap: 16px; margin-top: 40px;
}
.footer-brand { font-family: var(--mono); font-size: 9px; color: var(--teal); letter-spacing: 0.1em; text-transform: uppercase; white-space: nowrap; }
.footer-disc  { font-size: 9px; color: var(--muted); line-height: 1.5; flex: 1; }
.footer-author{ font-size: 9px; color: rgba(255,255,255,0.2); font-style: italic; white-space: nowrap; }

/* ── LOADING ── */
#loading {
  position: fixed; inset: 0; background: #FFFFFF;
  display: flex; flex-direction: column; align-items: center; justify-content: center;
  z-index: 999; gap: 20px;
}
.spinner {
  width: 30px; height: 30px; border: 2px solid var(--border);
  border-top-color: var(--teal); border-radius: 50%;
  animation: spin .65s linear infinite;
}
.loading-title { font-family: var(--mono); font-size: 12px; color: var(--navy); letter-spacing: 0.08em; }
.loading-bar-wrap { width: 200px; height: 2px; background: var(--border); border-radius: 2px; }
.loading-bar { height: 2px; background: var(--teal); border-radius: 2px; width: 0; transition: width .3s; }
@keyframes spin { to { transform: rotate(360deg); } }
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #F5F7FA; }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
</style>
</head>
<body>

<div id="loading">
  <div class="spinner"></div>
  <div class="loading-title">ZODA Dashboard · AFP Integra</div>
  <div class="loading-bar-wrap"><div class="loading-bar" id="lbar"></div></div>
</div>

<!-- HEADER -->
<div class="header">
  <img id="logo-img" class="header-logo" src="__LOGO_B64__" alt="AFP Integra"
       onerror="this.style.display='none';document.getElementById('logo-ph').style.display='flex'">
  <div id="logo-ph" class="header-logo-ph" style="display:none">AFP INTEGRA</div>
  <div class="header-sep"></div>
  <div class="header-brand">
    <div class="header-brand-title">ZODA</div>
    <div class="header-brand-sub">Monitor de Valorización · Renta Variable</div>
  </div>
  <div class="header-space"></div>
  <div class="header-right">
    <div class="header-updated" id="hdr-date"></div>
    <div class="header-author">Franco Olivares</div>
  </div>
</div>

<!-- NAV -->
<div class="nav">
  <button class="nav-tab active" onclick="scrollTo('s1')">01 · Z Histórico</button>
  <button class="nav-tab"        onclick="scrollTo('s2')">02 · Monitor</button>
  <button class="nav-tab"        onclick="scrollTo('s3')">03 · Régimen</button>
  <button class="nav-tab"        onclick="scrollTo('s4')">04 · EQRV</button>
  <button class="nav-tab"        onclick="scrollTo('s5')">05 · Múltiplos</button>
  <button class="nav-tab"        onclick="scrollTo('s6')">06 · YTD Fan</button>
</div>

<div class="main">

<!-- ══ 01 HISTÓRICO ══ -->
<div class="section" id="s1">
  <div class="section-header">
    <span class="section-num">01</span>
    <span class="section-title">Evolución Histórica del Z-Score</span>
    <span class="section-desc" id="s1-desc">Z compuesto · 1Y×33% + 3Y×33% + 5Y×33%</span>
  </div>
  <div class="controls">
    <div class="ctrl-group">
      <span class="ctrl-label">Empresa</span>
      <select id="sel-empresa" onchange="onEmpresa()"></select>
    </div>
    <div class="ctrl-group">
      <span class="ctrl-label">Perspectiva</span>
      <div class="tog">
        <button class="tog-btn active" id="btn-z-abs"    onclick="onZType('abs')">Absoluto</button>
        <button class="tog-btn"        id="btn-z-basket" onclick="onZType('basket')">vs Basket</button>
        <button class="tog-btn"        id="btn-z-bvl"    onclick="onZType('bvl')">vs BVL</button>
      </div>
    </div>
  </div>
  <!-- Panel de pesos por ventana temporal — solo Chart 1 -->
  <div class="weight-panel" id="weight-panel">
    <span class="weight-label-title">Pesos por ventana de tiempo
      <span id="weight-warn" class="weight-warn" style="display:none">⚠ La suma de pesos debe ser 1.00</span>
    </span>
    <div class="weight-row">
      <div class="weight-item">
        <label class="weight-lbl" for="w1y">1 año</label>
        <input type="number" id="w1y" class="weight-inp" value="0.33"
               min="0" max="1" step="0.01" oninput="onWeightChange()">
      </div>
      <div class="weight-item">
        <label class="weight-lbl" for="w3y">3 años</label>
        <input type="number" id="w3y" class="weight-inp" value="0.33"
               min="0" max="1" step="0.01" oninput="onWeightChange()">
      </div>
      <div class="weight-item">
        <label class="weight-lbl" for="w5y">5 años</label>
        <input type="number" id="w5y" class="weight-inp" value="0.33"
               min="0" max="1" step="0.01" oninput="onWeightChange()">
      </div>
      <div class="weight-item">
        <label class="weight-lbl" for="w10y">10 años</label>
        <input type="number" id="w10y" class="weight-inp" value="0.00"
               min="0" max="1" step="0.01" oninput="onWeightChange()"
               title="0 por defecto — historia insuficiente en LatAm">
      </div>
      <button class="weight-reset" onclick="resetWeights()">↺ Default</button>
      <span class="weight-sum" id="weight-sum">Σ = 0.99</span>
    </div>
  </div>
  <div class="plot-card plot-card-accent"><div id="plot-hist" style="height:360px;"></div></div>
  <div class="interp" id="interp-panel">
    <span class="interp-badge" id="interp-badge">—</span>
    <div>
      <div class="interp-short" id="interp-short">Selecciona una empresa</div>
      <div class="interp-long"  id="interp-long"></div>
      <div class="interp-meta"  id="interp-meta"></div>
    </div>
  </div>
</div>

<!-- ══ 02 BUBBLE ══ -->
<div class="section" id="s2">
  <div class="section-header">
    <span class="section-num">02</span>
    <span class="section-title">Monitor de Posición — Mapa de Valorización</span>
    <span class="section-desc">Burbuja ∝ posición consolidada Fondos 1+2+3 (S/. 100k)</span>
  </div>
  <div class="controls">
    <div class="ctrl-group">
      <span class="ctrl-label">Fecha</span>
      <input type="date" id="sel-fecha-bubble" onchange="onBubbleFecha()">
    </div>
    <div class="ctrl-group">
      <span class="ctrl-label">Eje Y</span>
      <div class="tog">
        <button class="tog-btn active" id="btn-y-basket" onclick="onYAxis('basket')">vs Basket</button>
        <button class="tog-btn"        id="btn-y-bvl"    onclick="onYAxis('bvl')">vs BVL</button>
      </div>
    </div>
  </div>
  <div class="leg-row" id="bubble-legend"></div>
  <div class="plot-card plot-card-accent"><div id="plot-bubble" style="height:520px;"></div></div>
</div>

<!-- ══ 03 HEATMAP ══ -->
<div class="section" id="s3">
  <div class="section-header">
    <span class="section-num">03</span>
    <span class="section-title">Régimen de Valorización — Mapa de Calor 12 meses</span>
    <span class="section-desc">Frecuencia semanal · Verde = barata · Rojo = cara</span>
  </div>
  <div class="plot-card plot-card-accent"><div id="plot-heat" style="min-height:480px;"></div></div>
  <p class="footnote">¹ Inspirado en el factor exposure dashboard de BlackRock Investment Institute.
  Ref.: Asness, C., Moskowitz, T. &amp; Pedersen, L. (2013).
  <em>Value and Momentum Everywhere. Journal of Finance, 68(3), 929–985.</em>
  Señales de valor sostenidas ≥ 3 meses (Z &lt; −1.5) presentan estadísticamente mayor probabilidad de reversión hacia la media.</p>
</div>

<!-- ══ 04 EQRV ══ -->
<div class="section" id="s4">
  <div class="section-header">
    <span class="section-num">04</span>
    <span class="section-title">Distribución Histórica del Z-Score — Estilo EQRV</span>
    <span class="section-desc">Valor actual vs percentiles históricos · Agrupado por sector</span>
  </div>
  <div class="controls">
    <div class="ctrl-group">
      <span class="ctrl-label">Ordenar por</span>
      <div class="tog">
        <button class="tog-btn active" id="btn-sort-sec" onclick="onEqrvSort('sector')">Sector</button>
        <button class="tog-btn"        id="btn-sort-z"   onclick="onEqrvSort('z')">Z actual ↓</button>
        <button class="tog-btn"        id="btn-sort-pct" onclick="onEqrvSort('pct')">Percentil ↓</button>
      </div>
    </div>
  </div>
  <div class="plot-card plot-card-accent"><div id="plot-eqrv" style="height:520px;"></div></div>
  <p class="footnote">² Inspirado en el <em>Equity Relative Valuation (EQRV)</em> de Bloomberg Terminal
  y en los gráficos de percentil de Forward PE de J.P. Morgan Asset Management — <em>Guide to the Markets Q1 2026.</em>
  Banda p10–p90 (rango amplio) · Caja p25–p75 (IQR) · Línea = mediana · Diamante = Z actual.</p>
</div>

<!-- ══ 05 MÚLTIPLOS ══ -->
<div class="section" id="s5">
  <div class="section-header">
    <span class="section-num">05</span>
    <span class="section-title">Evolución de Múltiplos de Valorización</span>
    <span class="section-desc">Datos diarios de Financials · Cons y Const · Min y Uti · hasta 3 equities comparables</span>
  </div>
  <div class="controls">
    <div class="ctrl-group">
      <span class="ctrl-label">Múltiplo</span>
      <select id="sel-multiplo" onchange="renderMultiplos()">
        <option value="Trail_PE">P/E Trailing</option>
        <option value="Fwd_PE">P/E Forward</option>
        <option value="Trail_PS">P/S Trailing</option>
        <option value="Fwd_PS">P/S Forward</option>
        <option value="Trail_PB">P/B Trailing</option>
        <option value="Fwd_PB">P/B Forward</option>
        <option value="Trail_EVEBITDA">EV/EBITDA Trailing</option>
        <option value="Fwd_EVEBITDA">EV/EBITDA Forward</option>
      </select>
    </div>
    <div class="ctrl-group">
      <span class="ctrl-label">Equity 1 (principal)</span>
      <select id="sel-m1" onchange="renderMultiplos()"></select>
    </div>
    <div class="ctrl-group">
      <span class="ctrl-label">Equity 2 (comparar)</span>
      <select id="sel-m2" onchange="renderMultiplos()">
        <option value="">— ninguno —</option>
      </select>
    </div>
    <div class="ctrl-group">
      <span class="ctrl-label">Equity 3 (comparar)</span>
      <select id="sel-m3" onchange="renderMultiplos()">
        <option value="">— ninguno —</option>
      </select>
    </div>
  </div>
  <div class="plot-card plot-card-accent"><div id="plot-mult" style="height:380px;"></div></div>
  <p class="footnote">³ Múltiplos calculados sobre datos semanales (viernes).
  Numeradores BBG: Market Cap (P/E, P/S, P/B, Fwd EV/EBITDA) y Enterprise Value (EV/EBITDA Trailing).
  Denominadores CIQ: Utilidad Neta, Ventas, Patrimonio y EBITDA trailing / forward 12 meses.
  Fuentes: Bloomberg L.P. y S&P Capital IQ.</p>
</div>


  <!-- ══ 06 REVERSIÓN A LA MEDIA ══ -->
  <div class="section" id="s6">
    <div class="section-header">
      <span class="section-num">06</span>
      <span class="section-title">Calendarized Return Fan — Retorno YTD vs historia</span>
      <span class="section-desc">Año actual vs bandas históricas · Goldman Sachs GIR style</span>
    </div>
    <div class="controls">
      <div class="ctrl-group">
        <span class="ctrl-label">Empresa</span>
        <select id="sel-fan" onchange="renderFan()"></select>
      </div>
    </div>
    <div class="plot-card plot-card-accent"><div id="plot-fan" style="height:400px;"></div></div>
    <p class="footnote">⁹ Doeswijk, R., Lam, T. &amp; Swinkels, L. (2020). <em>Historical Returns of the Market Portfolio.</em> Review of Asset Pricing Studies, 10(3), 521–567. Los retornos YTD extremos (fuera del P90 histórico) presentan mayor probabilidad de reversión inter-año. Permite cuantificar cuánto del rally ya está estadísticamente "consumido".</p>
  </div>

</div><!-- /main -->

<div class="footer">
  <span class="footer-brand">AFP Integra · ZODA</span>
  <span class="footer-disc">Uso interno exclusivo. Información elaborada a partir de datos de Bloomberg y Capital IQ.
  No constituye asesoramiento de inversión. Indicadores cuantitativos sujetos a revisión del equipo de inversiones.</span>
  <span class="footer-author">Franco Olivares</span>
</div>

<script>
// ── DATOS ────────────────────────────────────────────────────────────────────
const HIST_DATA   = __HIST_DATA__;
const INTERP_DATA = __INTERP_DATA__;
const BUBBLE_DATA = __BUBBLE_DATA__;
const HEAT_DATA   = __HEAT_DATA__;
const EQRV_DATA   = __EQRV_DATA__;
const MULT_DATA   = __MULT_DATA__;
const GEN_DATE    = "__GEN_DATE__";
const MULT_LABELS     = __MULT_LABELS__;
const CALENDAR_DATA   = __CALENDAR_DATA__;

// ── ESTADO ───────────────────────────────────────────────────────────────────
let E = { empresa:'', ztype:'abs', yaxis:'basket', bubbleFecha:'', eqrvSort:'sector' };

// Pesos por ventana (Chart 1) — default PESOS_Z = [0.33, 0.33, 0.33, 0.00]
let W = { '1Y': 0.33, '3Y': 0.33, '5Y': 0.33, '10Y': 0.00 };
const hiddenSec = new Set();

// ── LAYOUT BASE ──────────────────────────────────────────────────────────────
const LB = {
  paper_bgcolor:'#FFFFFF', plot_bgcolor:'#FFFFFF',
  font:{ family:"'Calibri','Segoe UI',sans-serif", size:11.5, color:'#6B7A99' },
  margin:{ t:20, r:20, b:42, l:58 },
  xaxis:{ showgrid:false, showline:true, linecolor:'#D0D7E5',
          tickcolor:'#D0D7E5', zeroline:false, tickfont:{color:'#6B7A99'} },
  yaxis:{ showgrid:false, showline:true, linecolor:'#D0D7E5',
          tickcolor:'#D0D7E5', zeroline:false, tickfont:{color:'#6B7A99'} },
  hoverlabel:{ bgcolor:'#1E2E6E', bordercolor:'#00AECB',
    font:{ family:"'IBM Plex Mono',monospace", size:11, color:'#FFFFFF' } },
  showlegend:false,
};
const CFG = { displayModeBar:true, modeBarButtonsToRemove:['select2d','lasso2d','autoScale2d'],
              displaylogo:false, responsive:true };

const C = { navy:'#1E2E6E', teal:'#00AECB', yellow:'#E3E829',
            bg:'#FFFFFF', surf:'#F5F7FA', border:'#D0D7E5',
            text:'#1a2340', muted:'#6B7A99',
            green:'#1a8a4a', red:'#c0392b', amber:'#b37a00' };

// ── HELPERS ──────────────────────────────────────────────────────────────────
function scrollTo(id) {
  document.getElementById(id)?.scrollIntoView({ behavior:'smooth', block:'start' });
  document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
  event.target.classList.add('active');
}
function tog(id, val, opts) {
  opts.forEach(o => document.getElementById(`btn-${id}-${o}`).classList.toggle('active', o===val));
}
function setProgress(pct) {
  document.getElementById('lbar').style.width = pct + '%';
}

// ── INIT ─────────────────────────────────────────────────────────────────────
window.addEventListener('load', () => {
  document.getElementById('hdr-date').textContent = 'Actualizado: ' + GEN_DATE;
  const logo = document.getElementById('logo-img');
  if (!logo.src || logo.src.endsWith('__LOGO_B64__') || logo.src.length < 50) {
    logo.style.display = 'none';
    document.getElementById('logo-ph').style.display = 'flex';
  }

  setProgress(20);
  populateEmpresaSelector();
  setProgress(35);
  initBubble();
  setProgress(50);
  populateMultiplosSelectors();
  setProgress(65);
  renderHeatmap();
  setProgress(80);
  renderEqrv();
  setProgress(92);

  resetWeights();  // initialises W, sum display and subtitle
  populateFanSelector();
  if (HIST_DATA.length) {
    E.empresa = HIST_DATA[0].alias;
    document.getElementById('sel-empresa').value = E.empresa;
    renderHistorico();
    renderBubble();
  }
  setProgress(100);
  setTimeout(() => { document.getElementById('loading').style.display = 'none'; }, 200);
});

// ── 01 HISTÓRICO ─────────────────────────────────────────────────────────────
function populateEmpresaSelector() {
  const sel = document.getElementById('sel-empresa');
  HIST_DATA.forEach(d => {
    const o = document.createElement('option');
    o.value = d.alias; o.textContent = `${d.alias}  —  ${d.sector}`;
    sel.appendChild(o);
  });
}
function onEmpresa() { E.empresa = document.getElementById('sel-empresa').value; renderHistorico(); }
function onZType(t) { E.ztype=t; tog('z',t,['abs','basket','bvl']); renderHistorico(); }

// ── Weight controls ────────────────────────────────────────────────────────
function onWeightChange() {
  W['1Y']  = parseFloat(document.getElementById('w1y').value)  || 0;
  W['3Y']  = parseFloat(document.getElementById('w3y').value)  || 0;
  W['5Y']  = parseFloat(document.getElementById('w5y').value)  || 0;
  W['10Y'] = parseFloat(document.getElementById('w10y').value) || 0;
  const s  = W['1Y'] + W['3Y'] + W['5Y'] + W['10Y'];
  const sumEl  = document.getElementById('weight-sum');
  const warnEl = document.getElementById('weight-warn');
  sumEl.textContent = 'Σ = ' + s.toFixed(2);
  const ok = Math.abs(s - 1.0) < 0.01;
  sumEl.style.color  = ok ? 'var(--navy)' : '#c0392b';
  warnEl.style.display = ok ? 'none' : 'inline';
  // Update subtitle dynamically
  const desc = document.getElementById('s1-desc');
  if (desc) {
    const parts = [
      W['1Y']  > 0 ? `1Y×${(W['1Y']*100).toFixed(0)}%`  : null,
      W['3Y']  > 0 ? `3Y×${(W['3Y']*100).toFixed(0)}%`  : null,
      W['5Y']  > 0 ? `5Y×${(W['5Y']*100).toFixed(0)}%`  : null,
      W['10Y'] > 0 ? `10Y×${(W['10Y']*100).toFixed(0)}%` : null,
    ].filter(Boolean);
    desc.textContent = 'Z compuesto · ' + (parts.length ? parts.join(' + ') : 'sin pesos');
  }
  renderHistorico();
}

function resetWeights() {
  document.getElementById('w1y').value  = '0.33';
  document.getElementById('w3y').value  = '0.33';
  document.getElementById('w5y').value  = '0.33';
  document.getElementById('w10y').value = '0.00';
  onWeightChange();
}

function _computeWeightedZ(d) {
  /**
   * Recalcula el Z compuesto ponderado en el cliente usando z_windows.
   *
   * z_windows[win] ya contiene el Z de ventana N = sum(w_mult_i * Z_mult_i_N),
   * es decir la contribución ponderada de todos los múltiplos para esa ventana.
   *
   * El Z final con pesos de usuario [p1,p3,p5,p10] es:
   *   Z_final(fecha) = p1*Z_1Y(fecha) + p3*Z_3Y(fecha) + p5*Z_5Y(fecha) + p10*Z_10Y(fecha)
   *
   * Si los pesos son los por defecto Y hay z_windows disponibles, usa z_windows
   * igualmente para consistencia (la curva puede diferir levemente de BD por
   * usar datos bi-semanales en lugar de diarios, pero es la misma lógica).
   * Si no hay z_windows (empresa sin múltiplos en CSV), cae en d.x/d.y de BD.
   */
  if (!d.z_windows || !Object.keys(d.z_windows).length) {
    return { x: d.x, y: d.y };  // fallback: serie pre-calculada de BD
  }

  // Union de todas las fechas de las ventanas activas
  const dateMap = {};
  for (const [win, wd] of Object.entries(d.z_windows)) {
    const p = W[win] || 0;
    if (p === 0 || !wd || !wd.x) continue;
    wd.x.forEach((date, i) => {
      if (!(date in dateMap)) dateMap[date] = 0;
      if (wd.y[i] !== null && !isNaN(wd.y[i])) {
        dateMap[date] += p * wd.y[i];   // Z_final = sum(p_i * Z_i)
      }
    });
  }

  const dates = Object.keys(dateMap).sort();
  return {
    x: dates,
    y: dates.map(date => Math.round(dateMap[date] * 10000) / 10000),
  };
}

// ── 01 HISTÓRICO ─────────────────────────────────────────────────────────────
function renderHistorico() {
  const d = HIST_DATA.find(x => x.alias === E.empresa);
  if (!d) return;
  let x, y, yLbl;
  if (E.ztype==='basket' && d.x_basket.length) { x=d.x_basket; y=d.y_basket; yLbl='Z vs Basket'; }
  else if (E.ztype==='bvl' && d.x_bvl.length)  { x=d.x_bvl;    y=d.y_bvl;    yLbl='Z vs BVL'; }
  else                                          { x=d.x;        y=d.y;        yLbl='Z Absoluto'; }
  const lc = (d.z_actual||0) > 0 ? C.red : C.green;
  const shapes=[
    {type:'line',x0:x[0],x1:x[x.length-1],y0:1.5,y1:1.5,line:{color:C.red,width:.9,dash:'dot'}},
    {type:'line',x0:x[0],x1:x[x.length-1],y0:-1.5,y1:-1.5,line:{color:C.green,width:.9,dash:'dot'}},
    {type:'line',x0:x[0],x1:x[x.length-1],y0:.5,y1:.5,line:{color:C.amber,width:.6,dash:'dot'}},
    {type:'line',x0:x[0],x1:x[x.length-1],y0:-.5,y1:-.5,line:{color:C.amber,width:.6,dash:'dot'}},
    {type:'line',x0:x[0],x1:x[x.length-1],y0:0,y1:0,line:{color:C.muted,width:.9}},
  ];
  const anns=[
    {x:x[x.length-1],y:1.5,text:'+1.5σ',showarrow:false,font:{size:9,color:C.red},xanchor:'right',yanchor:'bottom'},
    {x:x[x.length-1],y:-1.5,text:'−1.5σ',showarrow:false,font:{size:9,color:C.green},xanchor:'right',yanchor:'top'},
  ];
  Plotly.react('plot-hist',[
    {x,y,type:'scatter',mode:'none',fill:'tozeroy',fillcolor:lc+'18',hoverinfo:'skip'},
    {x,y,type:'scatter',mode:'lines',line:{color:lc,width:1.8},
     hovertemplate:'%{x}<br><b>Z = %{y:.3f}</b><extra></extra>'},
  ], {...LB, shapes, annotations:anns,
    yaxis:{...LB.yaxis,title:{text:yLbl,font:{size:10}}}, margin:{t:16,r:20,b:36,l:56}}, CFG);
  updateInterp(d);
}

function updateInterp(d) {
  const ip = INTERP_DATA[d.alias]; if (!ip) return;
  const b  = document.getElementById('interp-badge');
  b.textContent = ip.señal;
  b.style.background = ip.color + '1a';
  b.style.color = ip.color;
  b.style.border = `1px solid ${ip.color}40`;
  document.getElementById('interp-short').textContent = ip.corto;
  document.getElementById('interp-long').textContent  = ip.largo;
  const p = ip.percentil!=null ? ` · Percentil: ${ip.percentil.toFixed(0)}°` : '';
  document.getElementById('interp-meta').textContent = `${d.empresa} · ${d.sector}${p}`;
}

// ── 02 BUBBLE ────────────────────────────────────────────────────────────────
function initBubble() {
  const fechas = Object.keys(BUBBLE_DATA).sort();
  E.bubbleFecha = fechas[fechas.length-1];
  const inp = document.getElementById('sel-fecha-bubble');
  inp.value = E.bubbleFecha; inp.min = fechas[0]; inp.max = fechas[fechas.length-1];
  buildBubbleLegend();
}
function onBubbleFecha() {
  const fechas = Object.keys(BUBBLE_DATA).sort(); let best = fechas[0];
  const v = document.getElementById('sel-fecha-bubble').value;
  for (const f of fechas) { if (f <= v) best = f; }
  E.bubbleFecha = best; renderBubble();
}
function onYAxis(a) { E.yaxis=a; tog('y',a,['basket','bvl']); renderBubble(); }

function buildBubbleLegend() {
  const secs = {};
  Object.values(BUBBLE_DATA).forEach(pts => pts.forEach(p => { secs[p.sector]=p.color; }));
  const row = document.getElementById('bubble-legend'); row.innerHTML='';
  Object.entries(secs).forEach(([sec,col]) => {
    row.innerHTML+=`<div class="leg-item" onclick="toggleSec('${sec}')">
      <div class="leg-dot" style="background:${col}"></div><span>${sec}</span></div>`;
  });
}
function toggleSec(s) { hiddenSec.has(s)?hiddenSec.delete(s):hiddenSec.add(s); renderBubble(); }

function renderBubble() {
  const allPts = (BUBBLE_DATA[E.bubbleFecha]||[]).filter(p=>!hiddenSec.has(p.sector));
  if (!allPts.length) { Plotly.react('plot-bubble',[],{...LB},CFG); return; }

  const yKey    = E.yaxis==='basket' ? 'z_basket' : 'z_bvl';
  const yLabel  = E.yaxis==='basket' ? 'Z vs Basket (Stock/Basket)' : 'Z vs BVL (Stock/Índice)';
  const xLabel  = 'Valoración relativa Stock (Z-score)';

  // Filter to points with valid Y value
  const pts = allPts.filter(p => p[yKey] !== null && p[yKey] !== undefined);

  // ── Calibrar sizeref para que la burbuja máxima tenga ~55px ──────────
  // sizeref = 2 * maxSize / (maxPx²)  con sizemode='area'
  const sizes   = pts.map(p => p.size > 0 ? p.size : 0);
  const maxSize = Math.max(...sizes, 1);
  const maxPx   = 55;
  const sizeref = 2 * maxSize / (maxPx * maxPx);
  const minBubble = maxSize * 0.004;  // tamaño mínimo para índices/sin posición

  // ── Una sola traza 'scatter' con marcadores diferenciados por color ───
  // (mejor rendimiento y hover limpio que múltiples trazas por sector)
  const SECTOR_COLORS_HEX = {
    'Financials':'#2166ac','Mining':'#d6604d','Utilities':'#4dac26',
    'Consumer':'#7b3294','Construction Materials':'#e08214',
    'Industrials':'#4eb3d3','Health Care':'#e31a1c','Media':'#984ea3',
    'Index':'#aaaaaa','Other':'#aaaaaa'
  };

  // Separar empresas con posición (size>0) de índices/referencias (size=0)
  const ptsPos = pts.filter(p => p.size > 0);
  const ptsRef = pts.filter(p => p.size === 0);

  const mkTrace = (points, name, opacity, showLegend) => ({
    type: 'scatter',
    mode: 'markers+text',
    name: name,
    x: points.map(p => +p.z_abs),      // +p fuerza conversión a number
    y: points.map(p => +p[yKey]),
    text: points.map(p => p.alias),
    textposition: 'top center',
    textfont: { size: 8.5, color: '#333333', family: "'Calibri','Segoe UI',sans-serif" },
    cliponaxis: false,
    marker: {
      size:    points.map(p => Math.max(minBubble, p.size > 0 ? p.size : minBubble*0.3)),
      sizeref: sizeref,
      sizemode: 'area',
      color: points.map(p => SECTOR_COLORS_HEX[p.sector] || '#aaaaaa'),
      opacity: opacity,
      line: { color: points.map(p => SECTOR_COLORS_HEX[p.sector] || '#aaaaaa'),
              width: 1.2 },
    },
    hovertemplate:
      '<b>%{text}</b><br>' +
      'Z Stock: %{x:.2f}x<br>' +
      yLabel.split('(')[0].trim() + ': %{y:.2f}x<br>' +
      '<i>Exposición: ' + points.map(p => p.size > 0 ? 'S/. '+p.size.toFixed(0)+'M' : 'referencia').join('|') + '</i>' +
      '<extra>%{meta}</extra>',
    meta: points.map(p => p.sector),
    showlegend: false,
  });

  const traces = [];
  if (ptsPos.length) traces.push(mkTrace(ptsPos, 'Portafolio', 0.72, true));
  if (ptsRef.length) traces.push(mkTrace(ptsRef, 'Referencia', 0.45, false));

  // ── Leyenda manual de sectores como scatter dummy ────────────────────
  const sectoresPts = [...new Set(pts.map(p=>p.sector))];
  sectoresPts.forEach(sec => {
    traces.push({
      type:'scatter', mode:'markers', name: sec,
      x:[null], y:[null],
      marker:{ size:10, color: SECTOR_COLORS_HEX[sec]||'#aaa',
               opacity:0.8, line:{color:SECTOR_COLORS_HEX[sec]||'#aaa',width:1} },
      showlegend: !hiddenSec.has(sec),
    });
  });

  // ── Rango de ejes (automático + padding) ─────────────────────────────
  const xs = pts.map(p=>+p.z_abs),  ys = pts.map(p=>+p[yKey]);
  const xmin=Math.min(...xs), xmax=Math.max(...xs);
  const ymin=Math.min(...ys), ymax=Math.max(...ys);
  const xpad = Math.max((xmax-xmin)*0.18, 0.5);
  const ypad = Math.max((ymax-ymin)*0.22, 0.3);
  const X0=xmin-xpad, X1=xmax+xpad, Y0=ymin-ypad, Y1=ymax+ypad;

  // ── Cuadrantes y líneas de referencia ────────────────────────────────
  const shapes = [
    // Fondo cuadrantes (muy sutil — estilo equity research)
    {type:'rect',x0:0,x1:X1,y0:0,y1:Y1,fillcolor:'rgba(192,57,43,0.04)',line:{width:0},layer:'below'},
    {type:'rect',x0:X0,x1:0,y0:Y0,y1:0,fillcolor:'rgba(26,138,74,0.04)',line:{width:0},layer:'below'},
    {type:'rect',x0:X0,x1:0,y0:0,y1:Y1,fillcolor:'rgba(180,140,0,0.03)',line:{width:0},layer:'below'},
    {type:'rect',x0:0,x1:X1,y0:Y0,y1:0,fillcolor:'rgba(180,140,0,0.03)',line:{width:0},layer:'below'},
    // Ejes cruzados principales
    {type:'line',x0:X0,x1:X1,y0:0,y1:0,line:{color:'#555555',width:1.0},layer:'below'},
    {type:'line',x0:0,x1:0,y0:Y0,y1:Y1,line:{color:'#555555',width:1.0},layer:'below'},
    // Líneas ±1.5σ (referencia estadística)
    {type:'line',x0:X0,x1:X1,y0:1.5,y1:1.5,line:{color:'#c0392b',width:0.6,dash:'dash'},layer:'below'},
    {type:'line',x0:X0,x1:X1,y0:-1.5,y1:-1.5,line:{color:'#1a8a4a',width:0.6,dash:'dash'},layer:'below'},
    {type:'line',x0:1.5,x1:1.5,y0:Y0,y1:Y1,line:{color:'#c0392b',width:0.6,dash:'dash'},layer:'below'},
    {type:'line',x0:-1.5,x1:-1.5,y0:Y0,y1:Y1,line:{color:'#1a8a4a',width:0.6,dash:'dash'},layer:'below'},
  ];

  // ── Etiquetas de cuadrante (estilo PDF de referencia) ────────────────
  const qFont = {size:9.5, color:'#999999', family:"'Calibri','Segoe UI',sans-serif"};
  const annotations = [
    {x:X1-0.05,y:Y1-0.05,xanchor:'right',yanchor:'top',
     text:'Cara / Cara',showarrow:false,font:{...qFont,color:'#c0392b'},xref:'x',yref:'y'},
    {x:X0+0.05,y:Y0+0.05,xanchor:'left',yanchor:'bottom',
     text:'Barata / Barata',showarrow:false,font:{...qFont,color:'#1a8a4a'},xref:'x',yref:'y'},
    {x:X0+0.05,y:Y1-0.05,xanchor:'left',yanchor:'top',
     text:'Barata / Cara',showarrow:false,font:qFont,xref:'x',yref:'y'},
    {x:X1-0.05,y:Y0+0.05,xanchor:'right',yanchor:'bottom',
     text:'Cara / Barata',showarrow:false,font:qFont,xref:'x',yref:'y'},
    // Etiquetas de referencia ±1.5σ
    {x:1.52,y:Y0+0.05,xanchor:'left',yanchor:'bottom',text:'+1.5σ',
     showarrow:false,font:{size:7.5,color:'#c0392b'},xref:'x',yref:'y'},
    {x:1.52,y:Y0+0.05,xanchor:'left',yanchor:'bottom',text:'',
     showarrow:false,xref:'x',yref:'y'},
  ];

  const layout = {
    ...LB,
    paper_bgcolor: '#FFFFFF',
    plot_bgcolor:  '#FFFFFF',
    showlegend: true,
    legend: {
      orientation: 'h', x: 0, y: -0.13, xanchor: 'left',
      bgcolor: '#FFFFFF', bordercolor: '#D0D7E5', borderwidth: 1,
      font: { size: 10, color: '#1a2340', family: "'Calibri','Segoe UI',sans-serif" },
      itemsizing: 'constant',
    },
    xaxis: {
      ...LB.xaxis,
      title: { text: xLabel, font: { size: 10.5, color:'#444' } },
      range: [X0, X1],
      zeroline: false,
      showgrid: false, showline: true,
      linecolor: '#cccccc', tickcolor: '#cccccc',
      tickformat: '.1f', ticksuffix: 'x',
      tickfont: { size: 9, color: '#666' },
    },
    yaxis: {
      ...LB.yaxis,
      title: { text: yLabel, font: { size: 10.5, color:'#444' } },
      range: [Y0, Y1],
      zeroline: false,
      showgrid: false, showline: true,
      linecolor: '#cccccc', tickcolor: '#cccccc',
      tickformat: '.1f', ticksuffix: 'x',
      tickfont: { size: 9, color: '#666' },
    },
    shapes,
    annotations,
    margin: { t: 36, r: 24, b: 80, l: 68 },
    title: {
      text: `<b>Valuation Relativo Acciones RVL</b>   <span style="font-weight:normal;font-size:11px;color:#888">Z-scores: PE, PB, PS, EV/EBITDA · ${E.bubbleFecha}</span>`,
      font: { size: 13, color: '#1E2E6E', family: "'Calibri','Segoe UI',sans-serif" },
      x: 0.01, xanchor: 'left',
    },
    hoverlabel: {
      bgcolor: '#1E2E6E', bordercolor: '#00AECB',
      font: { family: "'IBM Plex Mono',monospace", size: 10.5, color: '#FFFFFF' },
      align: 'left',
    },
  };

  Plotly.react('plot-bubble', traces, layout, CFG);
}

// ── 03 HEATMAP ───────────────────────────────────────────────────────────────
function renderHeatmap() {
  const d = HEAT_DATA;
  if (!d||!d.z||!d.z.length) return;
  const zc = d.z.map(row=>row.map(v=>v===null?NaN:v));
  const yL = d.y.map((a,i)=>`${a} · ${d.sectores[i]}`);
  Plotly.react('plot-heat',[{
    type:'heatmap',x:d.x,y:yL,z:zc,
    colorscale:[[0,'#145228'],[.3,C.green],[.45,'#2d4f3a'],[.49,C.muted],
                [.51,C.muted],[.55,'#5c2323'],[.7,C.red],[1,'#6b1515']],
    zmid:0,zmin:-2.5,zmax:2.5,
    colorbar:{title:{text:'Z',font:{size:10}},thickness:10,len:.65,tickfont:{size:9},
              bgcolor:'#FFFFFF',bordercolor:C.border,
              tickvals:[-2,-1,0,1,2],ticktext:['−2','−1','0','+1','+2']},
    hovertemplate:'<b>%{y}</b><br>Semana: %{x}<br><b>Z = %{z:.3f}</b><extra></extra>',
    xgap:1,ygap:1,
  }],{...LB,height:Math.max(420,d.y.length*22+70),
    margin:{t:16,r:90,b:60,l:140},
    xaxis:{...LB.xaxis,type:'category',nticks:12,tickangle:-35,tickfont:{size:9}},
    yaxis:{...LB.yaxis,tickfont:{size:9.5,color:C.text},autorange:'reversed'}},CFG);
}

// ── 04 EQRV ──────────────────────────────────────────────────────────────────
function onEqrvSort(m) {
  E.eqrvSort=m;
  ['sector','z','pct'].forEach(x=>document.getElementById('btn-sort-'+x).classList.toggle('active',x===m));
  renderEqrv();
}
function renderEqrv() {
  if (!EQRV_DATA||!EQRV_DATA.length) return;
  const SO=['Financials','Consumer','Construction Materials','Industrials',
            'Health Care','Media','Mining','Utilities','Other'];
  let data=[...EQRV_DATA];
  if (E.eqrvSort==='z')   data.sort((a,b)=>b.current-a.current);
  else if (E.eqrvSort==='pct') data.sort((a,b)=>b.pct_rank-a.pct_rank);
  else data.sort((a,b)=>{
    const si=SO.indexOf(a.sector),sj=SO.indexOf(b.sector);
    return si!==sj?(si<0?99:si)-(sj<0?99:sj):b.current-a.current;
  });
  const aliases=data.map(d=>d.alias), n=aliases.length;
  const dotC=data.map(d=>d.current>1.5?C.red:d.current<-1.5?C.green:d.current>0.5?C.amber:d.current<-0.5?C.teal:C.muted);
  const shapes=[
    {type:'line',x0:0,x1:0,y0:-.5,y1:n-.5,xref:'x',yref:'y',line:{color:C.muted,width:.9}},
    {type:'line',x0:1.5,x1:1.5,y0:-.5,y1:n-.5,xref:'x',yref:'y',line:{color:C.red,width:.7,dash:'dot'}},
    {type:'line',x0:-1.5,x1:-1.5,y0:-.5,y1:n-.5,xref:'x',yref:'y',line:{color:C.green,width:.7,dash:'dot'}},
  ];
  let ps=0;
  for(let i=0;i<=n;i++){
    if(i===n||data[i].sector!==data[ps].sector){
      if(i>0) shapes.push({type:'line',x0:-4,x1:7,y0:i-.5,y1:i-.5,xref:'x',yref:'y',line:{color:C.border,width:.8}});
      if(i<n) ps=i;
    }
  }
  const secAnns=[];
  let s2=0;
  for(let i=0;i<=n;i++){
    if(i===n||data[i].sector!==data[s2].sector){
      secAnns.push({x:7.1,y:(s2+(i-1))/2,xref:'x',yref:'y',
        text:data[s2].sector,showarrow:false,xanchor:'left',align:'left',
        font:{size:8.5,color:data[s2].color||C.muted}});
      if(i<n) s2=i;
    }
  }
  Plotly.react('plot-eqrv',[
    {type:'bar',orientation:'h',x:data.map(d=>d.p90-d.p10),y:aliases,base:data.map(d=>d.p10),
     marker:{color:'rgba(0,174,203,0.10)',line:{width:0}},width:.55,hoverinfo:'skip',
     showlegend:true,name:'P10–P90'},
    {type:'bar',orientation:'h',x:data.map(d=>d.p75-d.p25),y:aliases,base:data.map(d=>d.p25),
     marker:{color:'rgba(30,46,110,0.50)',line:{width:0}},width:.44,hoverinfo:'skip',
     showlegend:true,name:'P25–P75'},
    {type:'scatter',mode:'markers',x:data.map(d=>d.p50),y:aliases,
     marker:{symbol:'line-ns',size:16,color:C.teal,line:{color:C.teal,width:2}},
     hovertemplate:'<b>%{y}</b><br>Mediana: %{x:.3f}<extra></extra>',showlegend:true,name:'Mediana'},
    {type:'scatter',mode:'markers+text',x:data.map(d=>d.current),y:aliases,
     text:data.map(d=>`${d.current>0?'+':''}${d.current.toFixed(2)}`),
     textposition:data.map(d=>d.current>=0?'middle right':'middle left'),
     textfont:{size:9.5,color:dotC},
     marker:{symbol:'diamond',size:10,color:dotC,line:{color:C.bg,width:1.2}},
     hovertemplate:'<b>%{y}</b><br>Z actual: %{x:.3f}<extra></extra>',showlegend:true,name:'Z actual'},
  ],{...LB,barmode:'overlay',height:Math.max(380,n*26+80),
    showlegend:true,
    legend:{bgcolor:'#FFFFFF',bordercolor:C.border,borderwidth:1,font:{size:10,color:C.text},orientation:'h',x:0,y:1.04,xanchor:'left'},
    xaxis:{...LB.xaxis,title:{text:'Z-Score',font:{size:10}},range:[-4,8],zeroline:false,
           tickvals:[-3,-2,-1,0,1,2,3,4,5],ticktext:['−3','−2','−1','0','+1','+2','+3','+4','+5']},
    yaxis:{...LB.yaxis,autorange:'reversed',tickfont:{size:10},ticklen:3},
    shapes,annotations:[...secAnns,
      {x:1.5,y:-.9,xref:'x',yref:'y',text:'+1.5σ',showarrow:false,font:{size:8,color:'#c0392b'},yanchor:'top'},
      {x:-1.5,y:-.9,xref:'x',yref:'y',text:'−1.5σ',showarrow:false,font:{size:8,color:'#1a8a4a'},yanchor:'top'}],
    margin:{t:50,r:95,b:50,l:75}},CFG);
}

// ── 05 MÚLTIPLOS ─────────────────────────────────────────────────────────────
function populateMultiplosSelectors() {
  const tickers = Object.keys(MULT_DATA).sort();
  ['sel-m1','sel-m2','sel-m3'].forEach((id, idx) => {
    const sel = document.getElementById(id);
    if (idx > 0) sel.innerHTML = '<option value="">— ninguno —</option>';
    tickers.forEach(t => {
      const o = document.createElement('option');
      o.value = t; o.textContent = t;
      sel.appendChild(o);
    });
    if (idx === 0 && tickers.length) sel.value = tickers[0];
  });
}

// Line colors for up to 3 equities
const M_COLORS = [C.teal, C.yellow, '#bc8cff'];
const M_DASH   = ['solid','dot','dashdot'];

function renderMultiplos() {
  const multiplo = document.getElementById('sel-multiplo').value;
  const tickers  = [
    document.getElementById('sel-m1').value,
    document.getElementById('sel-m2').value,
    document.getElementById('sel-m3').value,
  ].filter(Boolean);

  const traces = [];
  tickers.forEach((t, i) => {
    const td = MULT_DATA[t];
    if (!td) return;
    const md = td[multiplo];
    if (!md || !md.x.length) return;
    traces.push({
      x: md.x, y: md.y,
      type: 'scatter', mode: 'lines',
      name: t,
      line: { color: M_COLORS[i], width: 1.8, dash: M_DASH[i] },
      hovertemplate: `<b>${t}</b><br>%{x}<br>${MULT_LABELS[multiplo]}: %{y:.2f}x<extra></extra>`,
    });
  });

  const label = MULT_LABELS[multiplo] || multiplo;
  Plotly.react('plot-mult', traces, {
    ...LB, showlegend: traces.length > 1,
    legend: { bgcolor: '#FFFFFF', bordercolor: C.border, borderwidth: 1, font: { size: 10, color: C.text } },
    yaxis: { ...LB.yaxis, title: { text: label, font: { size: 10 } } },
    margin: { t: 16, r: 20, b: 36, l: 58 },
    shapes: [{type:'line',x0:traces[0]?.x[0]||'',x1:traces[0]?.x.slice(-1)[0]||'',
               y0:0,y1:0,line:{color:C.muted,width:.7}}],
  }, CFG);
}

// ════════════════════════════════════════════════════════════════════════════
// CHART 09 — CALENDARIZED RETURN FAN
// ════════════════════════════════════════════════════════════════════════════
function populateFanSelector() {
  const sel = document.getElementById('sel-fan');
  Object.keys(CALENDAR_DATA).sort().forEach(alias => {
    const o = document.createElement('option');
    o.value = alias; o.textContent = alias;
    sel.appendChild(o);
  });
  if (sel.options.length) {
    sel.value = sel.options[0].value;
    renderFan();
  }
}

function renderFan() {
  const alias = document.getElementById('sel-fan').value;
  const d = CALENDAR_DATA[alias];
  if (!d) return;

  const b = d.bands;
  const traces = [];

  // P10–P90 band (outermost, very faint)
  traces.push({
    type:'scatter', mode:'lines', name:'P10–P90',
    x: [...b.doy, ...b.doy.slice().reverse()],
    y: [...b.p90, ...b.p10.slice().reverse()],
    fill:'toself', fillcolor:'rgba(30,46,110,0.07)',
    line:{color:'transparent'}, hoverinfo:'skip', showlegend:true,
  });

  // P25–P75 box (tighter, slightly more opaque)
  traces.push({
    type:'scatter', mode:'lines', name:'P25–P75',
    x: [...b.doy, ...b.doy.slice().reverse()],
    y: [...b.p75, ...b.p25.slice().reverse()],
    fill:'toself', fillcolor:'rgba(30,46,110,0.14)',
    line:{color:'transparent'}, hoverinfo:'skip', showlegend:true,
  });

  // Median line
  traces.push({
    type:'scatter', mode:'lines', name:'Mediana histórica',
    x: b.doy, y: b.median,
    line:{ color: C.navy, width:1.4, dash:'dot' },
    hovertemplate:'Día %{x}<br>Mediana: %{y:.1f}%<extra></extra>',
  });

  // Previous year (reference)
  if (d.ytd_prev_year && d.ytd_prev_year.doy && d.ytd_prev_year.doy.length) {
    const py = parseInt(d.current_year) - 1;
    traces.push({
      type:'scatter', mode:'lines', name:`${py}`,
      x: d.ytd_prev_year.doy, y: d.ytd_prev_year.ytd,
      line:{ color: C.muted, width:1.2, dash:'dashdot' },
      hovertemplate:`Día %{x}<br>${py}: %{y:.1f}%<extra></extra>`,
      opacity: 0.7,
    });
  }

  // Current year (hero line — bold, colored by performance)
  if (d.ytd_current && d.ytd_current.doy && d.ytd_current.doy.length) {
    const cur = d.ytd_current;
    const lastRet = cur.ytd[cur.ytd.length-1];
    const heroColor = lastRet >= 0 ? '#1a8a4a' : '#c0392b';
    traces.push({
      type:'scatter', mode:'lines', name:`${d.current_year} (actual)`,
      x: cur.doy, y: cur.ytd,
      line:{ color: heroColor, width:2.6 },
      hovertemplate:`Día %{x}<br>${d.current_year}: %{y:.1f}%<extra></extra>`,
    });

    // Current point dot
    traces.push({
      type:'scatter', mode:'markers', name:'Hoy',
      x:[cur.doy[cur.doy.length-1]], y:[lastRet],
      marker:{ size:9, color: heroColor, line:{color:'white',width:1.5} },
      hovertemplate:`Hoy (día ${cur.doy[cur.doy.length-1]})<br>YTD ${d.current_year}: ${lastRet.toFixed(1)}%<extra></extra>`,
    });
  }

  // Compute where current year stands vs history
  let pctRankText = '';
  if (d.ytd_current && d.ytd_current.ytd && d.ytd_current.ytd.length) {
    const curRet = d.ytd_current.ytd[d.ytd_current.ytd.length-1];
    const curDoy = d.ytd_current.doy[d.ytd_current.doy.length-1];
    // Get historical returns at the same DOY
    const histAtDoy = [];
    d.hist_years.forEach(yr => {
      const yd = CALENDAR_DATA[alias];
      // We only have bands, not individual years — use bands
    });
    const p10_now   = b.p10[Math.min(curDoy, b.doy.length-1)];
    const p90_now   = b.p90[Math.min(curDoy, b.doy.length-1)];
    const med_now   = b.median[Math.min(curDoy, b.doy.length-1)];
    pctRankText = `YTD actual: ${curRet > 0 ? '+' : ''}${curRet.toFixed(1)}% · Mediana histórica a esta fecha: ${med_now > 0 ? '+' : ''}${med_now.toFixed(1)}% · Banda P10–P90: ${p10_now.toFixed(1)}% a ${p90_now.toFixed(1)}%`;
  }

  Plotly.react('plot-fan', traces, {
    ...LB, showlegend:true,
    legend:{ bgcolor:'#fff',bordercolor:C.border,borderwidth:1,font:{size:10},orientation:'h',x:0,y:1.04,xanchor:'left' },
    xaxis:{ ...LB.xaxis,
      title:{text:'Día del año (0 = 1 enero)',font:{size:10}},
      range:[0,365], tickvals:[0,31,59,90,120,151,181,212,243,273,304,334],
      ticktext:['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'],
    },
    yaxis:{ ...LB.yaxis,
      title:{text:'Retorno YTD acumulado (%)',font:{size:10}},
      ticksuffix:'%', zeroline:false,
    },
    shapes:[{ type:'line',x0:0,x1:365,y0:0,y1:0,line:{color:'#555',width:.8} }],
    margin:{ t:44, r:24, b:60, l:64 },
    title:{ text:`<b>${alias}</b> — Retorno YTD vs distribución histórica (${d.hist_years.length} años)${pctRankText ? '<br><sub>'+pctRankText+'</sub>' : ''}`,
            font:{size:12,color:C.navy},x:.01,xanchor:'left' },
  }, CFG);
}

</script>
</body>
</html>
"""


def exportar_exposicion_excel(exposiciones: dict, param: pd.DataFrame, bd: pd.DataFrame) -> None:
    """
    Exporta "QUEST - exposicion de cada activo.xlsx" con:
      - Una fila por equity en Parametría
      - Columnas: Alias | BBG | SBS | FechaExposicion | Fondo1_PxQ | Fondo2_PxQ | Fondo3_PxQ | Total_PEN | Total_MM_PEN
    Permite validar que los tamaños de burbuja del Chart 2 son correctos.

    Usa la fecha de exposición más reciente disponible en DQ_IN.
    Si un equity no tiene SBS o no aparece en DQ_IN, registra NaN en las columnas de exposición.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ openpyxl no disponible — Excel de exposición no generado.")
        return

    if not exposiciones:
        print("  ⚠ Sin datos de exposición — Excel de exposición no generado.")
        return

    # ── Cargar DQ_IN en bruto para extraer filas individuales por fondo ─────
    df_dq, _ = _load_dq_in()
    pxq_all  = df_dq[df_dq["Dato"].astype(str) == "PxQ"].copy()
    pxq_all["Fecha"] = pd.to_datetime(pxq_all["Fecha"], errors="coerce")
    pxq_all  = pxq_all.dropna(subset=["Fecha"])

    # Fecha de exposición más reciente
    last_fecha_str = max(exposiciones.keys())
    last_fecha_ts  = pd.Timestamp(last_fecha_str)

    # Filas PxQ en la fecha más cercana ≤ last_fecha
    fechas_dq = pxq_all["Fecha"].dt.normalize().unique()
    fechas_dq = sorted([f for f in fechas_dq if f <= last_fecha_ts])
    if not fechas_dq:
        print("  ⚠ Sin filas PxQ disponibles — Excel no generado.")
        return
    fecha_ref = fechas_dq[-1]
    pxq_fecha = pxq_all[pxq_all["Fecha"].dt.normalize() == fecha_ref].copy()

    # Nombre de cada fondo para las columnas
    fondos = pxq_fecha["Fondo"].astype(str).tolist()   # e.g. ["IN-FONDO1","IN-FONDO2","IN-FONDO3"]
    sbs_cols = [c for c in df_dq.columns
                if c not in ("column_0","Fondo","Dato","Fecha","SBS","Fecha_dt","")
                and not str(c).startswith("col_")
                and str(c) not in ("nan","None","TEST")]

    # ── Construir tabla de resultados ────────────────────────────────────────
    rows = []
    for _, r in param.iterrows():
        alias = r["alias"]
        bbg   = r["bbg_ticker"]
        sbs   = str(r.get("sbs", r.get("sbs_code", ""))).strip()
        if sbs in ("nan","None",""): sbs = ""
        z_col = r["z_col"]

        # Fondo exposures (una columna por fondo)
        fondo_vals = {}
        total_pen  = 0.0
        for fondo_name in fondos:
            fondo_row = pxq_fecha[pxq_fecha["Fondo"].astype(str) == fondo_name]
            val = np.nan
            if not fondo_row.empty and sbs and sbs in pxq_fecha.columns:
                raw = fondo_row[sbs].values[0] if len(fondo_row) > 0 else np.nan
                val = pd.to_numeric(raw, errors="coerce")
                val = float(val) if pd.notna(val) else np.nan
            fondo_vals[fondo_name] = val
            if pd.notna(val) and val > 0:
                total_pen += val

        fecha_exp = fecha_ref.date() if hasattr(fecha_ref, "date") else fecha_ref
        row = {
            "Alias":           alias,
            "BBG Ticker":      bbg,
            "Código SBS":      sbs,
            "Fecha Exposición": str(fecha_exp),
        }
        for fondo_name in fondos:
            row[f"{fondo_name} PxQ (S/.)"] = fondo_vals.get(fondo_name, np.nan)
        row["Total Exposición (S/.)"]  = total_pen if total_pen > 0 else np.nan
        row["Total Exposición (S/.M)"] = round(total_pen / 1e6, 2) if total_pen > 0 else np.nan

        # Z-score actual (para referencia cruzada con el Chart 2)
        z_emp = [c for c in bd.columns if c.startswith("Z_")
                 and "Basket" not in c and "BVL" not in c and "MXLA" not in c]
        z_last = bd.loc[bd[z_emp].notna().any(axis=1), z_col].dropna() if z_col in bd.columns else pd.Series()
        row["Z_score actual"] = round(float(z_last.iloc[-1]), 4) if not z_last.empty else np.nan
        rows.append(row)

    df_out = pd.DataFrame(rows)

    # ── Escribir con formato ─────────────────────────────────────────────────
    ruta_out = os.path.join(RUTA_SALIDA, "QUEST - exposicion de cada activo.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exposición"

    # Estilos
    navy_fill = PatternFill("solid", fgColor="1E2E6E")
    teal_fill = PatternFill("solid", fgColor="E3E829")
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    sub_font  = Font(name="Calibri", bold=True, color="1E2E6E", size=10)
    body_font = Font(name="Calibri", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_side = Side(style="thin", color="D0D7E5")
    thin_brd  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Fila de metadatos
    ws["A1"] = "AFP Integra · ZODA — Exposición por Activo (validación Chart 2)"
    ws["A1"].font = Font(name="Calibri", bold=True, color="1E2E6E", size=12)
    ws["A2"] = f"Fecha de referencia: {fecha_ref.date()}   |   Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", color="7E8083", size=9, italic=True)
    ws.append([])  # fila 3 vacía

    # Encabezados de columna (fila 4)
    headers = list(df_out.columns)
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_i, value=h)
        cell.fill  = navy_fill
        cell.font  = hdr_font
        cell.alignment = center
        cell.border = thin_brd

    # Datos (desde fila 5)
    for row_i, row in enumerate(df_out.itertuples(index=False), start=5):
        for col_i, val in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col_i)
            if isinstance(val, float) and np.isnan(val):
                cell.value = ""
            elif isinstance(val, (np.integer, np.floating)):
                cell.value = float(val)
            else:
                cell.value = val
            cell.font   = body_font
            cell.border = thin_brd
            cell.alignment = Alignment(horizontal="right" if col_i > 4 else "left",
                                        vertical="center")
            # Formato numérico para columnas de exposición
            if col_i > 4 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        # Cebra
        if row_i % 2 == 0:
            for col_i in range(1, len(headers)+1):
                ws.cell(row=row_i, column=col_i).fill = PatternFill("solid", fgColor="F5F7FA")

    # Anchos de columna
    col_widths = {"Alias":10,"BBG Ticker":20,"Código SBS":18,"Fecha Exposición":18}
    for col_i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = col_widths.get(h, 22)

    # Fila total al final
    last_row = 4 + len(df_out) + 1
    ws.cell(row=last_row, column=1, value="TOTAL PORTAFOLIO").font = sub_font
    total_col = headers.index("Total Exposición (S/.M)") + 1
    total_val = df_out["Total Exposición (S/.M)"].sum(min_count=1)
    ws.cell(row=last_row, column=total_col, value=round(total_val, 2)).font = sub_font

    ws.freeze_panes = "A5"
    wb.save(ruta_out)
    print(f"  ✓ Excel validación: {os.path.basename(ruta_out)}")
    print(f"    → {len(df_out)} activos | fecha ref: {fecha_ref.date()} | S/. {total_val:.0f}M total")


# ─── HELPER: extrae Px_last y CIQ trimestral de cualquier hoja ──────────────
_COLS_ALL_SHEETS = ["IQ_NI","IQ_TOTAL_REV","IQ_TOTAL_EQUITY","IQ_EBITDA",
                    "Px_last","CUR_MKT_CAP","CURR_ENTP_VAL"]

def _extract_px_and_ciq(sheet_df, ticker, fechas_d, fechas_tri):
    """Devuelve (px_series_diaria, ciq_df_trimestral) para un ticker.
    Manejo robusto de posición fija de columnas (misma lógica que ZODA_03_06_2026)."""
    mask = sheet_df.iloc[4] == ticker
    emp  = sheet_df.loc[:, mask]
    if emp.empty or len(emp.columns) < 5:
        return pd.Series(dtype=float), pd.DataFrame()
    cols = [_COLS_ALL_SHEETS[i] if i < len(_COLS_ALL_SHEETS) else f"col_{i}"
            for i in range(len(emp.columns))]
    emp_ = emp.iloc[5:].copy(); emp_.columns = cols
    # --- CIQ ---
    ciq_raw = emp_.iloc[:, :4].dropna(axis=0, how="all")
    n_tri   = min(len(fechas_tri), len(ciq_raw))
    ciq_out = ciq_raw.iloc[:n_tri].copy()
    ciq_out.index = pd.DatetimeIndex(fechas_tri.values[:n_tri])
    ciq_out = ciq_out.apply(pd.to_numeric, errors="coerce").fillna(0)
    # --- Px_last ---
    px_num = pd.to_numeric(emp_["Px_last"], errors="coerce")
    n_d    = min(len(fechas_d), len(px_num))
    px_s   = pd.Series(px_num.values[:n_d], index=pd.DatetimeIndex(fechas_d.values[:n_d]))
    px_s   = px_s[px_s > 0].dropna()
    return px_s, ciq_out


def _load_excel_sheets():
    """Carga las 3 hojas de datos + fechas. Usado por los 4 nuevos charts."""
    sheets = {}
    for sheet in ["Financials", "Cons y Const", "Min y Uti"]:
        df = _read(sheet, header=None, nrows=4020)
        df.columns = [f"c{i}" for i in range(len(df.columns))]
        sheets[sheet] = df
    # Fechas del sheet Financials (referencia)
    fin = sheets["Financials"]
    fechas_tri = pd.to_datetime(fin["c1"].iloc[5:], errors="coerce").dropna()
    mask_px    = fin.iloc[2] == "Px_last"
    px_col     = fin.columns[mask_px][0]
    fechas_d   = pd.to_datetime(fin.loc[5:, px_col], errors="coerce").dropna()
    return sheets, fechas_d, fechas_tri


def _ticker_sheet(ticker, sheets):
    """Devuelve la hoja donde existe el ticker."""
    for name, df in sheets.items():
        if (df.iloc[4] == ticker).any():
            return df
    return None


# ══════════════════════════════════════════════════════════════════════════════
# CHART A — Calendarized Return Fan
# Ref: Goldman Sachs Global Investment Research style + Doeswijk et al. (2020)
# ══════════════════════════════════════════════════════════════════════════════

def build_calendarized_data(bd, param, gics):
    """
    Para la empresa seleccionada, calcula el retorno YTD acumulado día a día
    para cada año calendario disponible. El año actual se superpone sobre las
    bandas históricas (P10, P25, mediana, P75, P90) de todos los años anteriores.

    Respaldo: Doeswijk, Lam & Swinkels (2020) — los retornos YTD tienen alta
    persistencia intra-año pero fuerte reversión inter-año cuando superan el P90
    histórico. Permite cuantificar si el rally actual ya está 'consumido'.
    """
    sheets, fechas_d, fechas_tri = _load_excel_sheets()
    result = {}

    for _, r in param.iterrows():
        alias = r["alias"]
        z_col = r["z_col"]
        bbg   = r["bbg_ticker"]
        if z_col not in bd.columns or z_col.startswith("nan"): continue
        if "Index" in z_col: continue

        sector = gics.loc[bbg,"GICS"] if bbg in gics.index else "Other"
        nombre = gics.loc[bbg,"Empresa"] if bbg in gics.index else alias

        sheet_df = _ticker_sheet(bbg, sheets)
        if sheet_df is None: continue
        px_s, _ = _extract_px_and_ciq(sheet_df, bbg, fechas_d, fechas_tri)
        if px_s.empty or len(px_s) < 100: continue

        # Compute YTD return for each calendar year
        current_year = px_s.index.max().year
        ytd_by_year  = {}

        for yr in sorted(px_s.index.year.unique()):
            yr_data = px_s[px_s.index.year == yr]
            if len(yr_data) < 20: continue
            base = float(yr_data.iloc[0])
            # Day-of-year (0-365) → YTD return %
            doys = [(d - pd.Timestamp(f"{yr}-01-01")).days for d in yr_data.index]
            rets = [round((float(v)/base - 1)*100, 3) for v in yr_data.values]
            ytd_by_year[str(yr)] = {"doy": doys, "ytd": rets}

        # Historical bands (all years except current)
        hist_years = [k for k in ytd_by_year if int(k) < current_year]
        if len(hist_years) < 3: continue

        # Align to common DOY grid (0-364)
        doy_grid = list(range(0, 365, 1))
        # For each historical year, interpolate YTD at each doy
        hist_matrix = []
        for yr_key in hist_years:
            d  = ytd_by_year[yr_key]
            s  = pd.Series(d["ytd"], index=d["doy"])
            s  = s[~s.index.duplicated(keep="first")]
            si = s.reindex(doy_grid).interpolate(method="index", limit_direction="both")
            hist_matrix.append(si.values.tolist())

        hist_arr = np.array(hist_matrix)  # shape: (n_years, 365)
        bands = {
            "p10":    np.nanpercentile(hist_arr, 10, axis=0).round(2).tolist(),
            "p25":    np.nanpercentile(hist_arr, 25, axis=0).round(2).tolist(),
            "median": np.nanpercentile(hist_arr, 50, axis=0).round(2).tolist(),
            "p75":    np.nanpercentile(hist_arr, 75, axis=0).round(2).tolist(),
            "p90":    np.nanpercentile(hist_arr, 90, axis=0).round(2).tolist(),
            "doy":    doy_grid,
        }

        result[alias] = dict(
            alias=alias, empresa=nombre, sector=sector,
            color=SECTOR_COLOR.get(sector, BRAND["muted"]),
            current_year=str(current_year),
            ytd_current=ytd_by_year.get(str(current_year), {}),
            ytd_prev_year=ytd_by_year.get(str(current_year-1), {}),
            hist_years=hist_years,
            bands=bands,
        )

    return result

# ─── GENERAR HTML ────────────────────────────────────────────────────────────
def generar_html(bd, param, gics, pos, exposiciones, mult_data, fecha_gen):
    print("  → Gráfico 1: histórico…")
    hist, interp = build_historico(bd, param, gics, mult_data)
    print("  → Gráfico 2: bubble chart…")
    bubble = build_bubble_data(bd, param, gics, exposiciones)
    print("  → Gráfico 3: heatmap…")
    heat   = build_heatmap_data(bd, param, gics)
    print("  → Gráfico 4: EQRV…")
    eqrv   = build_eqrv_data(bd, param, gics)
    print("  → Gráfico 6: Calendarized Return Fan…")
    calendar_data  = build_calendarized_data(bd, param, gics)

    z_emp = [c for c in bd.columns if c.startswith("Z_")
             and "Basket" not in c and "BVL" not in c
             and "MXLA" not in c and "Index" not in c]
    last_d = bd.loc[bd[z_emp].notna().any(axis=1),"Fecha"].max()
    last_s = last_d.strftime("%d/%m/%Y") if pd.notna(last_d) else ""
    gen_disp = f"{fecha_gen} (BD al {last_s})"

    logo_b64 = ""
    try:
        if os.path.exists(RUTA_LOGO):
            with open(RUTA_LOGO,"rb") as f:
                logo_b64 = "data:image/png;base64," + base64.b64encode(f.read()).decode()
    except Exception:
        pass

    html = HTML
    html = html.replace("__HIST_DATA__",   json.dumps(hist,        ensure_ascii=False))
    html = html.replace("__INTERP_DATA__", json.dumps(interp,      ensure_ascii=False))
    html = html.replace("__BUBBLE_DATA__", json.dumps(bubble,      ensure_ascii=False))
    html = html.replace("__HEAT_DATA__",   json.dumps(heat,        ensure_ascii=False))
    html = html.replace("__EQRV_DATA__",   json.dumps(eqrv,        ensure_ascii=False))
    html = html.replace("__MULT_DATA__",   json.dumps(mult_data,   ensure_ascii=False))
    html = html.replace("__MULT_LABELS__",    json.dumps(MULTIPLOS_LABELS,  ensure_ascii=False))
    html = html.replace("__CALENDAR_DATA__",  json.dumps(calendar_data,   ensure_ascii=False))
    html = html.replace("__LOGO_B64__",       logo_b64)
    html = html.replace("__GEN_DATE__",       gen_disp)
    return html

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    t0 = datetime.datetime.now()
    fecha_gen = t0.strftime("%d/%m/%Y %H:%M")
    fecha_tag = t0.strftime("%d.%m.%y")

    print("═"*60)
    print(f"  ZODA Dashboard v2 · AFP Integra")
    print(f"  {fecha_gen}")
    print("═"*60)

    print("\n[1/6] Leyendo BD…")
    bd   = cargar_bd()
    print(f"  → {len(bd)} filas | {bd['Fecha'].min().date()} → {bd['Fecha'].max().date()}")

    print("\n[2/6] Leyendo Parametría y GICS…")
    param = cargar_parametria()
    gics  = cargar_gics()
    print(f"  → {len(param)} activos mapeados")

    print("\n[3/6] Leyendo exposiciones DQ_IN…")
    exposiciones = cargar_exposiciones()   # {date_str → {sbs → PEN}}
    pos = cargar_posiciones()              # legacy DataFrame
    if exposiciones:
        last_exp = max(exposiciones.keys())
        n_activos = len(exposiciones[last_exp])
        print(f"  → Exposiciones hasta {last_exp} | {n_activos} instrumentos con posición")
        print("  → Exportando Excel de validación de exposiciones…")
        exportar_exposicion_excel(exposiciones, param, bd)
    else:
        print("  ⚠ Sin datos de posición disponibles")

    print("\n[4/6] Cargando múltiplos…")
    mult_data, desde_csv = cargar_multiplos_csv()
    if desde_csv:
        print(f"  → {len(mult_data)} tickers desde CSV pre-calculado")
        print(f"     (ruta: {os.path.basename(RUTA_MULTIPLOS_CSV)})")
    else:
        print(f"  ⚠  CSV no encontrado. Calculando en tiempo real…")
        print(f"     (ejecuta ZODA_03_06_2026.py para generar {os.path.basename(RUTA_MULTIPLOS_CSV)})")
        mult_data, _ = build_multiplos_data()
        print(f"  → {len(mult_data)} tickers calculados en tiempo real")

    print("\n[5/6] Generando HTML…")
    html = generar_html(bd, param, gics, pos, exposiciones, mult_data, fecha_gen)

    nombre = f"ZODA_Dashboard_{fecha_tag}.html"
    ruta   = os.path.join(RUTA_SALIDA, nombre)
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(html)

    tf     = datetime.datetime.now()
    tam_mb = os.path.getsize(ruta) / 1_048_576
    print(f"\n[6/6] Guardado.")
    print(f"  ✓ Archivo: {ruta}")
    print(f"  ✓ Tamaño:  {tam_mb:.1f} MB")
    print(f"  ✓ Tiempo:  {tf - t0}")
    print("═"*60)

if __name__ == "__main__":
    main()
