
# -*- coding: utf-8 -*-
"""
================================================================================
EXTENSIÓN — DASHBOARD DE ACTIVOS FÍSICOS Y DRILL-DOWN GEOGRÁFICO (AFP INTEGRA)
================================================================================
Módulo ADITIVO al dashboard principal (Dashboard_Physical_Risks.py), enganchado
exactamente igual que rf_extension: main() calcula 'districts', 'exp_clase',
'aum_cartera', 'emp_sector', 'geojson' y los pasa a generar_dashboard_activos(),
que produce un HTML APARTE. Si algo falla, NO afecta al dashboard original.

Qué genera
----------
Un HTML autocontenido (Leaflet vía CDN, sin tiles -> océano celeste propio) con:

  · Pestañas Resumen / Renta Variable / Renta Fija (mismo filtro de clase que el
    padre, vía la columna S TIPO_RENTA -> exp_clase).
  · POR PESTAÑA, TRES bloques en una misma fila (uno por peligro CENEPRED:
    Movimientos de Masa, Inundación, Sequías Severas). Cada bloque tiene:
       - Una TABLA con headers: Riesgo | AUM Total | Emisor | AUM por Emisor |
         Activo (top-3 activos por plata). La columna "Activo" se puede ocultar.
         Depto/Prov/Distrito NO son columnas: aparecen al pasar el cursor.
       - Un MAPA interactivo del Perú con drill-down jerárquico estricto
         País -> Departamento -> Provincia -> Distrito (click en el mapa o
         dropdowns en cascada). Distritos pintados en escala secuencial Azul
         Marino según el score del peligro (1-5). A partir de la capa
         Departamental se dibujan los activos físicos como íconos (mina / banco /
         círculo) cuyo tamaño es proporcional a la plata que representan.

  · Toggles globales: montos S/ on/off, conversión S/ <-> USD (tipo de cambio
    leído de la hoja "Tipo de Cambio", fila A=="DOL", columna D; default 3.5),
    ocultar columna Activo, sincronizar navegación de los 3 mapas.

Modelo de plata por activo (confirmado con el usuario)
------------------------------------------------------
La exposición ya viene asignada a nivel (emisor, distrito) en districts[].aportes
[{emisor, sector, monto, activos:[nombres]}]. El monto del distrito-emisor se
reparte EQUITATIVAMENTE entre los activos de ese distrito:
        plata_por_activo = (monto * fracción_de_clase) / nº_activos_del_distrito
Es una APROXIMACIÓN (no hay plata por unidad individual en la fuente) y así se
rotula en la metodología del HTML.
================================================================================
"""

import os
import re
import json
from datetime import datetime

# Paleta corporativa (coincide con el padre; se sobre-escribe si llega vía AFP).
AFP_AZUL = "#1E2E6E"
AFP_CYAN = "#00AECB"
AFP_AMAR = "#E3E829"
AFP_GRIS = "#DCDDDE"
AFP_PLOMO = "#7E8083"

# Niveles de riesgo que se listan en las tablas (igual criterio que el padre).
NIVELES_ALTOS = [5, 4, 3]

# Tipo de cambio por defecto si no se logra leer la hoja "Tipo de Cambio".
FX_DEFAULT = 3.5

# Mapeo sector -> tipo de ícono en el mapa.
#   Minería -> mina ; Financiero -> banco ; el resto -> círculo.
SECTOR_ICONO = {
    "Minería": "mina",
    "Financiero": "banco",
}
# Color del ícono por sector (para distinguir en el mapa).
SECTOR_COLOR = {
    "Minería": "#8c5a2b",       # marrón mina
    "Financiero": "#1E2E6E",    # azul AFP
    "Utilities": "#16a085",     # verde-agua
    "Construcción": "#7f8c8d",  # plomo cemento
    "Consumo": "#c0392b",
    "Salud": "#2980b9",
    "Infraestructura": "#d35400",
    "Educación": "#8e44ad",
    "Energía": "#f39c12",
}
SECTOR_COLOR_DEFAULT = "#555555"


# ==============================================================================
# 1) UTILIDADES — TIPO DE CAMBIO Y GEOMETRÍA
# ==============================================================================

def _norm_txt(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def leer_tipo_cambio(risk_path, logfn=print):
    """Lee el tipo de cambio USD de la hoja 'Tipo de Cambio': busca en la columna
    A la fila cuyo valor sea 'DOL' y devuelve la columna D (VALOR). Si no se
    encuentra o no es numérico, usa FX_DEFAULT y avisa con un mensaje claro."""
    info = {"valor": FX_DEFAULT, "fecha": None, "fuente": "default"}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(risk_path, read_only=True, data_only=True)
        hoja = None
        for nm in wb.sheetnames:
            if "TIPO" in _norm_txt(nm) and "CAMBIO" in _norm_txt(nm):
                hoja = nm
                break
        if hoja is None:
            wb.close()
            logfn("[Activos] Hoja 'Tipo de Cambio' no encontrada -> uso FX por "
                  f"defecto {FX_DEFAULT}.")
            return info
        ws = wb[hoja]
        encontrado = None
        for row in ws.iter_rows(min_row=1, max_row=120, max_col=4, values_only=True):
            a = row[0] if len(row) > 0 else None
            if a is not None and _norm_txt(a) == "DOL":
                valor = row[3] if len(row) > 3 else None
                fecha = row[2] if len(row) > 2 else None
                try:
                    valor = float(valor)
                    if valor > 0:
                        encontrado = (valor, fecha)
                        break
                except (TypeError, ValueError):
                    pass
        wb.close()
        if encontrado:
            info = {"valor": round(encontrado[0], 4),
                    "fecha": str(encontrado[1]) if encontrado[1] is not None else None,
                    "fuente": "leido"}
            logfn(f"[Activos] Tipo de cambio USD leído: {info['valor']} "
                  f"(fecha {info['fecha']}).")
        else:
            logfn("[Activos] No se halló la fila 'DOL' en columna A / valor en "
                  f"columna D -> uso FX por defecto {FX_DEFAULT}.")
    except Exception as e:
        logfn(f"[Activos] Error leyendo tipo de cambio ({e}) -> uso FX por "
              f"defecto {FX_DEFAULT}.")
    return info


def _iter_anillos_exteriores(geom):
    """Devuelve la lista de anillos EXTERIORES (lng,lat) de un Polygon/MultiPolygon."""
    if not geom:
        return []
    t = geom.get("type")
    c = geom.get("coordinates")
    if t == "Polygon":
        return [c[0]] if c else []
    if t == "MultiPolygon":
        return [poly[0] for poly in c if poly]
    return []


def _todas_coords(geom):
    """Itera TODOS los vértices (lng,lat) para calcular bbox."""
    if not geom:
        return
    t = geom.get("type")
    c = geom.get("coordinates")
    if t == "Polygon":
        for ring in c:
            for pt in ring:
                yield pt
    elif t == "MultiPolygon":
        for poly in c:
            for ring in poly:
                for pt in ring:
                    yield pt


def calcular_geometria(geojson):
    """Calcula, en Python puro (sin shapely), por cada distrito:
        - centroide [lat, lng]  (promedio de vértices del anillo exterior)
        - bbox [[sur,oeste],[norte,este]]
    y agrega bbox por provincia (clave 'DEP||PROV') y por departamento."""
    centroids = {}
    bbox_dist = {}
    bbox_prov = {}
    bbox_dep = {}

    def _upd_bbox(store, key, miny, minx, maxy, maxx):
        b = store.get(key)
        if b is None:
            store[key] = [[miny, minx], [maxy, maxx]]
        else:
            b[0][0] = min(b[0][0], miny); b[0][1] = min(b[0][1], minx)
            b[1][0] = max(b[1][0], maxy); b[1][1] = max(b[1][1], maxx)

    for f in geojson.get("features", []):
        geom = f.get("geometry")
        if not geom:
            continue
        prop = f.get("properties", {})
        u = str(prop.get("IDDIST", "")).split(".")[0].zfill(6)
        dep = prop.get("NOMBDEP")
        prov = prop.get("NOMBPROV")

        # centroide: promedio de vértices de los anillos exteriores
        sx = sy = 0.0
        n = 0
        for ring in _iter_anillos_exteriores(geom):
            for lng, lat in ring:
                sx += lng; sy += lat; n += 1
        if n:
            centroids[u] = [round(sy / n, 6), round(sx / n, 6)]

        # bbox
        minx = miny = float("inf")
        maxx = maxy = float("-inf")
        for lng, lat in _todas_coords(geom):
            minx = min(minx, lng); maxx = max(maxx, lng)
            miny = min(miny, lat); maxy = max(maxy, lat)
        if minx == float("inf"):
            continue
        bbox_dist[u] = [[round(miny, 6), round(minx, 6)],
                        [round(maxy, 6), round(maxx, 6)]]
        if dep:
            _upd_bbox(bbox_dep, dep, miny, minx, maxy, maxx)
        if dep and prov:
            _upd_bbox(bbox_prov, dep + "||" + prov, miny, minx, maxy, maxx)

    # redondeo final de bboxes agregadas
    for store in (bbox_prov, bbox_dep):
        for k, b in store.items():
            store[k] = [[round(b[0][0], 6), round(b[0][1], 6)],
                        [round(b[1][0], 6), round(b[1][1], 6)]]
    return centroids, bbox_dist, bbox_prov, bbox_dep


# ==============================================================================
# 2) CONSTRUCCIÓN DEL PAYLOAD Y FUNCIÓN PÚBLICA
# ==============================================================================

def _trim_districts(districts):
    """Reduce 'districts' a lo estrictamente necesario para este dashboard:
    ubigeo, geografía, los TRES scores por peligro y los aportes (emisor, sector,
    monto, activos)."""
    out = []
    for d in districts:
        ap = []
        for a in d.get("aportes", []):
            if a.get("monto", 0) <= 0:
                continue
            ap.append({
                "e": a.get("emisor"),
                "s": a.get("sector"),
                "m": round(float(a.get("monto", 0.0)), 2),
                "ac": list(a.get("activos", []) or []),
            })
        out.append({
            "u": d.get("ubigeo"),
            "dep": d.get("dep"), "prov": d.get("prov"), "dist": d.get("dist"),
            "smm": d.get("score_mm"), "sinu": d.get("score_inu"),
            "sseq": d.get("score_seq"),
            "ap": ap,
        })
    return out


def _sector_icono(sector):
    return SECTOR_ICONO.get(sector, "circulo")


def _color_nino(enfen):
    """Color/opacidad/etiqueta del rectángulo Región Niño 1+2 según el ÚLTIMO
    registro ENFEN cuyo estado no sea 'NO ACTIVO' / 'NO IDENTIFICADO'. Recibe la
    lista ENFEN ([{fecha, estado, conf}, ...]) ya parseada (mismo formato que usa
    el script padre). Devuelve un dict con las claves requeridas
    {color, opacity, label} y, de forma aditiva, el {estado, fecha} del registro
    usado (para el tooltip de hover en el mapa)."""
    neutro = {"color": "#aaaaaa", "opacity": 0.25, "label": "Condiciones Neutras",
              "estado": None, "fecha": None}
    if not enfen:
        return neutro

    def _sin_acentos(s):
        import unicodedata
        return "".join(c for c in unicodedata.normalize("NFD", _norm_txt(s))
                       if unicodedata.category(c) != "Mn")

    usado = None
    for rec in reversed(enfen):
        est = _sin_acentos(rec.get("estado"))
        if not est:
            continue
        if "NO ACTIVO" in est or "NO IDENTIFICADO" in est:
            continue
        usado = rec
        break
    if usado is None:
        return neutro
    est = _sin_acentos(usado.get("estado"))
    base = {"estado": usado.get("estado"), "fecha": usado.get("fecha")}
    if any(k in est for k in ("FUERTE", "MUY FUERTE", "EXTRAORDINARIO")):
        # 'MUY FUERTE' contiene 'FUERTE'; cualquiera de los tres -> rojo
        return {"color": "#dc2626", "opacity": 0.65, "label": "El Niño Fuerte", **base}
    if "MODERADO" in est:
        return {"color": "#F97316", "opacity": 0.55, "label": "El Niño Moderado", **base}
    if "DEBIL" in est:
        return {"color": "#facc15", "opacity": 0.45, "label": "El Niño Débil", **base}
    return {"color": "#aaaaaa", "opacity": 0.25, "label": "Condiciones Neutras", **base}


def generar_dashboard_activos(out_dir, districts, exp_clase, aum_cartera,
                              emp_sector, risk_path, geojson,
                              enfen=None, stamp=None, logfn=print):
    """Punto de entrada público (espejo de rf_extension.generar_extension)."""
    if stamp is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")

    fx = leer_tipo_cambio(risk_path, logfn=logfn)
    centroids, bbox_dist, bbox_prov, bbox_dep = calcular_geometria(geojson)

    # sectores presentes -> icono/color
    sectores = sorted({a.get("sector") for d in districts
                       for a in d.get("aportes", []) if a.get("sector")})
    sector_icono = {s: _sector_icono(s) for s in sectores}
    sector_color = {s: SECTOR_COLOR.get(s, SECTOR_COLOR_DEFAULT) for s in sectores}

    payload = {
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "districts": _trim_districts(districts),
        "exp_clase": exp_clase or {},
        "aum_cartera": aum_cartera or {"TODOS": 0, "RENTA VARIABLE": 0, "RENTA FIJA": 0},
        "emp_sector": emp_sector or {},
        "niveles_altos": NIVELES_ALTOS,
        "fx": fx,
        "centroids": centroids,
        "bbox_dist": bbox_dist,
        "bbox_prov": bbox_prov,
        "bbox_dep": bbox_dep,
        "sector_icono": sector_icono,
        "sector_color": sector_color,
        "nino": _color_nino(enfen),
        "afp": {"azul": AFP_AZUL, "cyan": AFP_CYAN, "amar": AFP_AMAR,
                "gris": AFP_GRIS, "plomo": AFP_PLOMO},
    }

    # GeoJSON minimal (solo lo necesario para pintar y navegar).
    geojson_min = {"type": "FeatureCollection", "features": [
        {"type": "Feature",
         "properties": {
             "IDDIST": str(f["properties"].get("IDDIST", "")).split(".")[0].zfill(6),
             "NOMBDIST": f["properties"].get("NOMBDIST"),
             "NOMBPROV": f["properties"].get("NOMBPROV"),
             "NOMBDEP": f["properties"].get("NOMBDEP")},
         "geometry": f["geometry"]}
        for f in geojson["features"] if f.get("geometry")]}

    html = _HTML_TEMPLATE.replace("__PAYLOAD__", json.dumps(payload, ensure_ascii=False))
    html = html.replace("__GEOJSON__", json.dumps(geojson_min, ensure_ascii=False))

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"Dashboard_Activos_Riesgo_{stamp}.html")
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    logfn(f"[Activos] HTML generado: {out_path}")
    return out_path


# ==============================================================================
# 3) TEMPLATE HTML + JS (Leaflet)
# ==============================================================================
_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Activos Físicos y Riesgo CENEPRED — AFP Integra</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
 :root{
   --azul:#1E2E6E; --cyan:#00AECB; --amar:#E3E829; --gris:#DCDDDE; --plomo:#7E8083;
   --grisbg:#f4f5f8; --txt:#1f2533; --linea:#DCDDDE; --celeste:#cfeefb;
   --nino:#ff7a18;
 }
 *{box-sizing:border-box;font-family:Calibri,'Segoe UI',Arial,sans-serif;}
 body{margin:0;background:#fff;color:var(--txt);font-size:14px;}
 .topbar{display:flex;align-items:center;justify-content:space-between;
   padding:14px 26px;border-bottom:3px solid var(--azul);background:#fff;}
 .topbar .center{text-align:center;flex:1;}
 .topbar .center h1{margin:0;font-size:19px;color:var(--azul);font-weight:700;}
 .topbar .center .sub{font-size:12px;color:var(--plomo);margin-top:2px;}
 .topbar .right{text-align:right;font-size:12px;color:var(--plomo);min-width:160px;}
 .topbar .right b{color:var(--azul);}
 .tabbar{display:flex;align-items:center;gap:4px;background:var(--grisbg);
   padding:8px 22px 0;border-bottom:1px solid var(--linea);flex-wrap:wrap;}
 .tab{padding:9px 18px;border:none;background:transparent;color:var(--azul);
   font-size:14px;font-weight:600;cursor:pointer;border-radius:7px 7px 0 0;
   border-bottom:3px solid transparent;}
 .tab:hover{background:#e8ebf3;}
 .tab.active{background:#fff;border-bottom:3px solid var(--cyan);color:var(--azul);}
 .modos{margin-left:auto;display:flex;align-items:center;gap:14px;padding-bottom:6px;
   flex-wrap:wrap;}
 .chk{font-size:12px;color:var(--plomo);display:inline-flex;align-items:center;gap:5px;
   cursor:pointer;}
 .seg{display:inline-flex;border:1px solid var(--azul);border-radius:7px;overflow:hidden;}
 .seg button{border:none;background:#fff;color:var(--azul);padding:6px 12px;
   font-size:12.5px;cursor:pointer;font-weight:600;}
 .seg button.on{background:var(--azul);color:#fff;}
 .wrap{max-width:1860px;margin:0 auto;padding:14px 18px 30px;}
 /* ---- Fila de 3 bloques (peligros) ---- */
 .grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;align-items:stretch;}
 @media(max-width:1200px){.grid3{grid-template-columns:1fr;}}
 .blk{background:#fff;border:1px solid var(--linea);border-radius:10px;
   padding:10px 12px;display:flex;flex-direction:column;}
 .blk h3{margin:0 0 2px;font-size:14.5px;color:var(--azul);display:flex;
   align-items:center;gap:8px;}
 .blk h3 .dot{width:12px;height:12px;border-radius:3px;display:inline-block;}
 .blk .note{font-size:11px;color:var(--plomo);margin:0 0 8px;}
 /* ---- Tablas ---- */
 .pivwrap{overflow:auto;max-height:340px;border:1px solid var(--linea);border-radius:8px;
   margin-bottom:10px;}
 table.piv{border-collapse:collapse;width:100%;font-size:11.5px;}
 table.piv thead th{position:sticky;top:0;background:var(--azul);color:#fff;
   padding:7px 8px;text-align:left;font-weight:600;white-space:nowrap;z-index:2;}
 table.piv td{padding:6px 8px;border-top:1px solid var(--linea);vertical-align:top;}
 table.piv td.riesgo{font-weight:800;font-size:15px;color:#fff;text-align:center;
   vertical-align:middle;width:34px;}
 .r5{background:#c0392b;} .r4{background:#e67e22;}
 .r3{background:#f1c40f;color:#5a4500 !important;}
 table.piv td.aumtot{background:#eef3ff;vertical-align:middle;font-weight:700;
   color:var(--azul);text-align:right;white-space:nowrap;}
 table.piv td.num{text-align:right;white-space:nowrap;}
 .pct{font-weight:700;color:var(--azul);font-size:12.5px;}
 .monto{display:block;font-size:10px;color:var(--plomo);font-weight:400;margin-top:1px;}
 body.ocultar-montos .monto{display:none;}
 .act{color:var(--txt);cursor:help;border-bottom:1px dotted #aab;}
 body.ocultar-activo .col-activo{display:none;}
 table.piv tbody tr:hover td:not(.riesgo):not(.aumtot){background:#f7f9ff;}
 /* ---- Mapa ---- */
 .mapbar{display:flex;flex-wrap:wrap;gap:6px;align-items:center;margin-bottom:6px;}
 .mapbar select{padding:4px 6px;border:1px solid var(--linea);border-radius:6px;
   font-size:11.5px;background:#fff;color:var(--txt);max-width:130px;}
 .mapbar select:disabled{opacity:.5;}
 .crumb{font-size:11px;color:var(--azul);font-weight:600;margin-right:auto;}
 .crumb a{color:var(--cyan);cursor:pointer;text-decoration:none;}
 .crumb a:hover{text-decoration:underline;}
 .mapa{width:100%;height:430px;border:1px solid var(--linea);border-radius:8px;
   background:var(--celeste);}
 .leaflet-container{background:var(--celeste);}
 .leaflet-tooltip.tt{font-size:11px;}
 /* ícono de activo (divIcon) */
 .act-ic{display:flex;align-items:center;justify-content:center;}
 .act-ic svg{filter:drop-shadow(0 1px 1px rgba(0,0,0,.35));}
 .act-badge{background:var(--azul);color:#fff;border-radius:50%;font-size:10px;
   font-weight:700;display:flex;align-items:center;justify-content:center;
   width:20px;height:20px;border:2px solid #fff;}
 /* leyenda */
 .leg{display:flex;flex-wrap:wrap;gap:10px;margin-top:6px;font-size:10.5px;
   color:var(--txt);align-items:center;}
 .leg .sw{display:inline-flex;align-items:center;gap:4px;}
 .leg i{width:13px;height:13px;border-radius:2px;display:inline-block;}
 details.metod{margin-top:8px;} details.metod summary{cursor:pointer;color:var(--azul);
   font-weight:600;font-size:12px;} details.metod p{font-size:11px;color:var(--plomo);}
 footer{border-top:1px solid var(--linea);margin-top:14px;padding:14px 26px;
   color:var(--plomo);font-size:11.5px;line-height:1.6;background:var(--grisbg);}
 footer b{color:var(--azul);}
</style></head>
<body>
<div class="topbar">
  <div style="min-width:120px"></div>
  <div class="center">
    <h1>Activos Físicos y Riesgo CENEPRED</h1>
    <div class="sub">Mesa de Inversiones · Top de activos por plata y navegación País → Departamento → Provincia → Distrito</div>
  </div>
  <div class="right">Tipo de cambio<br><b id="fxLbl">—</b><br><span id="gen"></span></div>
</div>

<div class="tabbar">
  <button class="tab active" data-tab="resumen">Resumen</button>
  <button class="tab" data-tab="rv">Renta Variable</button>
  <button class="tab" data-tab="rf">Renta Fija</button>
  <div class="modos">
    <label class="chk"><input type="checkbox" id="chkMonto"> Mostrar montos</label>
    <label class="chk"><input type="checkbox" id="chkActivo" checked> Columna "Activo"</label>
    <label class="chk"><input type="checkbox" id="chkSync" checked> Sincronizar mapas</label>
    <div class="seg" title="Moneda de los montos">
      <button id="mPEN" class="on">S/</button>
      <button id="mUSD">US$</button>
    </div>
  </div>
</div>

<div class="wrap">
  <div class="grid3" id="bloques"></div>
  <details class="metod">
    <summary>Metodología y supuestos</summary>
    <p>Cada bloque corresponde a un peligro CENEPRED y agrupa a los emisores por el
    score distrital de ESE peligro (1–5). El "AUM Total" por nivel y el "AUM por
    Emisor" usan como denominador el AUM de la cartera según la pestaña
    (Total / Renta Variable / Renta Fija, vía la col. S TIPO_RENTA). La plata de
    cada <b>activo físico individual</b> se estima repartiendo el monto del
    (emisor, distrito) en partes iguales entre los activos de ese distrito; es una
    <b>aproximación</b> (la fuente no trae plata por unidad). Los activos están
    georreferenciados a nivel <b>distrito</b>: en las capas Departamento y Provincia
    se muestran en el centroide de su distrito (no se inventan ubicaciones
    intra-provinciales). El tamaño del ícono es proporcional a √(plata) del activo.
    El rectángulo naranja marca la Región Niño 1+2 (0°–10°S).</p>
  </details>
</div>

<footer>
  <b>Fuentes:</b> CENEPRED (score distrital de peligros) · AFP Integra
  (Valorización de Portafolio) · Tipo de cambio: hoja "Tipo de Cambio" (DOL).<br>
  <span id="footgen"></span> · Documento autocontenido (Leaflet vía CDN, sin
  capa de teselas: el océano se dibuja localmente).
</footer>

<script>
const PAYLOAD = __PAYLOAD__;
const GEO = __GEOJSON__;
const AFP = PAYLOAD.afp;
const D = PAYLOAD.districts;
const EXP_CLASE = PAYLOAD.exp_clase || {};
const AUM = PAYLOAD.aum_cartera || {TODOS:0,'RENTA VARIABLE':0,'RENTA FIJA':0};
const EMP_SECTOR = PAYLOAD.emp_sector || {};
const NIV = PAYLOAD.niveles_altos || [5,4,3];
const FX = PAYLOAD.fx || {valor:3.5, fuente:'default'};
const CENTROID = PAYLOAD.centroids || {};
const BBOX_DEP = PAYLOAD.bbox_dep || {};
const BBOX_PROV = PAYLOAD.bbox_prov || {};
const BBOX_DIST = PAYLOAD.bbox_dist || {};
const SEC_ICONO = PAYLOAD.sector_icono || {};
const SEC_COLOR = PAYLOAD.sector_color || {};

const byU = {}; D.forEach(d=>byU[d.u]=d);

// ---- peligros (orden = orden de bloques)
const HZ = [
  {key:'mm',  col:'smm',  label:'Movimientos de Masa', dot:'#8c5a2b'},
  {key:'inu', col:'sinu', label:'Inundación',          dot:'#2980b9'},
  {key:'seq', col:'sseq', label:'Sequías Severas',     dot:'#d4a017'},
];

// ---- estado global
let TAB = 'resumen';
let MONEDA = 'PEN';          // 'PEN' | 'USD'
let SYNC = true;
const classKey = () => TAB==='rv'?'RENTA VARIABLE' : TAB==='rf'?'RENTA FIJA' : 'TODOS';

// ---- formato
function conv(v){ return MONEDA==='USD' ? (v/ (FX.valor||3.5)) : v; }
function fmtMonto(v){
  const x = conv(v||0);
  const sym = MONEDA==='USD' ? 'US$ ' : 'S/ ';
  return sym + x.toLocaleString('es-PE',{maximumFractionDigits:0});
}
const fmtP = v => (v*100).toLocaleString('es-PE',{maximumFractionDigits:1})+'%';

// ---- fracción de clase RV/RF de un emisor (igual criterio que el padre)
function classFrac(emisor){
  const ck = classKey();
  if(ck==='TODOS') return 1;
  const e = EXP_CLASE[emisor];
  if(!e || !e.TOTAL) return 0;
  return (e[ck]||0)/e.TOTAL;
}

// ======================================================================
//  PLATA POR ACTIVO  (reparto equitativo dentro del distrito-emisor)
//  Devuelve, para un distrito d: lista de {emisor, sector, activo, plata,
//  dep, prov, dist}. Filtra por la clase (pestaña) activa.
// ======================================================================
function activosDeDistrito(d){
  const out=[];
  (d.ap||[]).forEach(a=>{
    const m = a.m * classFrac(a.e);
    if(m<=0) return;
    const acts = (a.ac && a.ac.length) ? a.ac : ['(activo s/d)'];
    const cuota = m/acts.length;          // reparto equitativo
    acts.forEach(nombre=>{
      out.push({emisor:a.e, sector:a.s, activo:nombre, plata:cuota,
                dep:d.dep, prov:d.prov, dist:d.dist, u:d.u});
    });
  });
  return out;
}

// ======================================================================
//  TABLAS — una por peligro
// ======================================================================
function buildTablaHTML(hz){
  const col = hz.col;
  const ck = classKey();
  const denom = AUM[ck] || 1;
  // bucket[nivel][emisor] = {amt, activos:[{activo,plata,dep,prov,dist}]}
  const bucket={}; NIV.forEach(L=>bucket[L]={});
  D.forEach(d=>{
    const L = d[col];
    if(NIV.indexOf(L)<0) return;          // solo niveles altos 3-5
    (d.ap||[]).forEach(a=>{
      const m = a.m*classFrac(a.e);
      if(m<=0) return;
      const b = bucket[L][a.e] || (bucket[L][a.e]={amt:0,activos:[],sector:a.s});
      b.amt += m;
      const acts=(a.ac&&a.ac.length)?a.ac:['(activo s/d)'];
      const cuota=m/acts.length;
      acts.forEach(n=>b.activos.push({activo:n,plata:cuota,
        dep:d.dep,prov:d.prov,dist:d.dist}));
    });
  });
  // filas
  let rows=[];
  NIV.forEach(L=>{
    const ems=Object.entries(bucket[L]).sort((a,b)=>b[1].amt-a[1].amt);
    const aumL=ems.reduce((s,e)=>s+e[1].amt,0);
    ems.forEach(([emisor,b])=>{
      // agregar plata por nombre de activo (mismo activo en varios distritos suma)
      const porAct={};
      b.activos.forEach(x=>{
        const k=x.activo;
        if(!porAct[k]) porAct[k]={activo:k,plata:0,ubic:{}};
        porAct[k].plata+=x.plata;
        const uk=x.dist+' ('+x.prov+', '+x.dep+')';
        porAct[k].ubic[uk]=(porAct[k].ubic[uk]||0)+x.plata;
      });
      const top3=Object.values(porAct).sort((a,b)=>b.plata-a.plata).slice(0,3);
      rows.push({nivel:L, aumL, aumLpct:aumL/denom, emisor,
        emisorAmt:b.amt, sector:b.sector, top3,
        nTot:Object.keys(porAct).length});
    });
  });
  // armar HTML con rowspan en Riesgo y AUM Total
  let h='<thead><tr><th>Riesgo</th><th>AUM Total</th><th>Emisor</th>'+
        '<th>AUM por Emisor</th><th class="col-activo">Activo (top 3)</th></tr></thead><tbody>';
  if(!rows.length){
    h+='<tr><td colspan="5" style="padding:14px;color:var(--plomo)">'+
       'Sin emisores con score '+NIV.join('–')+' para este filtro.</td></tr></tbody>';
    return h;
  }
  const cnt={}, first={};
  rows.forEach((r,i)=>{cnt[r.nivel]=(cnt[r.nivel]||0)+1; if(first[r.nivel]===undefined)first[r.nivel]=i;});
  rows.forEach((r,i)=>{
    h+='<tr>';
    if(first[r.nivel]===i){
      h+='<td class="riesgo r'+r.nivel+'" rowspan="'+cnt[r.nivel]+'">'+r.nivel+'</td>';
      h+='<td class="aumtot" rowspan="'+cnt[r.nivel]+'"><span class="pct">'+fmtP(r.aumLpct)+
         '</span><span class="monto">'+fmtMonto(r.aumL)+'</span></td>';
    }
    h+='<td><b>'+r.emisor+'</b><br><span class="monto" style="color:var(--plomo)">'+
       (EMP_SECTOR[r.emisor]||r.sector||'')+'</span></td>';
    const emisorTot=(EXP_CLASE[r.emisor]&&EXP_CLASE[r.emisor].TOTAL)||0;
    const ep = emisorTot? r.emisorAmt/emisorTot : 0;
    h+='<td class="num"><span class="pct">'+fmtP(ep)+'</span><span class="monto">'+
       fmtMonto(r.emisorAmt)+'</span></td>';
    // activos top3 (con tooltip de ubicación)
    let actHtml = r.top3.map(a=>{
      const ub=Object.entries(a.ubic).sort((x,y)=>y[1]-x[1])
        .map(e=>e[0]+': '+fmtMonto(e[1])).join(' | ');
      const tip=(a.activo+' — '+fmtMonto(a.plata)+'  ·  '+ub).replace(/"/g,'&quot;');
      return '<span class="act" title="'+tip+'">'+a.activo+
             ' <span class="monto" style="display:inline">'+fmtMonto(a.plata)+'</span></span>';
    }).join('<br>');
    if(r.nTot>3) actHtml+='<br><span class="monto" style="display:inline">(+'+(r.nTot-3)+' activos más)</span>';
    h+='<td class="col-activo act-list">'+actHtml+'</td>';
    h+='</tr>';
  });
  h+='</tbody>';
  return h;
}

// ======================================================================
//  MAPAS — Leaflet con drill-down País -> Dep -> Prov -> Distrito
// ======================================================================
const PERU_BOUNDS = [[-18.5,-81.5],[0.5,-68.5]];
const NINO_RECT   = [[-10,-84.0],[0,-80.2]];   // 0-10°S, 80°W-costa (aprox.)
const SCORE_COLORS = ['#dce6f5','#a9c2e8','#6f95d4','#3a5fae','#1E2E6E']; // 1..5
function scoreColor(s){ return (s>=1&&s<=5)? SCORE_COLORS[s-1] : '#eef1f6'; }
const NINO = PAYLOAD.nino || {color:'#aaaaaa',opacity:0.25,label:'Condiciones Neutras',estado:null,fecha:null};

// ---- países vecinos (Natural Earth) en color tierra, para tapar el océano
//      celeste fuera del Perú. Se descarga UNA sola vez (promesa cacheada) y
//      cada mapa lo dibuja en su propio pane 'vecinos'.
let _vecinosProm=null;
function cargarVecinos(){
  if(_vecinosProm) return _vecinosProm;
  const SET=new Set(['Ecuador','Colombia','Brazil','Bolivia','Chile']);
  _vecinosProm=fetch('https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_50m_admin_0_countries.geojson')
    .then(r=>r.json())
    .then(gj=>({type:'FeatureCollection',
      features:(gj.features||[]).filter(f=>f.properties&&SET.has(f.properties.ADMIN))}))
    .catch(e=>{ console.warn('No se pudo cargar países vecinos:',e); return null; });
  return _vecinosProm;
}

// índice de provincias por departamento (para dropdowns)
const PROV_DE_DEP={}; const DIST_DE_PROV={};
GEO.features.forEach(f=>{
  const p=f.properties; const dep=p.NOMBDEP, prov=p.NOMBPROV, dist=p.NOMBDIST, u=p.IDDIST;
  if(!dep) return;
  (PROV_DE_DEP[dep]=PROV_DE_DEP[dep]||new Set()).add(prov);
  const k=dep+'||'+prov; (DIST_DE_PROV[k]=DIST_DE_PROV[k]||[]).push({dist,u});
});

// SVG de íconos (mina / banco / círculo), coloreado por sector
function iconSVG(tipo, color, px){
  const s=px;
  if(tipo==='mina'){
    return '<svg width="'+s+'" height="'+s+'" viewBox="0 0 24 24">'+
      '<path d="M2 20 L9 8 L13 14 L16 9 L22 20 Z" fill="'+color+'" stroke="#fff" stroke-width="1"/>'+
      '<circle cx="16" cy="6" r="2.2" fill="#ffd34d" stroke="#fff" stroke-width=".8"/></svg>';
  }
  if(tipo==='banco'){
    return '<svg width="'+s+'" height="'+s+'" viewBox="0 0 24 24">'+
      '<path d="M12 3 L22 8 L2 8 Z" fill="'+color+'" stroke="#fff" stroke-width="1"/>'+
      '<rect x="3" y="9" width="3" height="9" fill="'+color+'"/>'+
      '<rect x="10.5" y="9" width="3" height="9" fill="'+color+'"/>'+
      '<rect x="18" y="9" width="3" height="9" fill="'+color+'"/>'+
      '<rect x="2" y="19" width="20" height="2.5" fill="'+color+'" stroke="#fff" stroke-width=".6"/></svg>';
  }
  return '<svg width="'+s+'" height="'+s+'" viewBox="0 0 24 24">'+
    '<circle cx="12" cy="12" r="9" fill="'+color+'" stroke="#fff" stroke-width="1.4"/></svg>';
}

// Una instancia de mapa por peligro
class MapaPeligro{
  constructor(hz, el){
    this.hz=hz; this.col=hz.col;
    this.level='pais'; this.dep=null; this.prov=null; this.dist=null;
    this.map=L.map(el,{preferCanvas:true, zoomControl:true, attributionControl:false,
                       minZoom:4, maxZoom:13});
    this.map.fitBounds(PERU_BOUNDS);
    // pane océano (debajo de los distritos) para el rectángulo Niño 1+2
    this.map.createPane('oceano'); this.map.getPane('oceano').style.zIndex=350;
    // pane vecinos (encima del océano, debajo de los distritos): países limítrofes
    // en color tierra para que el celeste solo cubra el mar. Se crea una sola vez
    // al arrancar y NO se retoca en los re-renders del drill-down.
    this.map.createPane('vecinos'); this.map.getPane('vecinos').style.zIndex=360;
    cargarVecinos().then(data=>{ if(data) L.geoJSON(data,{pane:'vecinos',
      interactive:false, style:{fillColor:'#e8e0d0',fillOpacity:0.95,
        color:'#c9bfaf',weight:0.6}}).addTo(this.map); });
    // rectángulo Región Niño 1+2 (color/opacidad dinámicos según último ENFEN)
    const rectNino=L.rectangle(NINO_RECT,{pane:'oceano',color:NINO.color,weight:0,
      fillColor:NINO.color,fillOpacity:NINO.opacity,interactive:true}).addTo(this.map);
    rectNino.bindTooltip('Región Niño 1+2 · '+NINO.label,
      {permanent:true, direction:'top', className:'tt'});
    // tooltip flotante en hover: estado + fecha del registro ENFEN usado
    const ttHover=L.tooltip({direction:'top', className:'tt', sticky:true});
    rectNino.on('mouseover', e=>{
      const det=(NINO.estado? NINO.estado : 'Sin alerta El Niño activa')+
                (NINO.fecha? '<br>Último ENFEN: '+NINO.fecha : '');
      ttHover.setLatLng(e.latlng).setContent('<b>Región Niño 1+2</b><br>'+det);
      this.map.openTooltip(ttHover);
    });
    rectNino.on('mousemove', e=>ttHover.setLatLng(e.latlng));
    rectNino.on('mouseout', ()=>this.map.closeTooltip(ttHover));
    this.capaDist=null; this.capaMarkers=L.layerGroup().addTo(this.map);
    this.render();
  }
  featuresVisibles(){
    const f=GEO.features;
    if(this.level==='pais') return f;
    if(this.level==='dep') return f.filter(x=>x.properties.NOMBDEP===this.dep);
    if(this.level==='prov') return f.filter(x=>x.properties.NOMBDEP===this.dep&&x.properties.NOMBPROV===this.prov);
    return f.filter(x=>x.properties.IDDIST===this.dist);
  }
  render(){
    if(this.capaDist){ this.map.removeLayer(this.capaDist); }
    this.capaMarkers.clearLayers();
    const self=this;
    const feats=this.featuresVisibles();
    // capa de distritos
    this.capaDist=L.geoJSON({type:'FeatureCollection',features:feats},{
      pane:'overlayPane',
      style:f=>{
        const d=byU[f.properties.IDDIST];
        const sc=d? d[self.col] : null;
        return {color:'#ffffff', weight:0.5, fillColor:scoreColor(sc),
                fillOpacity:(sc?0.9:0.55)};
      },
      onEachFeature:(f,layer)=>{
        const p=f.properties; const d=byU[p.IDDIST];
        const sc=d? d[self.col] : null;
        let tip='<b>'+p.NOMBDIST+'</b><br>'+p.NOMBPROV+', '+p.NOMBDEP+
                '<br>Score '+self.hz.label+': <b>'+(sc??'s/d')+'</b>';
        if(d){
          const acts=activosDeDistrito(d);
          if(acts.length){
            const tot=acts.reduce((s,a)=>s+a.plata,0);
            tip+='<br>Plata en activos: '+fmtMonto(tot)+' ('+acts.length+' activo(s))';
          }
        }
        layer.bindTooltip(tip,{className:'tt',sticky:true});
        layer.on('click',()=>self.onClickDistrito(p));
        layer.on('mouseover',()=>layer.setStyle({weight:1.6,color:AFP.cyan}));
        layer.on('mouseout',()=>layer.setStyle({weight:0.5,color:'#fff'}));
      }
    }).addTo(this.map);
    // activos (solo desde capa departamental hacia abajo)
    if(this.level!=='pais') this.dibujarActivos(feats);
    // encuadre
    if(this.level==='pais'){ this.map.fitBounds(PERU_BOUNDS); }
    else {
      const bb=this.bboxActual();
      if(bb) this.map.fitBounds(bb,{padding:[12,12]});
    }
    this.syncUI();
  }
  bboxActual(){
    if(this.level==='dep') return BBOX_DEP[this.dep];
    if(this.level==='prov') return BBOX_PROV[this.dep+'||'+this.prov];
    if(this.level==='dist') return BBOX_DIST[this.dist];
    return null;
  }
  dibujarActivos(feats){
    const self=this;
    const visU=new Set(feats.map(f=>f.properties.IDDIST));
    // por distrito visible: recolectar activos, ordenar por plata, spiderfy
    visU.forEach(u=>{
      const d=byU[u]; if(!d) return;
      let acts=activosDeDistrito(d);
      if(!acts.length) return;
      const c=CENTROID[u]; if(!c) return;
      acts.sort((a,b)=>b.plata-a.plata);
      const MAXN = self.level==='dist'? 12 : 6;   // límite anti-aglomeración
      const extra = acts.length>MAXN ? acts.length-MAXN : 0;
      const muestra = acts.slice(0,MAXN);
      const maxPlata = Math.max.apply(null, acts.map(a=>a.plata)) || 1;
      // radio de dispersión según capa (más separación al acercar)
      const baseR = self.level==='dist'? 0.012 : (self.level==='prov'? 0.02 : 0.05);
      muestra.forEach((a,i)=>{
        const ang = (i/Math.max(1,muestra.length))*2*Math.PI;
        const rr = muestra.length>1 ? baseR*(0.4+0.6*(i%3)) : 0;
        const lat=c[0]+rr*Math.sin(ang), lng=c[1]+rr*Math.cos(ang);
        const tipo=SEC_ICONO[a.sector]||'circulo';
        const color=SEC_COLOR[a.sector]||'#555';
        // tamaño proporcional a sqrt(plata)
        const px=Math.max(14, Math.min(46, 14+30*Math.sqrt(a.plata/maxPlata)));
        const ic=L.divIcon({className:'act-ic', html:iconSVG(tipo,color,px),
          iconSize:[px,px], iconAnchor:[px/2,px/2]});
        const tip='<b>'+a.activo+'</b><br>'+a.emisor+' · '+(a.sector||'')+
          '<br>'+a.dist+' ('+a.prov+', '+a.dep+')'+
          '<br>Plata estimada: <b>'+fmtMonto(a.plata)+'</b>';
        L.marker([lat,lng],{icon:ic}).bindTooltip(tip,{className:'tt'})
          .addTo(self.capaMarkers);
      });
      if(extra>0){
        const ic=L.divIcon({className:'act-ic',
          html:'<div class="act-badge">+'+extra+'</div>', iconSize:[20,20],
          iconAnchor:[10,10]});
        L.marker([c[0]-baseR*0.7,c[1]+baseR*0.7],{icon:ic})
          .bindTooltip('+'+extra+' activos más en '+d.dist,{className:'tt'})
          .addTo(self.capaMarkers);
      }
    });
  }
  // ---- navegación jerárquica (sin saltar capas) ----
  onClickDistrito(p){
    if(this.level==='pais')      this.goTo('dep',  p.NOMBDEP);
    else if(this.level==='dep')  this.goTo('prov', p.NOMBDEP, p.NOMBPROV);
    else if(this.level==='prov') this.goTo('dist', p.NOMBDEP, p.NOMBPROV, null, p.IDDIST);
    // en 'dist' un click no profundiza más
  }
  goTo(level, dep, prov, _x, distU){
    if(SYNC){ MAPAS.forEach(m=>m._goToSelf(level,dep,prov,distU)); }
    else { this._goToSelf(level,dep,prov,distU); }
  }
  _goToSelf(level, dep, prov, distU){
    this.level=level; this.dep=dep||null; this.prov=prov||null; this.dist=distU||null;
    this.render();
  }
  syncUI(){
    // breadcrumb
    const cb=document.getElementById('crumb_'+this.hz.key);
    let html='<a data-go="pais">Perú</a>';
    if(this.dep) html+=' › <a data-go="dep">'+this.dep+'</a>';
    if(this.prov) html+=' › <a data-go="prov">'+this.prov+'</a>';
    if(this.dist){ const d=byU[this.dist]; html+=' › '+(d?d.dist:this.dist); }
    cb.innerHTML=html;
    cb.querySelectorAll('a').forEach(a=>a.onclick=()=>{
      const lv=a.dataset.go;
      if(lv==='pais') this.goTo('pais');
      else if(lv==='dep') this.goTo('dep',this.dep);
      else if(lv==='prov') this.goTo('prov',this.dep,this.prov);
    });
    // dropdowns
    const selDep=document.getElementById('selDep_'+this.hz.key);
    const selProv=document.getElementById('selProv_'+this.hz.key);
    selDep.value=this.dep||'';
    // poblar provincias del dep activo
    selProv.innerHTML='<option value="">(provincia)</option>';
    if(this.dep && PROV_DE_DEP[this.dep]){
      [...PROV_DE_DEP[this.dep]].sort().forEach(pv=>{
        const o=document.createElement('option'); o.value=pv; o.textContent=pv;
        selProv.appendChild(o);
      });
    }
    selProv.value=this.prov||'';
    selProv.disabled = !this.dep;
  }
}

let MAPAS=[];

// ======================================================================
//  CONSTRUCCIÓN DEL DOM (3 bloques)
// ======================================================================
function construirBloques(){
  const cont=document.getElementById('bloques');
  cont.innerHTML='';
  HZ.forEach(hz=>{
    const blk=document.createElement('div'); blk.className='blk';
    blk.innerHTML=
      '<h3><span class="dot" style="background:'+hz.dot+'"></span>'+hz.label+'</h3>'+
      '<div class="pivwrap"><table class="piv" id="piv_'+hz.key+'"></table></div>'+
      '<div class="mapbar">'+
        '<span class="crumb" id="crumb_'+hz.key+'"></span>'+
        '<select id="selDep_'+hz.key+'" title="Ir a departamento"></select>'+
        '<select id="selProv_'+hz.key+'" title="Ir a provincia" disabled></select>'+
      '</div>'+
      '<div class="mapa" id="mapa_'+hz.key+'"></div>'+
      '<div class="leg">'+
        '<span class="sw"><i style="background:'+SCORE_COLORS[0]+'"></i>1</span>'+
        '<span class="sw"><i style="background:'+SCORE_COLORS[1]+'"></i>2</span>'+
        '<span class="sw"><i style="background:'+SCORE_COLORS[2]+'"></i>3</span>'+
        '<span class="sw"><i style="background:'+SCORE_COLORS[3]+'"></i>4</span>'+
        '<span class="sw"><i style="background:'+SCORE_COLORS[4]+'"></i>5</span>'+
        '<span style="margin-left:6px">Escala de riesgo · íconos = activos (tamaño ∝ plata)</span>'+
      '</div>';
    cont.appendChild(blk);
  });
}

function poblarDropdowns(){
  const deps=Object.keys(PROV_DE_DEP).sort();
  HZ.forEach(hz=>{
    const selDep=document.getElementById('selDep_'+hz.key);
    selDep.innerHTML='<option value="">(departamento)</option>';
    deps.forEach(dp=>{ const o=document.createElement('option'); o.value=dp;
      o.textContent=dp; selDep.appendChild(o); });
    const m=MAPAS.find(x=>x.hz.key===hz.key);
    selDep.onchange=()=>{ const v=selDep.value;
      if(v) m.goTo('dep',v); else m.goTo('pais'); };
    const selProv=document.getElementById('selProv_'+hz.key);
    selProv.onchange=()=>{ const v=selProv.value;
      if(v) m.goTo('prov', m.dep, v); else if(m.dep) m.goTo('dep', m.dep); };
  });
}

function redibujarTablas(){
  HZ.forEach(hz=>{
    document.getElementById('piv_'+hz.key).innerHTML=buildTablaHTML(hz);
  });
}
function redibujarMapas(){ MAPAS.forEach(m=>m.render()); }

// ======================================================================
//  CONTROLES
// ======================================================================
function initControles(){
  document.querySelectorAll('.tab').forEach(t=>t.onclick=()=>{
    document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
    t.classList.add('active'); TAB=t.dataset.tab;
    redibujarTablas(); redibujarMapas();
  });
  const chkM=document.getElementById('chkMonto');
  chkM.onchange=()=>document.body.classList.toggle('ocultar-montos', !chkM.checked);
  document.body.classList.toggle('ocultar-montos', !chkM.checked);
  const chkA=document.getElementById('chkActivo');
  chkA.onchange=()=>document.body.classList.toggle('ocultar-activo', !chkA.checked);
  const chkS=document.getElementById('chkSync');
  chkS.onchange=()=>{ SYNC=chkS.checked; };
  const mPEN=document.getElementById('mPEN'), mUSD=document.getElementById('mUSD');
  mPEN.onclick=()=>{ MONEDA='PEN'; mPEN.classList.add('on'); mUSD.classList.remove('on');
    redibujarTablas(); redibujarMapas(); };
  mUSD.onclick=()=>{ MONEDA='USD'; mUSD.classList.add('on'); mPEN.classList.remove('on');
    redibujarTablas(); redibujarMapas(); };
}

// ======================================================================
//  ARRANQUE
// ======================================================================
document.getElementById('fxLbl').textContent =
  'S/ '+(FX.valor||3.5).toFixed(4)+' / US$'+(FX.fuente==='default'?' (default)':'');
document.getElementById('gen').textContent='Generado '+PAYLOAD.generado;
document.getElementById('footgen').innerHTML=
  'Última actualización automática: <b>'+PAYLOAD.generado+'</b>';

construirBloques();
MAPAS = HZ.map(hz=>new MapaPeligro(hz, document.getElementById('mapa_'+hz.key)));
poblarDropdowns();
redibujarTablas();
initControles();
// asegurar tamaño correcto de mapas tras el layout
setTimeout(()=>MAPAS.forEach(m=>m.map.invalidateSize()),200);
</script>
</body></html>"""