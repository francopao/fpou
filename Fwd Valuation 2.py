

# ==============================================================================
# FORWARD MULTIPLES — DASHBOARD v2   (¿la acción está cara o barata?)
# AFP Integra — Mesa RVL
# ------------------------------------------------------------------------------
# Novedades respecto a v1 (lo que pidió Franco):
#   - Se REORDENA: "Salud del dato" pasa al final de todo.
#   - NUEVO panel 1 (primero): evolución del múltiplo forward en el tiempo, con
#     el usuario eligiendo el múltiplo y hasta 3 empresas para comparar, y con
#     marcadores SUTILES de eventos relevantes (internacionales y locales).
#   - NUEVO panel (antes de "Salud"): múltiplo forward vs una variable FUNDAMENTAL
#     (producción minera / inversión minera / NIM-ROE bancario / despachos de
#     cemento), leídas de los 3 workbooks macro. Eje dual.
#
# Decisiones de datos importantes (verificadas celda por celda):
#   - Los 3 macros tienen los valores CACHEADOS (son output de fórmulas, pero
#     openpyxl data_only=True devuelve el número). Se usan los números, no las
#     fórmulas.
#   - Equivalencias BLINDADAS macro→empresa→ticker en un solo diccionario EQUIV,
#     tolerando los espacios finales de los nombres GICS. Nombres CORTOS en el
#     HTML.
#   - Mineria/Monitor: mensual 2025→2026. Financieras/PERBAN: mensual 2018→2026
#     (col C = banco en celdas combinadas → forward-fill; fila 2 = fechas).
#     Construcción/Data Graf: despachos trimestrales (etiquetas 1Q15…).
#
# Salida: un único HTML autocontenido (Plotly por CDN) con fondo blanco, paleta
# AFP Integra y Calibri. Los 2 paneles nuevos son interactivos vía JS embebido.
# ==============================================================================

# %% SECCIÓN 0 — Imports
import os, json, datetime, re
import numpy as np
import pandas as pd
from openpyxl import load_workbook


# %% SECCIÓN 1 — Rutas, paleta, parámetros
RUTA_FWD = (r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA"
            r"\Docs privados\ZODA\Forward\Copy of forward multiples 2.xlsx")
# Los 3 macros viven en "Proyecto Inputs Macro"
RUTA_MACRO = r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\Docs privados\Proyecto Inputs Macro"
RUTA_MINERIA = os.path.join(RUTA_MACRO, "BD Mineria.xlsm")
RUTA_FINAN   = os.path.join(RUTA_MACRO, "BD Financieras.xlsm")
RUTA_CONSTR  = os.path.join(RUTA_MACRO, "BD Construcción.xlsm")
RUTA_HTML = os.path.join(os.path.dirname(RUTA_FWD), "Forward_dashboard.html")

BRAND = {"navy": "#1E2E6E", "dark": "#002060", "teal": "#00AECB", "yellow": "#E3E829",
         "gray1": "#DCDDDE", "gray2": "#7E8083", "white": "#FFFFFF",
         "green": "#2EA86B", "red": "#D1495B", "amber": "#E0A320"}
SECTOR_COLOR = {"Financials": "#1E2E6E", "Mining": "#E0883E", "Utilities": "#2EA86B",
                "Consumer": "#8C6FD4", "Construction Materials": "#C99A3E",
                "Industrials": "#00AECB", "Health Care": "#D1495B", "Media": "#B07AA1",
                "Other": "#7E8083"}

METRICAS_FWD = ["BEST_PE_RATIO", "BEST_PX_SALES_RATIO", "BEST_CUR_EV_TO_EBITDA",
                "BEST_PX_BPS_RATIO", "BEST_EPS"]
NOMBRE_BONITO = {"BEST_PE_RATIO": "Fwd_PE", "BEST_PX_SALES_RATIO": "Fwd_PS",
                 "BEST_CUR_EV_TO_EBITDA": "Fwd_EV_EBITDA",
                 "BEST_PX_BPS_RATIO": "Fwd_PB", "BEST_EPS": "Fwd_EPS"}
LABEL_MULT = {"Fwd_PE": "P/E fwd", "Fwd_PS": "P/S fwd", "Fwd_EV_EBITDA": "EV/EBITDA fwd",
              "Fwd_PB": "P/B fwd", "Fwd_EPS": "EPS fwd"}
HOJAS_EQUITY = ["Financials", "Cons y Const", "Min y Uti"]
VACIOS = {"#N/A N/A", "NM", "(Invalid Formula Name)", "NA", "#VALUE!",
          "#N/A Field Not Applicable", "#N/A Invalid Security", "", "#REF!"}
MAX_FILAS = 2000
UMBRAL_2025 = pd.Timestamp("2025-01-01")

# --- Equivalencias BLINDADAS: macro → (corto, empresa_gics, bbg, grupo) --------
EQUIV = {
    "CERRO VERDE":  ("Cerro Verde",  "Sociedad Minera Cerro Verde",        "CVERDEC1 PE Equity", "mineria"),
    "BUENAVENTURA": ("Buenaventura", "Compañía de Minas Buenaventura",     "BVN US Equity",      "mineria"),
    "NEXA PERÚ":    ("Nexa Perú",    "Nexa Resources Perú",                "NEXAPEC1 PE Equity", "mineria"),
    "VOLCAN":       ("Volcan",       "Volcan Compañía Minera",             "VOLCABC1 PE Equity", "mineria"),
    "MINSUR":       ("Minsur",       "Minsur",                             "MINSURI1 PE Equity", "mineria"),
    "BCP":          ("Credicorp",    "Credicorp Ltd.",                     "BAP US Equity",      "financiero"),
    "INTERBANK":    ("Intercorp",    "Intercorp Financial Services Inc.",  "IFS PE Equity",      "financiero"),
    "BBVA":         ("BBVA Perú",    "Banco BBVA Perú",                    "BBVAC1 PE Equity",   "financiero"),
    "Unacem":       ("UNACEM",       "UNACEM Corp",                        "UNACEMC1 PE Equity", "cemento"),
    "Pacasmayo":    ("Pacasmayo",    "Cementos Pacasmayo",                 "CPACASC1 PE Equity", "cemento"),
}
BBG_A_CORTO = {v[2]: v[0] for v in EQUIV.values()}   # ticker → nombre corto

# --- Eventos (CURADOS y EDITABLES). Fechas verificadas; ajústalos a tu gusto ---
# scope: 'intl' (internacional) o 'local' (Perú). Se dibujan sutiles y se pueden
# apagar con el switch del panel 1.
EVENTOS = [
    {"date": "2024-10-01", "label": "Irán ataca Israel",      "scope": "intl"},
    {"date": "2025-04-02", "label": "Aranceles Trump",        "scope": "intl"},
    {"date": "2025-06-13", "label": "Escalada Israel–Irán",   "scope": "intl"},   # verificar fecha exacta
    {"date": "2025-10-10", "label": "Vacancia Boluarte",      "scope": "local"},
    {"date": "2026-04-12", "label": "Elecciones 1ª vuelta",   "scope": "local"},
    {"date": "2026-06-07", "label": "Elecciones 2ª vuelta",   "scope": "local"},
]


# %% SECCIÓN 2 — Lectura del Excel forward (parser validado)
def _hoja(wb, nombre):
    ws = wb[nombre]
    return pd.DataFrame(list(ws.iter_rows(min_row=1, max_row=MAX_FILAS, values_only=True)))


def _num(v):
    if isinstance(v, str) and v.strip() in VACIOS:
        return np.nan
    try:
        x = float(v)
    except (TypeError, ValueError):
        return np.nan
    return np.nan if x == 0 else x


def leer_equities(wb):
    paneles = {}
    for hoja in HOJAS_EQUITY:
        df = _hoja(wb, hoja); fmet, ftk = df.iloc[2], df.iloc[4]
        col_fecha = next((c for c in df.columns
                          if isinstance(ftk[c], str) and ftk[c] == fmet[c]), None)
        if col_fecha is None:
            continue
        fechas = pd.to_datetime(df[col_fecha].iloc[5:], errors="coerce").reset_index(drop=True)
        for c in df.columns:
            tk, me = ftk[c], fmet[c]
            if not (isinstance(tk, str) and tk.strip().endswith("Equity")):
                continue
            if me not in METRICAS_FWD:
                continue
            tk = tk.strip(); col = NOMBRE_BONITO[me]
            vals = df[c].iloc[5:].reset_index(drop=True).apply(_num)
            paneles.setdefault(tk, pd.DataFrame(index=fechas.values))[col] = \
                pd.Series(vals.values, index=fechas.values)
    return paneles


def leer_gics(wb):
    df = _hoja(wb, "GICS"); df.columns = [str(c).strip() for c in df.iloc[0]]; df = df.iloc[1:]
    mapa = {}
    for _, f in df.iterrows():
        bbg = str(f.get("BBG", "")).strip()
        if bbg and bbg.lower() != "nan":
            mapa[bbg] = {"Empresa": str(f.get("Empresa", "")).strip(),
                         "GICS": str(f.get("GICS", "")).strip()}
    return mapa


def leer_baskets(wb):
    df = _hoja(wb, "Baskets"); grupos = {"Financieras", "Con, Ind y HC", "Mineria y Utilities"}
    bk = {}
    for _, f in df.iterrows():
        if f.iloc[0] not in grupos:
            continue
        tks = [str(v).strip() for v in f.iloc[1:].tolist() if pd.notna(v) and str(v).strip()]
        if len(tks) >= 2:
            bk[tks[0]] = tks[1:]
    return bk


def corto_de(bbg, empresa):
    """Nombre corto y único para el HTML."""
    if bbg in BBG_A_CORTO:
        return BBG_A_CORTO[bbg]
    e = (empresa or bbg).strip()
    return e[:20]


# %% SECCIÓN 3 — Lectura de los 3 workbooks macro (variables fundamentales)
def _fechas_en_fila(fila, desde=4):
    """Devuelve [(col_idx, Timestamp)] para celdas datetime desde la col 'desde'."""
    out = []
    for j in range(desde, len(fila)):
        v = fila[j]
        if isinstance(v, (datetime.datetime, pd.Timestamp)):
            out.append((j, pd.Timestamp(v)))
    return out


def _serie_desde_fila(fechas_cols, fila):
    """Empareja columnas de fecha con sus valores numéricos."""
    pares = []
    for j, t in fechas_cols:
        v = _num(fila[j]) if j < len(fila) else np.nan
        if not np.isnan(v):
            pares.append([t.strftime("%Y-%m-%d"), round(float(v), 4)])
    return pares


def leer_mineria(ruta):
    """Producción nacional por mineral, producción por compañía×metal e inversión."""
    if not os.path.exists(ruta):
        print("  ⚠️  No encuentro BD Mineria; salto fundamentales mineros.")
        return {}
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Monitor"]
    rows = list(ws.iter_rows(min_row=1, max_row=130, max_col=25, values_only=True))
    fund = {}   # short_company -> {label: serie}

    # localizar filas de fecha (headers) y secciones
    cur_company = None
    seccion = None
    fechas_actual = None
    for i, row in enumerate(rows):
        c2 = row[2] if len(row) > 2 else None   # col C
        c3 = row[3] if len(row) > 3 else None   # col D (metal/unidad)
        # ¿fila de fechas? (tiene datetimes en cols 4+)
        fc = _fechas_en_fila(row)
        if fc and (isinstance(c2, str) or isinstance(c3, str)):
            fechas_actual = fc
        # secciones por encabezado en col C
        if isinstance(c2, str):
            cu = c2.strip().upper()
            if "INVERSIÓN MINERA" in cu:
                seccion = "inversion"; continue
            if "PRODUCCIÓN MINERA" in cu and "TOTAL" not in cu:
                seccion = "prod_emp"; continue
            if "PRODUCCIÓN" in cu:
                seccion = "prod_nac"; continue

        # PRODUCCIÓN NACIONAL: col C = mineral, col D = unidad
        if seccion == "prod_nac" and isinstance(c2, str) and isinstance(c3, str) \
           and fechas_actual and c2.strip().upper() not in ("PRODUCCIÓN TOTAL", "COMPAÑÍA"):
            serie = _serie_desde_fila(fechas_actual, row)
            if serie:
                fund.setdefault("Nacional", {})[f"Prod. {c2.strip().title()} ({c3.strip()})"] = serie

        # PRODUCCIÓN POR COMPAÑÍA: col C = compañía (ffill), col D = metal
        if seccion == "prod_emp":
            if isinstance(c2, str) and c2.strip():
                cur_company = c2.strip()
            metal = c3.strip() if isinstance(c3, str) else None
            if cur_company and metal and fechas_actual:
                if cur_company in EQUIV:
                    short = EQUIV[cur_company][0]
                    serie = _serie_desde_fila(fechas_actual, row)
                    if serie:
                        fund.setdefault(short, {})[f"Prod. {metal.title()} (t)"] = serie

        # INVERSIÓN MINERA: col C = rubro
        if seccion == "inversion" and isinstance(c2, str) and fechas_actual \
           and c2.strip().upper() not in ("RUBRO",):
            serie = _serie_desde_fila(fechas_actual, row)
            if serie:
                fund.setdefault("Nacional", {})[f"Inversión minera · {c2.strip().title()} ($mm)"] = serie

    # Adjuntar "Nacional" (cobre + inversión total) a cada minera para que pueda compararla
    nac = fund.get("Nacional", {})
    extras = {k: v for k, v in nac.items()
              if "Cobre" in k or "Total" in k}
    for macro, (short, *_rest, grupo) in EQUIV.items():
        if grupo == "mineria":
            fund.setdefault(short, {})
            for k, v in extras.items():
                fund[short][k] = v
    return fund


def leer_financieras(ruta):
    """NIM y ROE por banco (col C combinada → ffill; fila 2 = fechas)."""
    if not os.path.exists(ruta):
        print("  ⚠️  No encuentro BD Financieras; salto fundamentales bancarios.")
        return {}
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["PERBAN"]
    rows = list(ws.iter_rows(min_row=1, max_row=4053, max_col=104, values_only=True))
    fechas_cols = _fechas_en_fila(rows[1])   # fila idx1
    bancos = {EQUIV[k][0]: k for k in EQUIV if EQUIV[k][3] == "financiero"}  # short->macro
    macro_short = {EQUIV[k][0]: EQUIV[k][0] for k in EQUIV}  # identidad
    interes = {"NIM": "NIM",
               "Utilidad Neta Anualizada / Patrimonio Promedio": "ROE",
               "Utilidad Neta Anualizada / Activo Promedio": "ROA"}
    fund = {}
    cur = None
    for row in rows:
        c2 = row[2] if len(row) > 2 else None
        c3 = row[3] if len(row) > 3 else None
        if isinstance(c2, str) and c2.strip():
            cur = c2.strip()
        if cur in ("BCP", "INTERBANK", "BBVA") and isinstance(c3, str):
            ind = c3.strip()
            if ind in interes:
                short = EQUIV[cur][0]
                serie = _serie_desde_fila(fechas_cols, row)
                if serie:
                    fund.setdefault(short, {})[f"{interes[ind]} (%)"] = serie
    return fund


def leer_construccion(ruta):
    """Despachos de cemento (trimestral) para Unacem y Pacasmayo."""
    if not os.path.exists(ruta):
        print("  ⚠️  No encuentro BD Construcción; salto fundamentales de cemento.")
        return {}
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Data Graf"]
    rows = list(ws.iter_rows(min_row=1, max_row=60, max_col=200, values_only=True))

    def q_a_fecha(lbl):
        m = re.match(r"([1-4])Q(\d{2})", str(lbl).strip())
        if not m:
            return None
        q, yy = int(m.group(1)), 2000 + int(m.group(2))
        mes = {1: 3, 2: 6, 3: 9, 4: 12}[q]
        return pd.Timestamp(year=yy, month=mes, day=1)

    # fila de etiquetas trimestrales del bloque despachos (busca la que precede a 'Unacem')
    fila_lbls = None
    for i, row in enumerate(rows):
        d = row[3] if len(row) > 3 else None
        if isinstance(d, str) and d.strip() == "Unacem":
            fila_lbls = rows[i - 1]
            filas_emp = {"Unacem": row}
            # Pacasmayo suele estar en la fila siguiente
            if i + 1 < len(rows) and str(rows[i + 1][3]).strip() == "Pacasmayo":
                filas_emp["Pacasmayo"] = rows[i + 1]
            break
    if fila_lbls is None:
        return {}
    cols_fecha = [(j, q_a_fecha(fila_lbls[j])) for j in range(4, len(fila_lbls))
                  if q_a_fecha(fila_lbls[j]) is not None]
    fund = {}
    for macro, row in filas_emp.items():
        short = EQUIV[macro][0]
        serie = []
        for j, t in cols_fecha:
            v = _num(row[j]) if j < len(row) else np.nan
            if not np.isnan(v):
                serie.append([t.strftime("%Y-%m-%d"), round(float(v), 2)])
        if serie:
            fund.setdefault(short, {})["Despachos cemento (miles t)"] = serie
    return fund


# %% SECCIÓN 4 — Métricas analíticas (paneles 2/3/4) y salud (panel 6)
def clase_historia(serie, hoy):
    s = serie.dropna()
    if s.empty:
        return "VACIA"
    ini = s.index.min()
    cob = len(s) / max(len(pd.bdate_range(ini, s.index.max())), 1)
    if ini <= hoy - pd.DateOffset(years=2):
        cl = "OK_2A"
    elif ini <= UMBRAL_2025:
        cl = "OK_2025"
    else:
        cl = "INSUF"
    if cl in ("OK_2A", "OK_2025") and cob < 0.80:
        cl = "HUECOS"
    return cl


def _varN(serie, n):
    s = serie.dropna()
    if len(s) < 5:
        return np.nan
    v1 = s.iloc[-1]; v0 = s.asof(s.index[-1] - pd.DateOffset(months=n))
    if pd.isna(v0) or v0 <= 0 or v1 <= 0:
        return np.nan
    return v1 / v0 - 1.0


def construir_tablas(paneles, gics, baskets):
    hoy = pd.Timestamp(max(p.index.max() for p in paneles.values()))
    cartera = set(baskets.keys())
    filas, salud = [], []
    for tk, panel in paneles.items():
        info = gics.get(tk, {}); empresa = info.get("Empresa", tk) or tk
        sector = info.get("GICS", "Other") or "Other"
        short = corto_de(tk, empresa)
        for col in ["Fwd_PE", "Fwd_PS", "Fwd_EV_EBITDA", "Fwd_PB", "Fwd_EPS"]:
            cl = clase_historia(panel[col], hoy) if col in panel.columns else "VACIA"
            salud.append({"short": short, "sector": sector, "metrica": col,
                          "clase": cl, "en_cartera": tk in cartera})

        def ult(col):
            s = panel[col].dropna() if col in panel.columns else pd.Series(dtype=float)
            return s.iloc[-1] if not s.empty else np.nan
        pe, pb, ps = ult("Fwd_PE"), ult("Fwd_PB"), ult("Fwd_PS")
        roe = pb / pe if (pe and pb and pe > 0 and pb > 0) else np.nan
        filas.append({"short": short, "sector": sector, "en_cartera": tk in cartera,
                      "Fwd_PE": pe, "Fwd_PB": pb, "ROE": roe,
                      "crec": _varN(panel["Fwd_EPS"], 12) if "Fwd_EPS" in panel else np.nan,
                      "rev1": _varN(panel["Fwd_EPS"], 1) if "Fwd_EPS" in panel else np.nan,
                      "rev3": _varN(panel["Fwd_EPS"], 3) if "Fwd_EPS" in panel else np.nan,
                      "rev6": _varN(panel["Fwd_EPS"], 6) if "Fwd_EPS" in panel else np.nan})
    return pd.DataFrame(filas), pd.DataFrame(salud), hoy, cartera


# %% SECCIÓN 5 — Series semanales para el panel 1 (evolución)
def series_para_evolucion(paneles, gics):
    """{short: {Fwd_PE:[[date,val]...], ...}} remuestreado semanal para no pesar."""
    out, meta = {}, {}
    for tk, panel in paneles.items():
        info = gics.get(tk, {}); empresa = info.get("Empresa", tk) or tk
        short = corto_de(tk, empresa); sector = info.get("GICS", "Other") or "Other"
        d = {}
        for col in ["Fwd_PE", "Fwd_PS", "Fwd_EV_EBITDA", "Fwd_PB", "Fwd_EPS"]:
            if col not in panel.columns:
                continue
            s = panel[col].dropna()
            if s.empty:
                continue
            s.index = pd.to_datetime(s.index)
            sw = s.resample("W-FRI").last().dropna()
            d[col] = [[t.strftime("%Y-%m-%d"), round(float(v), 3)] for t, v in sw.items()]
        if d:
            out[short] = d
            meta[short] = {"sector": sector}
    return out, meta


# %% SECCIÓN 6 — Figuras estáticas (paneles 2/3/4/6) con Plotly
def _ols(x, y):
    m = (~np.isnan(x)) & (~np.isnan(y))
    if m.sum() < 5:
        return None
    b1, b0 = np.polyfit(x[m], y[m], 1)
    r2 = 1 - np.sum((y[m] - (b0 + b1 * x[m])) ** 2) / np.sum((y[m] - y[m].mean()) ** 2)
    return b0, b1, r2


def fig_scatter(df, xcol, ycol, xlab, ylab, xr, yr):
    import plotly.graph_objects as go
    d = df.dropna(subset=[xcol, ycol]).copy()
    d = d[d[xcol].between(*xr) & d[ycol].between(*yr)]
    fig = go.Figure()
    for sector, g in d[~d.en_cartera].groupby("sector"):
        fig.add_trace(go.Scatter(x=g[xcol], y=g[ycol], mode="markers", name=sector,
            marker=dict(size=8, color=SECTOR_COLOR.get(sector, BRAND["gray2"]), opacity=0.45),
            text=g["short"], hovertemplate="<b>%{text}</b><br>" + xlab +
            ": %{x:.1%}<br>" + ylab + ": %{y:.2f}<extra></extra>"))
    fit = _ols(d[xcol].values, d[ycol].values)
    if fit:
        b0, b1, r2 = fit; xs = np.linspace(d[xcol].min(), d[xcol].max(), 50)
        fig.add_trace(go.Scatter(x=xs, y=b0 + b1 * xs, mode="lines",
            name=f"Ajuste (R²={r2:.2f})",
            line=dict(color=BRAND["navy"], width=2, dash="dash"), hoverinfo="skip"))
    c = d[d.en_cartera]
    fig.add_trace(go.Scatter(x=c[xcol], y=c[ycol], mode="markers+text", name="En cartera",
        text=c["short"], textposition="top center",
        textfont=dict(size=9, color=BRAND["navy"], family="Calibri"),
        marker=dict(size=14, color=BRAND["teal"], line=dict(width=1.6, color=BRAND["navy"])),
        hovertemplate="<b>%{text}</b> (cartera)<extra></extra>"))
    fig.update_layout(height=500, plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Calibri, sans-serif", size=13, color=BRAND["dark"]),
        margin=dict(l=60, r=25, t=15, b=50),
        legend=dict(orientation="h", y=-0.14, font=dict(size=10)),
        xaxis=dict(title=xlab, range=list(xr), showgrid=False, zeroline=True,
                   zerolinecolor=BRAND["gray1"], tickformat=".0%", ticks="outside",
                   linecolor=BRAND["gray2"]),
        yaxis=dict(title=ylab, range=list(yr), showgrid=False, ticks="outside",
                   linecolor=BRAND["gray2"]))
    return fig


def fig_revisiones(df):
    import plotly.graph_objects as go
    d = df[df.en_cartera].dropna(subset=["rev3"]).sort_values("rev3").copy()
    CAP = 0.50
    d["xp"] = d["rev3"].clip(-CAP, CAP)
    col = [BRAND["green"] if v >= 0 else BRAND["red"] for v in d["rev3"]]
    fig = go.Figure(go.Bar(y=d["short"], x=d["xp"], orientation="h", marker_color=col,
        customdata=d[["rev3", "rev1", "rev6"]].values,
        hovertemplate="<b>%{y}</b><br>3m: %{customdata[0]:.1%}<br>" +
                      "1m: %{customdata[1]:.1%}<br>6m: %{customdata[2]:.1%}<extra></extra>"))
    fig.update_layout(height=max(360, 26 * len(d) + 70), plot_bgcolor="white",
        paper_bgcolor="white", font=dict(family="Calibri, sans-serif", size=12, color=BRAND["dark"]),
        margin=dict(l=10, r=25, t=15, b=45),
        xaxis=dict(title="Revisión EPS fwd · 3m (recortado ±50%)", range=[-0.55, 0.55],
                   tickformat=".0%", showgrid=False, zeroline=True, zerolinecolor=BRAND["gray2"]),
        yaxis=dict(showgrid=False))
    return fig


def fig_salud(salud):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    orden = ["OK_2A", "OK_2025", "HUECOS", "INSUF", "VACIA"]
    colc = {"OK_2A": BRAND["navy"], "OK_2025": BRAND["teal"], "HUECOS": BRAND["yellow"],
            "INSUF": BRAND["amber"], "VACIA": BRAND["gray1"]}
    et = {"OK_2A": "OK ≥2a", "OK_2025": "OK 2025", "HUECOS": "huecos",
          "INSUF": "insuf.", "VACIA": "sin dato"}
    mo = ["Fwd_PE", "Fwd_PS", "Fwd_EV_EBITDA", "Fwd_PB", "Fwd_EPS"]
    fig = make_subplots(rows=1, cols=2, column_widths=[0.42, 0.58],
        subplot_titles=("Universo: cobertura por métrica", "Cartera: calidad por métrica"),
        horizontal_spacing=0.12)
    piv = salud.groupby(["metrica", "clase"]).size().unstack(fill_value=0).reindex(mo)
    for cl in orden:
        if cl in piv.columns:
            fig.add_trace(go.Bar(y=[m.replace("Fwd_", "") for m in piv.index], x=piv[cl],
                name=et[cl], orientation="h", marker_color=colc[cl],
                hovertemplate=et[cl] + ": %{x}<extra></extra>"), row=1, col=1)
    cart = salud[salud.en_cartera].copy()
    zmap = {"OK_2A": 4, "OK_2025": 3, "HUECOS": 2, "INSUF": 1, "VACIA": 0}
    mat = cart.pivot_table(index="short", columns="metrica", values="clase",
                           aggfunc="first").reindex(columns=mo)
    Z = mat.apply(lambda c: c.map(lambda x: zmap.get(x, 0)))
    fig.add_trace(go.Heatmap(z=Z.values, x=[m.replace("Fwd_", "") for m in mo], y=Z.index,
        zmin=0, zmax=4, colorscale=[[0, BRAND["gray1"]], [.25, BRAND["amber"]],
        [.5, BRAND["yellow"]], [.75, BRAND["teal"]], [1, BRAND["navy"]]],
        showscale=False, xgap=2, ygap=2, text=mat.values,
        hovertemplate="%{y} · %{x}: %{text}<extra></extra>"), row=1, col=2)
    fig.update_layout(height=620, barmode="stack", plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Calibri, sans-serif", size=12, color=BRAND["dark"]),
        margin=dict(l=10, r=20, t=40, b=40),
        legend=dict(orientation="h", y=-0.08, font=dict(size=10)))
    fig.update_xaxes(showgrid=False, ticks="outside")
    fig.update_yaxes(showgrid=False, autorange="reversed", row=1, col=2)
    return fig


# %% SECCIÓN 7 — Ensamblado del HTML (con JS para los 2 paneles interactivos)
def construir_html(figs_div, series_ev, meta_ev, fund, hoy, n_univ, n_cart):
    import plotly.io as pio
    div = lambda f: pio.to_html(f, full_html=False, include_plotlyjs=False,
                                config={"displayModeBar": False})

    # Listas para selectores
    empresas_ev = sorted(series_ev.keys(), key=lambda s: (meta_ev[s]["sector"], s))
    empresas_fund = [EQUIV[k][0] for k in EQUIV if EQUIV[k][0] in fund]
    default_ev = [c for c in ["Credicorp", "Cerro Verde", "UNACEM"] if c in series_ev][:3]

    PALETA_LINEAS = [BRAND["navy"], BRAND["teal"], BRAND["amber"]]

    JS_DATA = {
        "series": series_ev, "meta": meta_ev, "fund": fund,
        "eventos": EVENTOS, "empresasEv": empresas_ev, "empresasFund": empresas_fund,
        "defaultEv": default_ev, "labelMult": LABEL_MULT,
        "paleta": PALETA_LINEAS, "brand": BRAND,
    }

    opts_mult = "".join(f'<option value="{k}">{v}</option>' for k, v in LABEL_MULT.items())
    opts_emp_ev = "".join(
        f'<option value="{c}"{" selected" if c in default_ev else ""}>{c}</option>'
        for c in empresas_ev)
    opts_emp_fund = "".join(f'<option value="{c}">{c}</option>' for c in empresas_fund)

    plantilla = r"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Forward Multiples — ¿Cara o barata? · AFP Integra</title>
<script src="https://cdn.plot.ly/plotly-2.30.0.min.js"></script>
<style>
 *{box-sizing:border-box} body{margin:0;background:#fff;color:#002060;font-family:Calibri,'Segoe UI',sans-serif}
 .header{background:#1E2E6E;color:#fff;padding:18px 30px;display:flex;justify-content:space-between;align-items:center}
 .header h1{margin:0;font-size:20px;font-weight:600} .header .sub{font-size:12px;color:#00AECB;margin-top:3px}
 .accent{height:4px;background:#E3E829} .meta{text-align:right;font-size:11px;color:rgba(255,255,255,.75)}
 .wrap{max-width:1100px;margin:0 auto;padding:22px 18px 60px}
 .panel{border:1px solid #DCDDDE;border-radius:8px;margin-bottom:22px;background:#fff;overflow:hidden}
 .panel-head{padding:14px 18px 6px;border-bottom:1px solid #DCDDDE}
 .panel-title{font-size:15px;font-weight:700;color:#1E2E6E;display:block}
 .panel-desc{font-size:12px;color:#7E8083;line-height:1.5;display:block;margin-top:4px}
 .panel-body{padding:8px 8px 0} .panel-read{font-size:11.5px;background:#f6f7fb;border-top:1px solid #DCDDDE;padding:10px 18px;line-height:1.5}
 .panel-read b{color:#1E2E6E}
 .ctrls{display:flex;flex-wrap:wrap;gap:14px;align-items:flex-end;padding:12px 18px 4px}
 .ctrl{display:flex;flex-direction:column;gap:3px} .ctrl label{font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:#1E2E6E;font-weight:600}
 select{font-family:Calibri,sans-serif;font-size:13px;border:1px solid #DCDDDE;border-radius:5px;padding:5px 8px;color:#002060;background:#fff}
 select[multiple]{min-width:190px} .switch{display:flex;align-items:center;gap:6px;font-size:12px;color:#002060;cursor:pointer}
 .hint{font-size:10.5px;color:#7E8083} .foot{text-align:center;font-size:10.5px;color:#7E8083;margin-top:10px}
</style></head><body>
 <div class="header"><div><h1>Forward Multiples — ¿la acción está cara o barata?</h1>
   <div class="sub">Mesa Renta Variable Local · valoración relativa con estimados Bloomberg</div></div>
   <div class="meta">Fecha de datos: __HOY__<br>Universo: __NUNIV__ equities · Cartera: __NCART__ nombres</div></div>
 <div class="accent"></div>
 <div class="wrap">

  <section class="panel">
   <div class="panel-head"><span class="panel-title">01 · Evolución del múltiplo forward</span>
    <span class="panel-desc">Elige el múltiplo y hasta 3 empresas para comparar su trayectoria. Las líneas verticales sutiles marcan hitos (internacionales y locales) que mueven la valoración.</span></div>
   <div class="ctrls">
     <div class="ctrl"><label>Múltiplo</label><select id="ev-mult">__OPTSMULT__</select></div>
     <div class="ctrl"><label>Empresas (máx. 3)</label><select id="ev-emp" multiple size="6">__OPTSEMPEV__</select></div>
     <label class="switch"><input type="checkbox" id="ev-eventos" checked> Mostrar eventos</label>
     <span class="hint">Ctrl/⌘+clic para multiseleccionar</span>
   </div>
   <div class="panel-body"><div id="chart-ev" style="height:480px"></div></div>
   <div class="panel-read"><b>Cómo leerlo.</b> Un múltiplo que sube sostenido = el mercado está pagando cada vez más caro ese nombre; compáralo contra sus pares y contra los hitos marcados para entender si el movimiento es idiosincrático o de contexto.</div>
  </section>

  __FIG_PE_CREC__
  __FIG_PB_ROE__
  __FIG_REV__

  <section class="panel">
   <div class="panel-head"><span class="panel-title">05 · Múltiplo forward vs variable fundamental</span>
    <span class="panel-desc">¿El múltiplo acompaña a su motor real? Cruza el múltiplo forward (eje izquierdo) con un fundamental de la empresa (eje derecho): producción/inversión minera, NIM o ROE bancario, o despachos de cemento.</span></div>
   <div class="ctrls">
     <div class="ctrl"><label>Empresa</label><select id="fu-emp">__OPTSEMPFUND__</select></div>
     <div class="ctrl"><label>Múltiplo</label><select id="fu-mult">__OPTSMULT2__</select></div>
     <div class="ctrl"><label>Fundamental</label><select id="fu-var"></select></div>
   </div>
   <div class="panel-body"><div id="chart-fund" style="height:470px"></div></div>
   <div class="panel-read"><b>Cómo leerlo.</b> Si el múltiplo sube mientras el fundamental cae (o viceversa), el mercado puede estar adelantándose o equivocándose: una divergencia sostenida es una señal de revisar la tesis.</div>
  </section>

  __FIG_SALUD__

  <div class="foot">AFP Integra (Grupo Sura) · Múltiplos forward (BEST_*) de Bloomberg · Fundamentales: Proyecto Inputs Macro · Caro/barato evaluado contra fundamentales, no en aislado.</div>
 </div>

<script>
const D = __JSDATA__;
const fmt = {tickfont:{family:'Calibri',size:11,color:'#002060'},titlefont:{family:'Calibri',size:12,color:'#1E2E6E'}};
function baseLayout(){return {plot_bgcolor:'#fff',paper_bgcolor:'#fff',
  font:{family:'Calibri,sans-serif',size:12,color:'#002060'},
  margin:{l:60,r:60,t:18,b:40},showlegend:true,
  legend:{orientation:'h',y:-0.16,font:{size:11}},hovermode:'x unified'};}

/* ---------- Panel 1: evolución ---------- */
function eventosShapes(){
  if(!document.getElementById('ev-eventos').checked) return {shapes:[],anns:[]};
  const shapes=[],anns=[];
  D.eventos.forEach(e=>{
    const col = e.scope==='local' ? '#1E2E6E' : '#7E8083';
    shapes.push({type:'line',x0:e.date,x1:e.date,yref:'paper',y0:0,y1:1,
      line:{color:col,width:1,dash:'dot'},opacity:0.45});
    anns.push({x:e.date,y:1,yref:'paper',yanchor:'bottom',text:e.label,showarrow:false,
      textangle:-90,font:{size:8.5,color:col},xanchor:'left',xshift:2,opacity:0.7});
  });
  return {shapes,anns};
}
function renderEv(){
  const mult=document.getElementById('ev-mult').value;
  let sel=Array.from(document.getElementById('ev-emp').selectedOptions).map(o=>o.value);
  if(sel.length>3){ sel=sel.slice(0,3);
    Array.from(document.getElementById('ev-emp').options).forEach(o=>{o.selected=sel.includes(o.value);}); }
  const traces=sel.map((c,i)=>{
    const s=(D.series[c]||{})[mult]||[];
    return {x:s.map(p=>p[0]),y:s.map(p=>p[1]),mode:'lines',name:c,
      line:{color:D.paleta[i%D.paleta.length],width:2}};
  });
  const ev=eventosShapes();
  const lay=baseLayout();
  lay.shapes=ev.shapes; lay.annotations=ev.anns;
  lay.xaxis=Object.assign({showgrid:false,linecolor:'#7E8083'},fmt);
  lay.yaxis=Object.assign({title:D.labelMult[mult],showgrid:false,linecolor:'#7E8083'},fmt);
  Plotly.react('chart-ev',traces,lay,{displayModeBar:false});
}

/* ---------- Panel 5: múltiplo vs fundamental ---------- */
function poblarFund(){
  const emp=document.getElementById('fu-emp').value;
  const sel=document.getElementById('fu-var');
  sel.innerHTML='';
  Object.keys(D.fund[emp]||{}).forEach(k=>{
    const o=document.createElement('option'); o.value=k; o.textContent=k; sel.appendChild(o);
  });
  renderFund();
}
function renderFund(){
  const emp=document.getElementById('fu-emp').value;
  const mult=document.getElementById('fu-mult').value;
  const fvar=document.getElementById('fu-var').value;
  const sm=(D.series[emp]||{})[mult]||[];
  const sf=((D.fund[emp]||{})[fvar])||[];
  const t1={x:sm.map(p=>p[0]),y:sm.map(p=>p[1]),mode:'lines',name:D.labelMult[mult],
    line:{color:'#1E2E6E',width:2},yaxis:'y'};
  const t2={x:sf.map(p=>p[0]),y:sf.map(p=>p[1]),mode:'lines+markers',name:fvar,
    line:{color:'#00AECB',width:2},marker:{size:5},yaxis:'y2'};
  const lay=baseLayout();
  lay.xaxis=Object.assign({showgrid:false,linecolor:'#7E8083'},fmt);
  lay.yaxis=Object.assign({title:D.labelMult[mult],showgrid:false,linecolor:'#1E2E6E'},fmt);
  lay.yaxis2=Object.assign({title:fvar,overlaying:'y',side:'right',showgrid:false,
    linecolor:'#00AECB'},fmt);
  Plotly.react('chart-fund',[t1,t2],lay,{displayModeBar:false});
}

document.getElementById('ev-mult').addEventListener('change',renderEv);
document.getElementById('ev-emp').addEventListener('change',renderEv);
document.getElementById('ev-eventos').addEventListener('change',renderEv);
document.getElementById('fu-emp').addEventListener('change',poblarFund);
document.getElementById('fu-mult').addEventListener('change',renderFund);
document.getElementById('fu-var').addEventListener('change',renderFund);
renderEv(); poblarFund();
</script>
</body></html>"""

    reemplazos = {
        "__HOY__": str(hoy.date()), "__NUNIV__": str(n_univ), "__NCART__": str(n_cart),
        "__OPTSMULT__": opts_mult, "__OPTSMULT2__": opts_mult,
        "__OPTSEMPEV__": opts_emp_ev, "__OPTSEMPFUND__": opts_emp_fund,
        "__FIG_PE_CREC__": figs_div["pe_crec"], "__FIG_PB_ROE__": figs_div["pb_roe"],
        "__FIG_REV__": figs_div["rev"], "__FIG_SALUD__": figs_div["salud"],
        "__JSDATA__": json.dumps(JS_DATA, ensure_ascii=False),
    }
    html = plantilla
    for k, v in reemplazos.items():
        html = html.replace(k, v)
    return html


def _wrap_panel(num, titulo, desc, leer, fig_div):
    return f"""<section class="panel"><div class="panel-head">
 <span class="panel-title">{num} · {titulo}</span><span class="panel-desc">{desc}</span></div>
 <div class="panel-body">{fig_div}</div>
 <div class="panel-read"><b>Cómo leerlo.</b> {leer}</div></section>"""


# %% SECCIÓN 8 — MAIN
def main():
    print("=" * 70); print("  FORWARD DASHBOARD v2"); print("=" * 70)
    wb = load_workbook(RUTA_FWD, read_only=True, data_only=True)
    print("[1/6] Excel forward..."); paneles = leer_equities(wb)
    gics = leer_gics(wb); baskets = leer_baskets(wb)
    print(f"  -> {len(paneles)} equities, {len(baskets)} en cartera")

    print("[2/6] Workbooks macro (fundamentales)...")
    fund = {}
    for sub in (leer_mineria(RUTA_MINERIA), leer_financieras(RUTA_FINAN), leer_construccion(RUTA_CONSTR)):
        for emp, series in sub.items():
            fund.setdefault(emp, {}).update(series)
    fund.pop("Nacional", None)   # "Nacional" ya se repartió a las mineras
    print(f"  -> fundamentales para: {sorted(fund.keys())}")

    print("[3/6] Tablas analíticas...")
    df, salud, hoy, cartera = construir_tablas(paneles, gics, baskets)

    print("[4/6] Series semanales (evolución)...")
    series_ev, meta_ev = series_para_evolucion(paneles, gics)

    print("[5/6] Figuras estáticas...")
    import plotly.io as pio
    dv = lambda f: pio.to_html(f, full_html=False, include_plotlyjs=False,
                               config={"displayModeBar": False})
    figs_div = {
        "pe_crec": _wrap_panel("02", "P/E forward vs crecimiento de EPS",
            "El múltiplo alto solo se justifica con más crecimiento esperado (Yee 2004).",
            "Puntos por encima de la recta = caros para su crecimiento. Teal con borde = tu cartera.",
            dv(fig_scatter(df, "crec", "Fwd_PE", "Crecimiento EPS fwd (12m)", "P/E forward", (-0.5, 1.2), (0, 50)))),
        "pb_roe": _wrap_panel("03", "P/B forward vs ROE",
            "La recta clásica P/B–ROE (Damodaran). ROE = Fwd_PB / Fwd_PE, sale de tus múltiplos.",
            "Puntos por encima de la recta = caros para su rentabilidad. Muy informativo en financieras.",
            dv(fig_scatter(df, "ROE", "Fwd_PB", "ROE implícito (Fwd_PB / Fwd_PE)", "P/B forward", (0, 0.55), (0, 10)))),
        "rev": _wrap_panel("04", "Revisiones del EPS forward",
            "¿El motor del múltiplo (utilidades esperadas) mejora o empeora? (Earnings revisions).",
            "Verde = analistas subiendo EPS; rojo = bajando. Evita el value trap.",
            dv(fig_revisiones(df))),
        "salud": _wrap_panel("06", "Salud del dato",
            "¿En qué confiar? El forward P/B casi no existe fuera de financieras.",
            "Azul/teal = dato confiable; amarillo/ámbar = parcial; gris = sin dato.",
            dv(fig_salud(salud))),
    }

    print("[6/6] Ensamblando HTML...")
    html = construir_html(figs_div, series_ev, meta_ev, fund, hoy, len(paneles), len(cartera))
    with open(RUTA_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n  ✓ Dashboard: {RUTA_HTML}  ({len(html)/1e6:.2f} MB)")
    return df, fund


if __name__ == "__main__":
    DF, FUND = main()