
"""
ENFEN PDF Extractor → Excel  (v3)
- Lee PDFs en la carpeta ENFEN
- Si el Excel ya existe, detecta qué filas están registradas y solo procesa PDFs nuevos
- Columna Fecha (date) con formato DD/MM/YYYY en Excel
- Extractores robustos: maneja superíndices, formato antiguo/nuevo, comunicados "No Activo"
- Advertencias en consola y hoja secundaria del Excel
"""

import os, re
import pdfplumber
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
PDF_DIR  = r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\ESG\Riesgos Fisicos\FEN\ENFEN"
XLSX_OUT = os.path.join(PDF_DIR, "ENFEN_Comunicados_Resumen.xlsx")

MESES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}
MESES_INV = {v: k for k, v in MESES.items()}

HEADERS = [
    "Fecha",
    "Año", "Mes Nro", "Mes", "N° Comunicado",
    "Estado de Alerta",
    "Magnitud El Niño Costero",
    "Magnitud Pacífico Central",
    "Pronóstico Lluvias",
    "Pronóstico Hidrológico",
    "Resumen Ejecutivo",
]
COL_W = [14, 6, 8, 12, 14, 35, 28, 28, 32, 32, 80]

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def parse_filename(fname):
    m = re.search(r'_N(\d+)_(\w+)_(\d{4})', fname, re.IGNORECASE)
    if m:
        mes_str = m.group(2).lower()
        return int(m.group(3)), MESES.get(mes_str, 0), mes_str, int(m.group(1))
    m = re.search(r'Comunicado_(\w+)_(\d{4})', fname, re.IGNORECASE)
    if m:
        mes_str = m.group(1).lower()
        mes_num = MESES.get(mes_str, 0)
        return int(m.group(2)), mes_num, mes_str, mes_num
    return None, None, None, None

def leer_pdf(path):
    try:
        with pdfplumber.open(path) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages[:3])
    except Exception:
        return None

# ─── LIMPIEZA DE SUPERÍNDICES ─────────────────────────────────────────────────

def limpiar_superindices(texto):
    """
    pdfplumber extrae superíndices como dígitos en línea propia.
    Casos observados:
      - Dígito solo ANTES de la línea 'Estado de sistema de alerta'
      - Dígito solo DESPUÉS de fecha/código y antes de 'Estado'
      - Dígito pegado AL FINAL del valor de alerta en esa misma línea
    """
    # Dígito en línea sola inmediatamente antes de "Estado"
    texto = re.sub(r'\n\d{1,2}\n(Estado)', r'\n\1', texto)
    # Dígito en línea sola entre una línea de fecha/callao y "Estado"
    texto = re.sub(r'(Callao[^\n]*\n)\d{1,2}\n(Estado)', r'\1\2', texto, flags=re.IGNORECASE)
    # Superíndice pegado al final del valor en la misma línea de alerta
    texto = re.sub(r'(alerta\s*[:\-][^\n]+?)\d{1,2}(\s*\n)', r'\1\2', texto, flags=re.IGNORECASE)
    # Superíndice pegado tras palabra clave del valor (p.ej. "Costero1\n")
    texto = re.sub(r'(Costero|Activo|Vigilancia|Preparaci[óo]n|Niña)\d{1,2}(\s*\n)',
                   r'\1\2', texto, flags=re.IGNORECASE)
    return texto

# ─── EXTRACTORES ─────────────────────────────────────────────────────────────

def extraer_estado_alerta(texto):
    t = limpiar_superindices(texto)

    # 1. Header directo: "Estado d[e/el/de] sistema/del sistema de alerta: X"
    m = re.search(
        r'Estado d[eo][ls]?\s+sistema de alerta\s*[:\-]\s*(.+?)(?:\n|$)',
        t, re.IGNORECASE
    )
    if m:
        val = re.sub(r'\s*\d+\s*$', '', m.group(1)).strip().strip('"').strip('\u201c\u201d')
        if val:
            return val

    # 2. En cuerpo: "estado de sistema de alerta a X"
    m2 = re.search(
        r'estado de sistema de alerta a\s*["\u201c\u201d]?([\w\s\u00c0-\u024f]+?)["\u201c\u201d]?[\.\n]',
        t, re.IGNORECASE
    )
    if m2:
        return m2.group(1).strip()

    # 3. "mantiene el estado de X" (formatos con comillas tipográficas o normales)
    m3 = re.search(
        r'mantiene el estado de\s*[\u201c\u201d"\']([\w\s\u00c0-\u024f]+?)[\u201c\u201d"\']\s*[,\.\n]?',
        t, re.IGNORECASE
    )
    if m3:
        return m3.group(1).strip()

    # 4. "cambia el estado de sistema de alerta a X"
    m4 = re.search(
        r'cambia el estado de sistema de alerta a\s*["\u201c\u201d]?([\w\s]+?)["\u201c\u201d]?[\.\n]',
        t, re.IGNORECASE
    )
    if m4:
        return m4.group(1).strip()

    return "No identificado"

def extraer_magnitud_costero(texto):
    # Comunicado "No Activo" con evento finalizado
    if re.search(r'El Ni[ñn]o Costero ha finalizado', texto, re.IGNORECASE):
        return "No aplica (evento finalizado)"

    # Comunicado "No Activo" sin evento activo → buscar condición actual Niño 1+2
    if re.search(r'No Activo', texto, re.IGNORECASE):
        m = re.search(
            r'condiciones?\s+(neutras?|cálidas?\s*d[eé]bil[es]*|cálidas?\s*moderad[ao]s?'
            r'|frías?\s*\w*|normales?)'
            r'(?:\s+en\s+la\s+regi[óo]n\s+Ni[ñn]o\s*1\+2'
            r'|\s+en\s+la\s+costa'
            r'|\s+frente\s+a\s+la\s+costa)?',
            texto, re.IGNORECASE
        )
        if m:
            return m.group(1).strip().capitalize()
        return "No aplica (No Activo)"

    patrones = [
        r'El Ni[ñn]o Costero.*?magnitud\s+([\w\s]+?)(?:\s+por lo pronto|\.|,|\n)',
        r'El Ni[ñn]o costero.*?magnitud\s+([\w\s]+?)(?:\s+por lo pronto|\.|,|\n)',
        r'magnitud\s+(d[eé]bil|moderada?|fuerte|extraordinaria?|muy fuerte)\b',
        r'sea\s+(d[eé]bil|moderado|moderada|fuerte|extraordinario)',
        r'magnitud entre\s+([\w\s]+?)(?:\.|,|\n)',
        r'condiciones\s+(neutras?)\s+en la regi[óo]n\s+Ni[ñn]o 1\+2',
    ]
    for pat in patrones:
        m = re.search(pat, texto, re.IGNORECASE | re.DOTALL)
        if m:
            val = re.sub(r'\s*(en los|a partir|hasta|durante|que se)\b.*', '',
                         m.group(1).strip(), flags=re.IGNORECASE)
            if len(val) < 40:
                return val.strip()
    return "No identificado"

def extraer_magnitud_pacifico(texto):
    # Comunicado "No Activo"
    if re.search(r'No Activo', texto, re.IGNORECASE):
        # Buscar condición Pacífico Central explícita
        m = re.search(
            r'Pac[íi]fico [Cc]entral[^.]{0,80}condiciones?\s+(neutras?|cálidas?\s*\w*|frías?\s*\w*)',
            texto, re.IGNORECASE | re.DOTALL
        )
        if m:
            return m.group(1).strip().capitalize()
        m2 = re.search(
            r'condiciones neutras.*?Pac[íi]fico [Cc]entral',
            texto, re.IGNORECASE | re.DOTALL
        )
        if m2:
            return "Neutras"
        # Perspectivas con condiciones para Pacífico
        m3 = re.search(
            r'Pac[íi]fico central[^\n]*condiciones?\s+([\w\s]+?)(?:de julio|hasta|\.|,|\n)',
            texto, re.IGNORECASE
        )
        if m3:
            val = m3.group(1).strip()
            if len(val) < 30:
                return val.capitalize()
        return "No aplica (No Activo)"

    patrones = [
        r'Pac[íi]fico central.*?(?:magnitud|condici[óo]n)\s+([\w\s]+?)(?:\.|,|\n)',
        r'regi[óo]n Ni[ñn]o 3\.4.*?(?:magnitud|condici[óo]n)\s+([\w\s]+?)(?:\.|,|\n)',
        r'El Ni[ñn]o en el Pac[íi]fico central.*?sea\s+([\w\s]+?)(?:\.|,|\n)',
        r'condiciones?\s+(neutras?|cálidas?\s*\w*|frías?\s*\w*)\s+(?:en el\s+)?Pac[íi]fico',
    ]
    for pat in patrones:
        m = re.search(pat, texto, re.IGNORECASE | re.DOTALL)
        if m:
            val = re.sub(r'\s*(en los|a partir|hasta|durante)\b.*', '',
                         m.group(1).strip(), flags=re.IGNORECASE)
            if 2 < len(val) < 50:
                return val.strip()
    if re.search(r'condici[óo]n neutra', texto, re.IGNORECASE):
        return "Neutra"
    return "No identificado"

def extraer_pronostico_lluvias(texto):
    # Comunicado de invierno/No Activo que explícitamente descarta lluvias
    if re.search(
        r'no se espera(?:n)? la ocurrencia de lluvias|'
        r'no se esperan? lluvias',
        texto, re.IGNORECASE
    ):
        return "Sin lluvias esperadas (estacionalidad)"

    # Sin sección de lluvias en comunicado No Activo de invierno
    if re.search(r'No Activo', texto, re.IGNORECASE):
        if not re.search(r'lluvi', texto, re.IGNORECASE):
            return "No aplica (período seco / No Activo)"

    patrones = [
        r'pron[óo]stico vigente.*?indica lluvias de\s+([\w\s]+?)(?:en|principalmente|\.|,|\n)',
        r'lluvias de\s+([\w\s]+?)\s+en la costa',
        r'lluvias\s+(normal a superior|por encima|deficitarias?|d[eé]biles?|moderadas?|muy fuertes?|superiores?|normales?)',
        r'ocurrencia de lluvias\s+([\w\s]+?)(?:\s+en|\s+para|\.|,)',
        r'lluvias?\s+([\w\s]{3,30}?)\s+en\s+(?:la costa|Tumbes|Piura|la sierra)',
        r'lluvias\s+entre\s+([\w\s]+?)(?:\.|,|\n)',
    ]
    for pat in patrones:
        m = re.search(pat, texto, re.IGNORECASE | re.DOTALL)
        if m:
            val = m.group(1).strip()
            if len(val) < 60:
                return val
    return "No identificado"

def extraer_hidrologico(texto):
    patrones = [
        r'pron[óo]stico hidrol[óo]gico.*?predominen caudales\s+([\w\s]+?)(?:\.|,|\n)',
        r'caudales.*?(rango[\w\s]*normal[\w\s]*?)(?:\.|,|\n|;)',
        r'caudales.*?(por encima de los promedios[\w\s]*?)(?:\.|,|\n|;)',
        r'caudales.*?(por debajo[\w\s]*?)(?:\.|,|\n|;)',
        r'caudales.*?(?:dentro del\s+)?([\w\s]+rango[\w\s]+)(?:\.|,|\n)',
        r'hidrol[óo]gico.*?rango\s+([\w\s]+?)(?:\.|,|\n)',
    ]
    for pat in patrones:
        m = re.search(pat, texto, re.IGNORECASE | re.DOTALL)
        if m:
            val = re.sub(r'\s*(siendo esta|para la zona|principalmente|a excepci[óo]n)\b.*',
                         '', m.group(1).strip(), flags=re.IGNORECASE)
            val = val.strip()
            if 3 < len(val) < 100:
                return val

    # Fallback para No Activo: caudales en cuerpo del texto
    m2 = re.search(
        r'caudales de los r[íi]os[^\n.;]{5,120}(?:por encima|por debajo|dentro del rango|'
        r'decreci|aument|normal)',
        texto, re.IGNORECASE
    )
    if m2:
        val = m2.group(0).strip()
        if len(val) < 120:
            return val[:100]

    if re.search(r'No Activo', texto, re.IGNORECASE):
        return "No aplica (No Activo)"

    return "No identificado"

def extraer_resumen(texto):
    m = re.search(
        r'RESUMEN EJECUTIVO\s*\n([\s\S]{100,1500}?)(?=\n[A-Z\u00C0-\u024F]{4}|\Z)',
        texto, re.IGNORECASE
    )
    if m:
        bloque = re.sub(r'\n\d+\s', ' ', m.group(1).strip())
        return bloque[:1000]
    # Formato antiguo / No Activo: párrafo principal después del header de alerta
    t = limpiar_superindices(texto)
    m2 = re.search(
        r'Estado d[eo][ls]?\s+sistema de alerta[^\n]+\n([\s\S]{100,800}?)(?=\n[A-Z\u00C0-\u024F]{4}|\Z)',
        t, re.IGNORECASE
    )
    if m2:
        return m2.group(1).strip()[:1000]
    return "Resumen no encontrado"

# ─── ESTILOS ──────────────────────────────────────────────────────────────────

def make_styles():
    thin = Side(style="thin", color="B0B0B0")
    return {
        "font_header":  Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "font_normal":  Font(name="Arial", size=9),
        "font_bold":    Font(name="Arial", bold=True, size=9),
        "fill_header":  PatternFill("solid", fgColor="1F4E79"),
        "fill_even":    PatternFill("solid", fgColor="D6E4F0"),
        "fill_odd":     PatternFill("solid", fgColor="EBF3FB"),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left":   Alignment(horizontal="left",   vertical="top",    wrap_text=True),
        "border":       Border(left=thin, right=thin, top=thin, bottom=thin),
    }

def apply_header_row(ws, st):
    for col_i, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws.cell(row=1, column=col_i, value=h)
        c.font = st["font_header"]
        c.fill = st["fill_header"]
        c.border = st["border"]
        c.alignment = st["align_center"]
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def apply_data_row(ws, row_i, rec, st):
    fill = st["fill_even"] if row_i % 2 == 0 else st["fill_odd"]
    vals = [
        rec["fecha"],
        rec["año"], rec["mes_num"], rec["mes_str"].capitalize(), rec["n_comunicado"],
        rec["estado_alerta"], rec["magnitud_costero"], rec["magnitud_pacifico"],
        rec["pronostico_lluvias"], rec["hidrologico"], rec["resumen"],
    ]
    for col_i, val in enumerate(vals, 1):
        c = ws.cell(row=row_i, column=col_i, value=val)
        c.font   = st["font_bold"] if col_i <= 5 else st["font_normal"]
        c.border = st["border"]
        c.fill   = fill
        c.alignment = st["align_center"] if col_i <= 5 else st["align_left"]
        if col_i == 1 and isinstance(val, date):
            c.number_format = "DD/MM/YYYY"
    ws.row_dimensions[row_i].height = 60

# ─── LECTURA DEL EXCEL EXISTENTE ─────────────────────────────────────────────

def leer_claves_existentes(xlsx_path):
    if not os.path.exists(xlsx_path):
        return set()
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        header_map = {str(c.value).strip(): c.column - 1
                      for c in ws[1] if c.value}
        col_a = header_map.get("Año")
        col_m = header_map.get("Mes Nro")
        if col_a is None or col_m is None:
            wb.close(); return set()
        claves = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                a = int(row[col_a]) if row[col_a] else None
                m = int(row[col_m]) if row[col_m] else None
                if a and m:
                    claves.add((a, m))
            except (TypeError, ValueError):
                pass
        wb.close()
        return claves
    except Exception as e:
        print(f"  ⚠  No se pudo leer el Excel existente: {e}")
        return set()

# ─── GUARDAR / ACTUALIZAR EXCEL ───────────────────────────────────────────────

def guardar_excel(nuevos_registros, advertencias_nuevas, xlsx_path):
    st = make_styles()

    # Leer registros existentes
    registros_existentes = []
    if os.path.exists(xlsx_path):
        try:
            wb_old = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws_old = wb_old.active
            header_map = {str(c.value).strip(): c.column - 1
                          for c in ws_old[1] if c.value}
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                def g(key, default=None):
                    idx = header_map.get(key)
                    return row[idx] if idx is not None and idx < len(row) else default
                try:
                    año     = int(g("Año") or 0)
                    mes_num = int(g("Mes Nro") or 0)
                    if not año or not mes_num:
                        continue
                    fecha_val = g("Fecha")
                    fecha = fecha_val.date() if hasattr(fecha_val, 'date') else \
                            fecha_val if isinstance(fecha_val, date) else \
                            date(año, mes_num, 1)
                    registros_existentes.append({
                        "fecha":             fecha,
                        "año":               año,
                        "mes_num":           mes_num,
                        "mes_str":           str(g("Mes", MESES_INV.get(mes_num, ""))),
                        "n_comunicado":      g("N° Comunicado", mes_num),
                        "estado_alerta":     g("Estado de Alerta", ""),
                        "magnitud_costero":  g("Magnitud El Niño Costero", ""),
                        "magnitud_pacifico": g("Magnitud Pacífico Central", ""),
                        "pronostico_lluvias":g("Pronóstico Lluvias", ""),
                        "hidrologico":       g("Pronóstico Hidrológico", ""),
                        "resumen":           g("Resumen Ejecutivo", ""),
                    })
                except (TypeError, ValueError):
                    pass
            wb_old.close()
        except Exception as e:
            print(f"  ⚠  Error leyendo Excel existente: {e}")

    # Leer advertencias existentes
    adv_existentes = []
    if os.path.exists(xlsx_path):
        try:
            wb_adv = load_workbook(xlsx_path, data_only=True, read_only=True)
            if "Advertencias" in wb_adv.sheetnames:
                for row in wb_adv["Advertencias"].iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        adv_existentes.append(str(row[0]))
            wb_adv.close()
        except Exception:
            pass

    # Combinar sin duplicados
    claves = {(r["año"], r["mes_num"]) for r in registros_existentes}
    for rec in nuevos_registros:
        if (rec["año"], rec["mes_num"]) not in claves:
            registros_existentes.append(rec)
            claves.add((rec["año"], rec["mes_num"]))

    adv_set = set(adv_existentes) | set(advertencias_nuevas)
    todas_adv = sorted(adv_set)

    todos = sorted(registros_existentes, key=lambda r: (r["año"], r["mes_num"]))

    # Escribir
    wb = Workbook()
    ws = wb.active
    ws.title = "Comunicados ENFEN"
    apply_header_row(ws, st)
    for row_i, rec in enumerate(todos, 2):
        apply_data_row(ws, row_i, rec, st)

    if todas_adv:
        ws2 = wb.create_sheet("Advertencias")
        ws2["A1"] = "Archivos / campos con extracción incompleta"
        ws2["A1"].font = Font(name="Arial", bold=True, size=10)
        ws2.column_dimensions["A"].width = 90
        for i, adv in enumerate(todas_adv, 2):
            ws2.cell(row=i, column=1, value=adv).font = Font(name="Arial", size=9)

    wb.save(xlsx_path)
    return len(todos), len(nuevos_registros)

# ─── PROCESAR UN PDF ──────────────────────────────────────────────────────────

# Campos que se consideran "No identificado" verdadero
# (excluimos los "No aplica" que son respuestas válidas para comunicados No Activo)
_NO_ID = "No identificado"

def procesar_pdf(fname, pdf_dir):
    año, mes_num, mes_str, n_comunicado = parse_filename(fname)
    if año is None:
        return None, f"[NOMBRE NO PARSEABLE] {fname}"

    texto = leer_pdf(os.path.join(pdf_dir, fname))
    if texto is None:
        return None, f"[ERROR LECTURA PDF] {fname}"

    estado_alerta      = extraer_estado_alerta(texto)
    magnitud_costero   = extraer_magnitud_costero(texto)
    magnitud_pacifico  = extraer_magnitud_pacifico(texto)
    pronostico_lluvias = extraer_pronostico_lluvias(texto)
    hidrologico        = extraer_hidrologico(texto)
    resumen            = extraer_resumen(texto)

    campos_ni = [
        c for c, v in [
            ("Estado de Alerta",        estado_alerta),
            ("Magnitud Costero",        magnitud_costero),
            ("Magnitud Pacífico",       magnitud_pacifico),
            ("Pronóstico Lluvias",      pronostico_lluvias),
            ("Pronóstico Hidrológico",  hidrologico),
        ] if v == _NO_ID
    ]
    adv = f"[CAMPOS NO IDENTIFICADOS] {fname} → {', '.join(campos_ni)}" if campos_ni else None

    rec = {
        "fecha":             date(año, mes_num, 1),
        "año":               año,
        "mes_num":           mes_num,
        "mes_str":           mes_str,
        "n_comunicado":      n_comunicado,
        "estado_alerta":     estado_alerta,
        "magnitud_costero":  magnitud_costero,
        "magnitud_pacifico": magnitud_pacifico,
        "pronostico_lluvias":pronostico_lluvias,
        "hidrologico":       hidrologico,
        "resumen":           resumen,
    }
    return rec, adv

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    try:
        todos_pdfs = sorted([
            f for f in os.listdir(PDF_DIR)
            if f.lower().endswith(".pdf") and "ENFEN_Comunicado" in f
        ])
    except FileNotFoundError:
        print(f"❌  Carpeta no encontrada: {PDF_DIR}")
        return

    SEP = "─" * 64
    print(f"\n{SEP}")
    print(f"  ENFEN EXTRACTOR  v3")
    print(SEP)
    print(f"  📂  Carpeta    : {PDF_DIR}")
    print(f"  📋  PDFs total : {len(todos_pdfs)}")

    claves_ya = leer_claves_existentes(XLSX_OUT)
    if claves_ya:
        print(f"  📊  Excel existente con {len(claves_ya)} registros → se omitirán esos PDFs")
    else:
        print(f"  📊  Excel no encontrado o vacío → se procesarán todos los PDFs")

    pendientes = []
    omitidos   = 0
    for fname in todos_pdfs:
        año, mes_num, _, _ = parse_filename(fname)
        if año and mes_num and (año, mes_num) in claves_ya:
            omitidos += 1
        else:
            pendientes.append(fname)

    print(f"  ⏭   Omitidos   : {omitidos}  (ya registrados en el Excel)")
    print(f"  🔎  Pendientes : {len(pendientes)}")
    print(f"{SEP}\n")

    if not pendientes:
        print("🎉  El Excel ya está al día. No hay PDFs nuevos que procesar.")
        return

    nuevos_registros = []
    advertencias     = []

    for fname in pendientes:
        rec, adv = procesar_pdf(fname, PDF_DIR)
        if rec is None:
            if adv:
                advertencias.append(adv)
                print(f"  ❌  {fname}\n     {adv}")
            continue

        estado  = "✅" if adv is None else "⚠ "
        ni_info = f"\n     ⚠  No identificados: {adv.split('→')[1].strip()}" if adv else ""
        print(f"  {estado}  {fname}")
        print(f"       Alerta       : {rec['estado_alerta']}")
        print(f"       Mag. Costero : {rec['magnitud_costero']}")
        print(f"       Mag. Pacífico: {rec['magnitud_pacifico']}{ni_info}")

        nuevos_registros.append(rec)
        if adv:
            advertencias.append(adv)

    total_filas, nuevas = guardar_excel(nuevos_registros, advertencias, XLSX_OUT)

    print(f"\n{SEP}")
    print(f"  RESULTADO FINAL")
    print(SEP)
    print(f"  📥  Nuevos registros añadidos : {nuevas}")
    print(f"  📊  Total filas en el Excel   : {total_filas}")
    print(f"  ⚠   Advertencias              : {len(advertencias)}")
    print(f"  💾  Guardado en:")
    print(f"      {XLSX_OUT}")
    if advertencias:
        print(f"\n  DETALLE ADVERTENCIAS (solo campos verdaderamente no identificados):")
        for a in advertencias:
            print(f"    {a}")
    print(f"{SEP}\n")

if __name__ == "__main__":
    main()
